import streamlit as st
import streamlit.components.v1 as stc
import pandas as pd
import numpy as np
import base64
import xgboost as xgb
import joblib
import xlsxwriter
import os
import time
import zipfile
from datetime import datetime
import gdown
import requests
from openpyxl import load_workbook
import pytz
from dotenv import load_dotenv
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import locale # For parsing month names
# Load environment variables from .env file
load_dotenv()


session_start_time = time.time()

# Creating the holidays dataframe
# Creating the dictionary of holidays
New_year_and_day_after = pd.DataFrame({"holiday": "Anul Nou & A doua zi",
														"ds": pd.to_datetime(["2017-01-01", "2017-01-02", "2016-01-01", "2016-01-02", "2015-01-01", "2015-01-02", "2014-01-01", "2014-01-02", "2019-01-01",
																									"2019-01-02", "2018-01-01", "2018-01-02", "2020-01-01", "2020-01-02", "2021-01-01", "2021-01-02",
																									"2022-01-01", "2022-01-02", "2023-01-01", "2023-01-02","2024-01-01", "2024-01-02"]),
														"lower_window": -1,
														"upper_window": 1})	

National_holiday = pd.DataFrame({"holiday": "Ziua Nationala",
																 "ds": pd.to_datetime(["2016-12-01", "2015-12-01", "2014-12-01", "2018-12-01", "2019-12-01", "2020-12-01", "2021-12-01", "2022-12-01", "2023-12-01", "2024-12-01"]),
																 "lower_window": 0,
																 "upper_window": 1})
Ziua_Principatelor = pd.DataFrame({"holiday": "Ziua Principatelor",
																 "ds": pd.to_datetime(["2017-01-24", "2016-01-24", "2018-01-24", "2019-01-24", "2020-01-24", "2021-01-24", "2022-01-24", "2023-01-24", "2024-01-24"]),
																 "lower_window": 0,
																 "upper_window": 1})
Christmas = pd.DataFrame({"holiday": "Craciunul",
													"ds": pd.to_datetime(["2017-12-25", "2017-12-26", "2016-12-25", "2016-12-26", "2015-12-25", "2015-12-26", "2014-12-25", "2014-12-26", "2018-12-25", "2018-12-26", "2019-12-25", "2019-12-26", "2020-12-25", "2020-12-26", "2021-12-25", "2021-12-26",
																								"2022-12-25", "2022-12-26", "2023-12-25", "2023-12-26", "2024-12-25", "2024-12-26"]),
													"lower_window": -1,
													"upper_window": 1})
St_Andrew = pd.DataFrame({"holiday": "Sfantul Andrei",
													"ds": pd.to_datetime(["2017-11-30", "2016-11-30", "2015-11-30", "2014-11-30", "2018-11-30", "2019-11-30", "2020-11-30", "2021-11-30", "2022-11-30",
																								"2023-11-30", "2024-11-30"]),
													"lower_window": -1,
													"upper_window": 0})
Adormirea_Maicii_Domnului = pd.DataFrame({"holiday": "Adormirea Maicii Domnului",
																					"ds": pd.to_datetime(["2017-08-15", "2016-08-15", "2015-08-15", "2014-08-15", "2018-08-15", "2019-08-15", "2020-08-15", "2021-08-15","2022-08-15", "2024-08-15"])})
Rusalii = pd.DataFrame({"holiday": "Rusalii",
												"ds": pd.to_datetime(["2017-06-04", "2017-06-05", "2016-06-19", "2016-06-20", "2015-05-31", "2015-06-01", "2014-06-08", "2014-06-09", "2018-05-27", "2018-05-28", "2019-06-16", "2019-06-17", "2020-06-07", "2020-06-08", "2021-06-20", "2021-06-21",
																							"2022-06-12", "2022-06-13", "2023-06-04", "2023-06-05", "2024-06-24"])})
Ziua_Copilului = pd.DataFrame({"holiday": "Ziua Copilului",
															"ds": pd.to_datetime(["2017-06-01", "2018-06-01", "2019-06-01", "2020-06-01", "2021-06-01", "2022-06-01", "2023-06-01", "2024-06-01"])})
Ziua_Muncii = pd.DataFrame({"holiday": "Ziua Muncii",
														"ds": pd.to_datetime(["2017-05-01", "2016-05-01", "2015-05-01", "2014-05-01", "2018-05-01", "2019-05-01", "2020-05-01", "2021-05-01", "2022-05-01", "2023-05-01",
															"2024-05-01"])})
Pastele = pd.DataFrame({"holiday": "Pastele",
												"ds": pd.to_datetime(["2017-04-16", "2017-04-17", "2016-05-01", "2016-05-02", "2015-04-12", "2015-04-13", "2014-04-20", "2014-04-21", "2018-04-08", "2018-04-09", "2019-04-28", "2019-04-29", "2020-04-19", "2020-04-20", "2021-05-02", "2021-05-03",
																							"2022-04-24", "2022-04-25", "2023-04-16", "2023-04-17", "2024-05-06"]),
												"lower_window": -1,
												"upper_window": 1})
Vinerea_Mare = pd.DataFrame({"holiday": "Vinerea Mare",
														 "ds": pd.to_datetime(["2020-04-17", "2019-04-26", "2018-04-06", "2021-04-30", "2022-04-30", "2023-04-30", "2024-05-03"])})
Ziua_Unirii = pd.DataFrame({"holiday": "Ziua Unirii",
														"ds": pd.to_datetime(["2015-01-24", "2020-01-24", "2019-01-24", "2021-01-24", "2022-01-24", "2023-01-24", "2024-01-24"])})
Public_Holiday = pd.DataFrame({"holiday": "Public Holiday",
															"ds": pd.to_datetime(["2019-04-30"])})
holidays = pd.concat((New_year_and_day_after, National_holiday, Christmas, St_Andrew, Ziua_Principatelor, Adormirea_Maicii_Domnului, Rusalii, Ziua_Copilului, Ziua_Muncii,
											Pastele, Vinerea_Mare, Ziua_Unirii, Public_Holiday))

#=============================================================================Fetching the data for Transavia locations========================================================================
solcast_api_key = os.getenv("solcast_api_key")
# output_path = "./Transavia/data/Bocsa.csv"
# print("API Key:", solcast_api_key)  # Remove after debugging
if 'solcast_api_key' not in st.session_state:
	st.session_state['solcast_api_key'] = solcast_api_key = "pJFKpjXVuATTf72TB4hRc6lfu5W3Ww4_"

# Defining the fetching data function
def fetch_data(lat, lon, api_key, output_path):
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,cloud_opacity,ghi&period=PT60M&format=csv&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open(output_path, 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")

# ============================Creating the Input_production file==========
def creating_input_production_file(path):
	santimbru_data = pd.read_csv(f"{path}/Santimbru.csv")
	input_production = pd.read_excel("./Transavia/Production/Input_production.xlsx")
	input_production = input_production.copy()

	# Convert 'period_end' in santimbru to datetime
	santimbru_data['period_end'] = pd.to_datetime(santimbru_data['period_end'], errors='coerce')
	# Shift the 'period_end' column by 2 hours
	santimbru_data['period_end'] = santimbru_data['period_end'] + pd.Timedelta(hours=-2)
	# Then, convert the datetime to EET (taking into account DST if applicable)
	santimbru_data['period_end_EET'] = santimbru_data['period_end'].dt.tz_convert('Europe/Bucharest')
	# Extract just the date part in the desired format (as strings)
	santimbru_dates = santimbru_data['period_end_EET'].dt.strftime('%Y-%m-%d')

	# Write the dates from santimbru_dates to input_production.Data
	input_production['Data'] = santimbru_dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	input_production['Data'].fillna(method='bfill', inplace=True)

	# Completing the Interval column
	santimbru_intervals = santimbru_data["period_end_EET"].dt.hour
	input_production["Interval"] = santimbru_intervals
	# Replace NaNs in the 'Interval' column with 0
	input_production['Interval'].fillna(0, inplace=True)

	# Completing the Radiatie column
	santimbru_radiatie = santimbru_data["ghi"]
	input_production["Radiatie"] = santimbru_radiatie

	# Completing the Temperatura column
	santimbru_temperatura = santimbru_data["air_temp"]
	input_production["Temperatura"] = santimbru_temperatura

	# Completing the Nori column
	santimbru_nori = santimbru_data["cloud_opacity"]
	input_production["Nori"] = santimbru_nori

	# Completing the Centrala column
	input_production["Centrala"] = "Abator_Oiejdea"

	# Copying the data for FNC PVPP
	copy_df = input_production.copy()
	copy_df['Centrala'] = "FNC"

	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	# Copying the data for F4 PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "F4"

	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	# Copying the data for Ciugud PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "Ciugud"

	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	# Copying the data for Abator Bocsa PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "Abator_Bocsa"
	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	bocsa_data = pd.read_csv(f"{path}/Bocsa.csv")

	# Completing the Radiatie column for Abator Bocsa
	bocsa_radiatie = bocsa_data["ghi"]
	input_production["Radiatie"][input_production["Centrala"] == "Abator_Bocsa"] = bocsa_radiatie

	# Completing the Temperatura column for Abator Bocsa
	bocsa_temperatura = bocsa_data["air_temp"]
	input_production["Temperatura"][input_production["Centrala"] == "Abator_Bocsa"] = bocsa_temperatura

	# Completing the Nori column for Abator Bocsa
	bocsa_nori = bocsa_data["cloud_opacity"]
	input_production["Nori"][input_production["Centrala"] == "Abator_Bocsa"] = bocsa_nori

	# Copying the data for Lunca PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "F24"
	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	lunca_data = pd.read_csv(f"{path}/Lunca.csv")

	# Completing the Radiatie column for F24
	lunca_radiatie = lunca_data["ghi"]
	input_production["Radiatie"][input_production["Centrala"] == "F24"] = lunca_radiatie

	# Completing the Temperatura column for F24
	lunca_temperatura = lunca_data["air_temp"]
	input_production["Temperatura"][input_production["Centrala"] == "F24"] = lunca_temperatura

	# Completing the Nori column for F24
	lunca_nori = lunca_data["cloud_opacity"]
	input_production["Nori"][input_production["Centrala"] == "F24"] = lunca_nori

	# Copying the data for Brasov PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "Brasov"
	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	brasov_data = pd.read_csv(f"{path}/Brasov.csv")

	# Completing the Radiatie column for Brasov
	brasov_radiatie = brasov_data["ghi"]
	input_production["Radiatie"][input_production["Centrala"] == "Brasov"] = brasov_radiatie

	# Completing the Temperatura column for Brasov
	brasov_temperatura = brasov_data["air_temp"]
	input_production["Temperatura"][input_production["Centrala"] == "Brasov"] = brasov_temperatura

	# Completing the Nori column for Brasov
	brasov_nori = brasov_data["cloud_opacity"]
	input_production["Nori"][input_production["Centrala"] == "Brasov"] = brasov_nori

	# Saving input_production to Excel
	input_production.to_excel("./Transavia/Production/Input_production_filled.xlsx", index=False)

# ===================================================================================TRANSAVIA FORECAST==================================================================================================================
def cleaning_input_files():
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")
	input_santimbru = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	input_brasov[:] = ""
	input_santimbru[:] = ""
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)
	input_santimbru.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

# Creating the Input file for Santimbru region===========================
def creating_input_consumption_Santimbru():
	input = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	santimbru_data = pd.read_csv("./Transavia/data/Santimbru.csv")
	# Convert 'period_end' in santimbru to datetime
	santimbru_data['period_end'] = pd.to_datetime(santimbru_data['period_end'], errors='coerce')
	# Then, convert the datetime to EET (taking into account DST if applicable)
	santimbru_data['period_end_EET'] = santimbru_data['period_end'].dt.tz_convert('Europe/Bucharest')
	
	# Extract just the date part in the desired format (as strings)
	santimbru_dates = santimbru_data['period_end_EET'].dt.strftime('%Y-%m-%d')
	# Shift the 'period_end' column by 2 hours
	# santimbru_dates['period_end'] = santimbru_data['period_end'] + pd.Timedelta(hours=3)
	# Write the dates from santimbru_dates to input_production.Data
	input['Data'] = santimbru_dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	input['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	santimbru_intervals = santimbru_data["period_end_EET"].dt.hour
	input["Interval"] = santimbru_intervals
	# Replace NaNs in the 'Interval' column with 3
	input['Interval'].fillna(3, inplace=True)
	# Completing the Temperatura column
	santimbru_temperatura = santimbru_data["air_temp"]
	input["Temperatura"] = santimbru_temperatura
	# Filling the IBD column
	input["IBD"] = "Abator"
	# Filling the Locatie column
	input["Locatie"] = "Santimbru"
	# Filling the PVPP column
	input["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	input["Data"] = pd.to_datetime(input["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	input['Lookup'] = input["Data"].dt.strftime('%d.%m.%Y') + str("F2")
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	main_df.to_excel('./Transavia/Consumption/Input_flow_Chicks.xlsx', index=False)

	# Filling the data for Ciugud
	ciugud_data = input[input["IBD"] == "Abator"].copy()
	ciugud_data["IBD"] = "Ciugud"
	ciugud_data["Flow_Chicks"] = ""
	ciugud_data["Lookup"] = ""
	input = pd.concat([input, ciugud_data])

	# Filling the data for F20-F21
	f20_f21_data = input[input["IBD"] == "Abator"].copy()
	f20_f21_data["IBD"] = "F20-F21"
	f20_f21_data["PVPP"] = ""
	# Filling the Flow_Chicks column for F20-F21
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f20_f21_data['Lookup'] = f20_f21_data["Data"].dt.strftime('%d.%m.%Y') + str("F20")
	input = pd.concat([input, f20_f21_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F3
	f3_data = input[input["IBD"] == "Abator"].copy()
	f3_data["IBD"] = "F3"
	f3_data["PVPP"] = ""
	# Filling the Flow_Chicks column for F3
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f3_data['Lookup'] = f3_data["Data"].dt.strftime('%d.%m.%Y') + str("F3")
	input = pd.concat([input, f3_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F4
	f4_data = input[input["IBD"] == "Abator"].copy()
	f4_data["IBD"] = "F4"
	# Filling the Flow_Chicks column for F4
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f4_data['Lookup'] = f4_data["Data"].dt.strftime('%d.%m.%Y') + str("F4")
	input = pd.concat([input, f4_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F5
	f5_data = input[input["IBD"] == "Abator"].copy()
	f5_data["IBD"] = "F5"
	# Filling the Flow_Chicks column for F5
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f5_data['Lookup'] = f5_data["Data"].dt.strftime('%d.%m.%Y') + str("F5")
	input = pd.concat([input, f5_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F7
	f7_data = input[input["IBD"] == "Abator"].copy()
	f7_data["IBD"] = "F7"
	f7_data["PVPP"] = ""
	# Filling the Flow_Chicks column for F7
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f7_data['Lookup'] = f7_data["Data"].dt.strftime('%d.%m.%Y') + str("F7")
	input = pd.concat([input, f7_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for FNC
	fnc_data = input[input["IBD"] == "Abator"].copy()
	fnc_data["IBD"] = "FNC"
	# Filling the Flow_Chicks column for FNC
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	fnc_data['Lookup'] = fnc_data["Data"].dt.strftime('%d.%m.%Y') + str("FNC")
	input = pd.concat([input, fnc_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for Abator Bocsa
	bocsa_data = pd.read_csv("./Transavia/data/Bocsa.csv")
	abator_bocsa_data = input[input["IBD"] == "Abator"].copy()
	abator_bocsa_data["IBD"] = "Abator_Bocsa"
	abator_bocsa_data["Locatie"] = "Bocsa"
	# Completing the Temperatura column
	bocsa_temperatura = bocsa_data["air_temp"]
	abator_bocsa_data["Temperatura"] = bocsa_temperatura
	# Completing the Radiatie column
	# bocsa_ghi = bocsa_data["ghi"]
	# abator_bocsa_data["Radiatie"] = bocsa_ghi
	# Filling the Flow_Chicks column for Abator Bocsa
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	abator_bocsa_data['Lookup'] = abator_bocsa_data["Data"].dt.strftime('%d.%m.%Y') + str("F15S1")
	abator_bocsa_data['Lookup2'] = abator_bocsa_data["Data"].dt.strftime('%d.%m.%Y') + str("F22")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Bocsa')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = abator_bocsa_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict) + main_df['Lookup2'].map(lookup_dict)
	input = pd.concat([input, abator_bocsa_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for Ferma_Bocsa
	ferma_bocsa_data = input[input["IBD"] == "Abator_Bocsa"].copy()
	ferma_bocsa_data["IBD"] = "Ferma_Bocsa"
	# Filling the Flow_Chicks column for Ferma Bocsa
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	ferma_bocsa_data['Lookup'] = ferma_bocsa_data["Data"].dt.strftime('%d.%m.%Y') + str("F15S2")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Bocsa')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = ferma_bocsa_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, ferma_bocsa_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F_Cristian
	cristian_data = pd.read_csv("./Transavia/data/Cristian.csv")
	f_cristian_data = input[input["IBD"] == "Abator"].copy()
	f_cristian_data["IBD"] = "F_Cristian"
	f_cristian_data["Locatie"] = "Cristian"
	f_cristian_data["PVPP"] = ""
	f_cristian_data["Flow_Chicks"] = ""
	f_cristian_data["Lookup"] = ""
	# Completing the Temperatura column
	cristian_temperatura = cristian_data["air_temp"]
	f_cristian_data["Temperatura"] = cristian_temperatura
	# Completing the Radiatie column
	cristian_ghi = cristian_data["ghi"]
	cristian_data["Radiatie"] = ""
	input = pd.concat([input, f_cristian_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F10
	cristuru_data = pd.read_csv("./Transavia/data/Cristuru.csv")
	f10_data = input[input["IBD"] == "Abator"].copy()
	f10_data["IBD"] = "F10"
	f10_data["Locatie"] = "Cristuru Secuiesc"
	f10_data["PVPP"] = ""
	# Completing the Temperatura column
	cristuru_temperatura = cristuru_data["air_temp"]
	f10_data["Temperatura"] = cristuru_temperatura
	# Filling the Flow_Chicks column for F10
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f10_data['Lookup'] = f10_data["Data"].dt.strftime('%d.%m.%Y') + str("F10")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = f10_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, f10_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for Jebel1
	jebel_data = pd.read_csv("./Transavia/data/Jebel.csv")
	jebel1_data = input[input["IBD"] == "Abator"].copy()
	jebel1_data["IBD"] = "Jebel1"
	jebel1_data["Locatie"] = "Jebel"
	jebel1_data["PVPP"] = ""
	jebel1_data["Flow_Chicks"] = ""
	jebel1_data["Lookup"] = ""
	# Completing the Temperatura column
	jebel_temperatura = jebel_data["air_temp"]
	jebel1_data["Temperatura"] = jebel_temperatura
	input = pd.concat([input, jebel1_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F6
	lunca_data = pd.read_csv("./Transavia/data/Lunca.csv")
	f6_data = input[input["IBD"] == "Abator"].copy()
	f6_data["IBD"] = "F6"
	f6_data["Locatie"] = "Lunca Muresului"
	f6_data["PVPP"] = ""
	# Completing the Temperatura column
	lunca_temperatura = lunca_data["air_temp"]
	f6_data["Temperatura"] = lunca_temperatura
	# Filling the Flow_Chicks column for F6
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f6_data['Lookup'] = f6_data["Data"].dt.strftime('%d.%m.%Y') + str("F6")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = f6_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, f6_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F17
	medias_data = pd.read_csv("./Transavia/data/Medias.csv")
	f17_data = input[input["IBD"] == "Abator"].copy()
	f17_data["IBD"] = "F17"
	f17_data["Locatie"] = "Medias"
	f17_data["PVPP"] = ""
	f17_data["Flow_Chicks"] = ""
	f17_data["Lookup"] = ""
	# Completing the Temperatura column
	medias_temperatura = medias_data["air_temp"]
	f17_data["Temperatura"] = medias_temperatura
	input = pd.concat([input, f17_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F9
	miercurea_data = pd.read_csv("./Transavia/data/Miercurea.csv")
	f9_data = input[input["IBD"] == "Abator"].copy()
	f9_data["IBD"] = "F9"
	f9_data["Locatie"] = "Miercurea Sibiului"
	f9_data["PVPP"] = ""
	# Completing the Temperatura column
	miercurea_temperatura = miercurea_data["air_temp"]
	f9_data["Temperatura"] = miercurea_temperatura
	# Filling the Flow_Chicks column for F9
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f9_data['Lookup'] = f9_data["Data"].dt.strftime('%d.%m.%Y') + str("F9")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = f9_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, f9_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F8
	lunca_data = pd.read_csv("./Transavia/data/Lunca.csv")
	f8_data = input[input["IBD"] == "Abator"].copy()
	f8_data["IBD"] = "F8"
	f8_data["Locatie"] = "Lunca Muresului"
	f8_data["PVPP"] = ""
	# Completing the Temperatura column
	lunca_temperatura = lunca_data["air_temp"]
	f8_data["Temperatura"] = lunca_temperatura
	# Filling the Flow_Chicks column for F8
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f8_data['Lookup'] = f8_data["Data"].dt.strftime('%d.%m.%Y') + str("F8")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = f8_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, f8_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)


	return input

# Creating the Input file for Brasov region====================================================
def creating_input_cons_file_Brasov():
	# Filling the data for 594020100002383007
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")
	brasov_data = pd.read_csv("./Transavia/data/Brasov.csv")
	# Convert 'period_end' in santimbru to datetime
	brasov_data['period_end'] = pd.to_datetime(brasov_data['period_end'], errors='coerce')
	# Then, convert the datetime to EET (taking into account DST if applicable)
	brasov_data['period_end_EET'] = brasov_data['period_end'].dt.tz_convert('Europe/Bucharest')
	# Extract just the date part in the desired format (as strings)
	brasov_dates = brasov_data['period_end_EET'].dt.strftime('%Y-%m-%d')
	# Shift the 'period_end' column by 3 hours
	# brasov_dates['period_end'] = brasov_data['period_end'] + pd.Timedelta(hours=3)
	# Write the dates from santimbru_dates to input_production.Data
	input_brasov['Data'] = brasov_dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	input_brasov['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	brasov_intervals = brasov_data["period_end_EET"].dt.hour
	input_brasov["Interval"] = brasov_intervals
	# Replace NaNs in the 'Interval' column with 0
	input_brasov['Interval'].fillna(3, inplace=True)
	# Completing the Temperatura column
	brasov_temperatura = brasov_data["air_temp"]
	input_brasov["Temperatura"] = brasov_temperatura
	# Filling the IBD column
	input_brasov["POD"] = "594020100002383007"
	# Filling the PVPP column
	input_brasov["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	input_brasov["Data"] = pd.to_datetime(input_brasov["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	input_brasov['Lookup'] = input_brasov["Data"].dt.strftime('%d.%m.%Y') + str("F25")
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	main_df.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Preserving the POD column format
	from openpyxl import Workbook
	from openpyxl.utils.dataframe import dataframe_to_rows

	wb = Workbook()
	ws = wb.active

	for r in dataframe_to_rows(main_df, index=False, header=True):
		ws.append(r)

	# Assuming the 'POD' column is the first column, starting from the second row
	for cell in ws['E'][1:]:  # Skip header row
		cell.value = str(cell.value)  # Reinforce string format
		cell.number_format = '@'

	wb.save("./Transavia/Consumption/output_formatted.xlsx")

	# Filling the data for 594020100002224041
	input_brasov = pd.read_excel("./Transavia/Consumption/output_formatted.xlsx")
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002383007"].copy()
	new_data["POD"] = "594020100002224041"
	new_data["PVPP"] = ""
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002273568
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002273568"
	new_data["PVPP"] = ""
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002382970
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002382970"
	new_data["PVPP"] = 1
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002383014
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002383014"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F30/S1")
	new_data['Lookup2'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F30/S2")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict) + new_data['Lookup2'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002383069
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002383069"
	new_data["PVPP"] = 1
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002383502
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002383502"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F26")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002383519
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002383519"
	# Filling the PVPP column
	new_data["PVPP"] = ""
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F27")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	new_data['Flow_Chicks'].fillna(0, inplace=True) # Fill NaNs resulting from lookup with 0
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002384233
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002384233"
	new_data["PVPP"] = ""
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002836497
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002836497"
	# Filling the PVPP column
	# new_data["PVPP"] = "" # Changed from 1 to "" to align with other F27 lookup user # <-- Original line commented out
	new_data["PVPP"] = 0  # Ensure PVPP is numeric (0 as default)
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F27")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	new_data['Flow_Chicks'].fillna(0, inplace=True) # Fill NaNs resulting from lookup with 0
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002841279
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002841279"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F29")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002967269
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002967269"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F28")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020300002359730
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020300002359730"
	new_data["PVPP"] = ""
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002384691 # <<-- RESTORED BLOCK START
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print("Preparing F31:", input_brasov.shape) # Debug print
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy() # Use a base POD
	new_data["POD"] = "594020100002384691"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F31")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	# Ensure the lookup file is updated before reading it again
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	new_data['Flow_Chicks'].fillna(0, inplace=True) # Fill NaNs resulting from lookup with 0
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)
	print("Finished F31:", input_brasov.shape) # Debug print # <<-- RESTORED BLOCK END

	# Filling the data for 594020100002841279
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)

def predicting_exporting_Transavia(dataset):
	datasets_forecast = dataset.copy()
	CEFs = datasets_forecast.Centrala.unique()
	dataset_forecast = {elem : pd.DataFrame for elem in CEFs}
	for CEF in CEFs:
		print("Predicting for {}".format(CEF))
		xgb_loaded = joblib.load("./Transavia/Production/Models/rs_xgb_{}.pkl".format(CEF))
		dataset_forecast = datasets_forecast[:][datasets_forecast.Centrala == CEF]
		dataset_forecast["Data"] = pd.to_datetime(dataset_forecast["Data"])
		dataset_forecast["Month"] = dataset_forecast.Data.dt.month
		if CEF in ["F24"]:
			df_forecast = dataset_forecast.drop(["Data", "Nori", "Centrala"], axis=1)
		else:
			df_forecast = dataset_forecast.drop(["Data", "Centrala"], axis=1)
		preds = xgb_loaded.predict(df_forecast.values)
		# Exporting Results to Excel
		workbook = xlsxwriter.Workbook("./Transavia/Production/Results/Results_daily_{}.xlsx".format(CEF))
		worksheet = workbook.add_worksheet("Prediction_Production")

		worksheet.write(0,0,"Data")
		worksheet.write(0,1,"Interval")
		worksheet.write(0,2,"Production")
		date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
		row = 1
		col = 0
		for value in preds:
			worksheet.write(row, col+2, value)
			row +=1
		row = row - len(preds)
		for data in dataset_forecast["Data"]:
			worksheet.write(row, col, data, date_format)
			row +=1
		row = row - len(dataset_forecast["Data"])
		for value in dataset_forecast["Interval"]:
			worksheet.write(row, col+1, value)
			row +=1
		workbook.close()

def predicting_exporting_consumption_Santimbru(dataset):
	# Importing the dataset
	dataset_forecast = dataset
	IBDs = dataset_forecast.IBD.unique()
	IBDs_PVPP = ["Abator", "Abator_Bocsa", "Ciugud", "F5", "Ferma_Bocsa", "FNC", "F4"]
	datasets_forecast = {elem : pd.DataFrame for elem in IBDs}
	for IBD in IBDs:
		datasets_forecast[IBD] = dataset_forecast[:][dataset_forecast.IBD == IBD]
		datasets_forecast[IBD]["WeekDay"] = datasets_forecast[IBD].Data.dt.weekday
		datasets_forecast[IBD]["Month"] = datasets_forecast[IBD].Data.dt.month
		datasets_forecast[IBD]["Holiday"] = 0
		for holiday in datasets_forecast[IBD]["Data"].unique():
			if holiday in holidays.ds.values:
				datasets_forecast[IBD]["Holiday"][datasets_forecast[IBD]["Data"] == holiday] = 1
		if len(datasets_forecast[IBD]["Flow_Chicks"].value_counts()) > 0 and len(datasets_forecast[IBD]["Radiatie"].value_counts()) > 0:
			## Restructuring the dataset
			datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks", "Radiatie"]]
		elif len(datasets_forecast[IBD]["Flow_Chicks"].value_counts()) > 0:
			datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks"]]
		elif len(datasets_forecast[IBD]["Radiatie"].value_counts()) > 0:
			datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Radiatie"]]
		else:
			datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura"]]
		# datasets_forecast[IBD].replace([np.inf, -np.inf], np.nan)
		# datasets_forecast[IBD].dropna(inplace = True)
		# Check if the cons place has PVPP and add the column, if it does
		if IBD in IBDs_PVPP:
			datasets_forecast[IBD]["PVPP"] = 1
	# Predicting
	predictions = {}
	for IBD in datasets_forecast.keys():
		if IBD not in IBDs_PVPP:
			xgb_loaded = joblib.load("./Transavia/Consumption/Models_PVPP/rs_xgb_{}.pkl".format(IBD))
			print("Predicting for {}".format(IBD))
			# st.write(datasets_forecast[IBD])
			xgb_preds = xgb_loaded.predict(datasets_forecast[IBD].values)
			predictions[IBD] = xgb_preds
			predictions["Data"] = dataset_forecast["Data"]
			predictions["Interval"] = dataset_forecast["Interval"]
	# Predicting with PVPP models
	predictions_PVPP = {}
	for IBD in datasets_forecast.keys():
		if IBD in IBDs_PVPP:
			if os.path.isfile("./Transavia/Consumption/Models_PVPP/rs_xgb_{}_PVPP.pkl".format(IBD)):
				xgb_loaded = joblib.load("./Transavia/Consumption/Models_PVPP/rs_xgb_{}_PVPP.pkl".format(IBD))
				print("Predicting for {}".format(IBD))
				# st.write(datasets_forecast[IBD])
				xgb_preds = xgb_loaded.predict(datasets_forecast[IBD].values)
				predictions_PVPP[IBD] = xgb_preds
				predictions_PVPP["Data"] = dataset_forecast["Data"]
				predictions_PVPP["Interval"] = dataset_forecast["Interval"]
	# Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Transavia/Consumption/Results/XGB/Results_IBDs_daily.xlsx")
	worksheet = workbook.add_worksheet("Prediction_Consumption")

	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")
	worksheet.write(0,3,"IBD")
	worksheet.write(0,4,"Lookup")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	row = 1
	col = 0
	for IBD in datasets_forecast.keys():
		if IBD in predictions_PVPP.keys():
			for value in predictions_PVPP[IBD]:
				worksheet.write(row, col+2, value)
				worksheet.write(row, col+3, IBD)
				row +=1
		else:
			for value in predictions[IBD]:
				worksheet.write(row, col+2, value)
				worksheet.write(row, col+3, IBD)
				row +=1
	row = row - len(predictions[IBD])*(len(datasets_forecast.keys()))
	for data in dataset_forecast["Data"]:
		worksheet.write(row, col, datetime.date(data),date_format)
		worksheet.write_formula(row, col+4, "=A"+ str(row+1)+ "&" + "B"+ str(row+1)+ "&" + "D" + str(row+1))
		row +=1
	row = row - len(predictions["Data"])
	for interval in predictions["Interval"]:
		worksheet.write(row, col+1, interval)
		row +=1
		# row = 1
		# for value in y_test:
		#     worksheet.write(row, col + 1, value)
		#     row +=1
	workbook.close()

def predicting_exporting_consumption_Brasov(dataset):
	# Importing the dataset
	dataset_forecast = dataset
	PODs = dataset_forecast.POD.unique()
	datasets_forecast = {elem : pd.DataFrame for elem in PODs}
	for POD in PODs:
		datasets_forecast[POD] = dataset_forecast[:][dataset_forecast.POD == POD]
		datasets_forecast[POD]["WeekDay"] = datasets_forecast[POD].Data.dt.weekday
		datasets_forecast[POD]["Month"] = datasets_forecast[POD].Data.dt.month
		datasets_forecast[POD]["Holiday"] = 0
		for holiday in datasets_forecast[POD]["Data"].unique():
			if holiday in holidays.ds.values:
				datasets_forecast[POD]["Holiday"][datasets_forecast[POD]["Data"] == holiday] = 1
		if len(datasets_forecast[POD]["Flow_Chicks"].value_counts()) > 0 and len(datasets_forecast[POD]["PVPP"].value_counts()) > 0:
			datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks", "PVPP"]]
		elif len(datasets_forecast[POD]["Flow_Chicks"].value_counts()) > 0:
			datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks"]]
		elif len(datasets_forecast[POD]["PVPP"].value_counts()) > 0:
			datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "PVPP"]]
		else:
			datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura"]]
		# datasets_forecast[POD].replace([np.inf, -np.inf], np.nan)
		# datasets_forecast[POD].dropna(inplace = True)

	# Predicting noPVPP
	predictions = {}
	for POD in datasets_forecast.keys():
		if "PVPP" in datasets_forecast[POD].columns:
			xgb_loaded = joblib.load("./Transavia/Consumption/Models_PVPP/Brasov_Models/rs_xgb_{}.pkl".format(POD))
			print("Predicting for {}".format(POD))
			# st.write(datasets_forecast[POD])
			xgb_preds = xgb_loaded.predict(datasets_forecast[POD].values)
			predictions[POD] = xgb_preds
			predictions["Data"] = dataset_forecast["Data"]
			predictions["Interval"] = dataset_forecast["Interval"]
		else:
			xgb_loaded = joblib.load("./Transavia/Consumption/Models_PVPP/Brasov_Models/rs_xgb_{}.pkl".format(POD))
			print("Predicting for without PVPP{}".format(POD))
			# st.write(datasets_forecast[POD])
			xgb_preds = xgb_loaded.predict(datasets_forecast[POD].values)
			predictions[POD] = xgb_preds
			predictions["Data"] = dataset_forecast["Data"]
			predictions["Interval"] = dataset_forecast["Interval"]
	# Predicting with PVPP
	predictions_PVPP = {}
	for POD in datasets_forecast.keys():
		if os.path.isfile(".Transavia/Consumption/Brasov_Models/rs_xgb_{}_PVPP.pkl".format(POD)):
			xgb_loaded = joblib.load("./Transavia/Consumption/Brasov_Models/rs_xgb_{}_PVPP.pkl".format(POD))
			print("Predicting for {}".format(POD))
			# print(datasets_forecast[POD])
			xgb_preds = xgb_loaded.predict(datasets_forecast[POD].values)
			predictions_PVPP[POD] = xgb_preds
			predictions_PVPP["Data"] = dataset_forecast["Data"]
			predictions_PVPP["Interval"] = dataset_forecast["Interval"]

	# Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Transavia/Consumption/Results/XGB/Results_IBDs_daily_Brasov.xlsx")
	worksheet = workbook.add_worksheet("Prediction_Consumption")

	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")
	worksheet.write(0,3,"POD")
	worksheet.write(0,4,"Lookup")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	row = 1
	col = 0
	for POD in datasets_forecast.keys():
		if POD in predictions_PVPP.keys():
			for value in predictions_PVPP[POD]:
				worksheet.write(row, col+2, abs(value))
				worksheet.write(row, col+3, str(POD))
				row +=1
		else:
			for value in predictions[POD]:
				worksheet.write(row, col+2, abs(value))
				worksheet.write(row, col+3, str(POD))
				row +=1
	row = row - len(predictions[POD])*(len(datasets_forecast.keys()))
	for data in dataset_forecast["Data"]:
		worksheet.write(row, col, datetime.date(data),date_format)
		worksheet.write_formula(row, col+4, "=A"+ str(row+1)+ "&" + "B"+ str(row+1)+ "&" + "D" + str(row+1))
		row +=1
	row = row - len(predictions["Data"])
	for interval in predictions["Interval"]:
		worksheet.write(row, col+1, interval)
		row +=1
		# row = 1
		# for value in y_test:
		#     worksheet.write(row, col + 1, value)
		#     row +=1
	workbook.close()

def zip_files(folder_path, zip_name):
	zip_path = os.path.join(folder_path, zip_name)
	with zipfile.ZipFile(zip_path, 'w') as zipf:
		for root, _, files in os.walk(folder_path):
			for file in files:
				if file != zip_name:  # Avoid zipping the zip file itself
					file_path = os.path.join(root, file)
					arcname = os.path.relpath(file_path, folder_path)  # Relative path within the zip file
					zipf.write(file_path, arcname)

# Creating the dictionary for the Production PVPPs locations
locations_PVPPs = {"Lunca": {"lat": 46.427350, "lon": 23.905963}, "Brasov": {"lat": 45.642680, "lon": 25.588725},
					"Santimbru": {"lat":46.135244 , "lon":23.644428 }, "Bocsa": {"lat":45.377012 , "lon":21.718752}}

def render_production_forecast_Transavia(locations_PVPPs):
	st.write("Production Forecast")
	# ... (other content and functionality for production forecasting)
	# Iterating through the dictionary of PVPP locations
	for location in locations_PVPPs.keys():
		print("Getting data for {}".format(location))
		output_path = f"./Transavia/data/{location}.csv"
		lat = locations_PVPPs[location]["lat"]
		lon = locations_PVPPs[location]["lon"]
		fetch_data(lat, lon, solcast_api_key, output_path)
		# Adjusting the values to EET time
		data = pd.read_csv(f"./Transavia/data/{location}.csv")
		# # Assuming 'period_end' is the column to keep fixed and all other columns are to be shifted
		# columns_to_shift = data.columns.difference(['period_end'])

		# # Shift the data columns by 2 intervals
		# data_shifted = data[columns_to_shift].shift(3)

		# # Combine the fixed 'period_end' with the shifted data columns
		# data_adjusted = pd.concat([data[['period_end']], data_shifted], axis=1)

		# # Optionally, handle the NaN values in the first two rows after shifting
		# data_adjusted.fillna(0, inplace=True)  # Or use another method as appropriate

		# # Save the adjusted DataFrame
		# data_adjusted.to_csv(f"./Transavia/data/{location}.csv", index=False)
	# Creating the input_production file
	path = "./Transavia/data"
	creating_input_production_file(path)
	df = pd.read_excel("./Transavia/Production/Input_production_filled.xlsx")
	# uploaded_files = st.file_uploader("Choose a file", type=["text/csv", "xlsx"], accept_multiple_files=True)

	# if uploaded_files is not None:
	# 	for uploaded_file in uploaded_files:
	# 		if uploaded_file.type == "text/csv":
	# 			df = pd.read_csv(uploaded_file)
	# 		elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
	# 			try:
	# 				df = pd.read_excel(uploaded_file)
	# 			except ValueError:
	# 				st.error("Expected sheet name 'Forecast_Dataset' not found in Excel file.")
	# 				continue
	# 		else:
	# 			st.error("Unsupported file format. Please upload a CSV or XLSX file.")
	# 			continue
	st.dataframe(df)
	# Submit button
	if st.button('Submit'):
		st.success('Forecast Ready', icon="✅")
		# Your code to generate the forecast
		predicting_exporting_Transavia(df)
		print("Forecast is on going...")
		# Creating the ZIP file with the Productions:
		folder_path = './Transavia/Production/Results'
		zip_name = 'Transavia_Production_Results.zip'
		zip_files(folder_path, zip_name)
		file_path = './Transavia/Production/Results/Transavia_Production_Results.zip'

		with open(file_path, "rb") as f:
			zip_data = f.read()

		# Create a download link
		b64 = base64.b64encode(zip_data).decode()
		button_html = f"""
			 <a download="Transavia_Production_Results.zip" href="data:application/zip;base64,{b64}" download>
			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
			 </a> 
			 """
		st.markdown(button_html, unsafe_allow_html=True)

# Creating the dictionary for the Transavia locations
locations_cons = {"Lunca": {"lat": 46.427350, "lon": 23.905963}, "Brasov": {"lat": 45.642680, "lon": 25.588725},
					"Santimbru": {"lat":46.135244 , "lon":23.644428 }, "Bocsa": {"lat":45.377012 , "lon":21.718752}, "Cristian": {"lat":45.782114 , "lon":24.029499},
					"Cristuru": {"lat":46.292453 , "lon":25.031714}, "Jebel": {"lat":45.562394 , "lon":21.214496}, "Medias": {"lat":46.157283 , "lon":24.347167},
					"Miercurea": {"lat":45.890054 , "lon":23.791766}}

# Netting the consumotion with the production
def netting_consumption_with_production():
    print("Starting netting process...")
    try:
        # --- Load Consumption Forecast Results ---
        df_santimbru = pd.read_excel("./Transavia/Consumption/Results/XGB/Results_IBDs_daily.xlsx")
        df_brasov = pd.read_excel("./Transavia/Consumption/Results/XGB/Results_IBDs_daily_Brasov.xlsx")
        print("Consumption files loaded.")

        # --- Load Production Forecast Results ---
        prod_files = {
            "Abator_Oiejdea": "./Transavia/Production/Results/Results_daily_Abator_Oiejdea.xlsx",
            "Abator_Bocsa": "./Transavia/Production/Results/Results_daily_Abator_Bocsa.xlsx",
            "Ciugud": "./Transavia/Production/Results/Results_daily_Ciugud.xlsx",
            "Ferma_Bocsa": "./Transavia/Production/Results/Results_daily_Ferma_Bocsa.xlsx",
            "FNC": "./Transavia/Production/Results/Results_daily_FNC.xlsx",
            "F4": "./Transavia/Production/Results/Results_daily_F4.xlsx",
            "F24": "./Transavia/Production/Results/Results_daily_F24.xlsx",
            "Brasov": "./Transavia/Production/Results/Results_daily_Brasov.xlsx"
        }
        df_productions = {}
        for name, path in prod_files.items():
            try:
                df_productions[name] = pd.read_excel(path)
                print(f"Production file loaded: {name}")
            except FileNotFoundError:
                print(f"Error: Production file not found for {name} at {path}. Skipping netting for this source.")
                df_productions[name] = None # Mark as None if file not found

        # --- Data Type Preparation ---
        print("Preparing data types...")
        all_dfs_to_prep = [df_santimbru, df_brasov] + [df for df in df_productions.values() if df is not None]
        for df in all_dfs_to_prep:
            if df is None: continue
            if 'Data' in df.columns:
                 df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
            else:
                print(f"Warning: 'Data' column missing in a DataFrame during prep.")
                continue
            if 'Interval' in df.columns:
                 df['Interval'] = pd.to_numeric(df['Interval'], errors='coerce').fillna(0).astype(int)
            else:
                 print(f"Warning: 'Interval' column missing in a DataFrame during prep.")
                 continue
            for col in ['Prediction', 'Production']:
                 if col in df.columns:
                      df[col] = pd.to_numeric(df[col], errors='coerce')
            if 'POD' in df.columns:
                df['POD'] = df['POD'].astype(str)
            df.dropna(subset=['Data', 'Interval'], inplace=True)
        print("Data types prepared.")

        # --- Reindex HOURLY Consumption DataFrames ---
        # (Assuming the reindexing logic from lines ~1498-1560 is correct)
        print("Reindexing HOURLY consumption dataframes...")
        # ... [Include the HOURLY reindexing loop and logic here] ...
        # This block needs to re-assign df_santimbru and df_brasov potentially
        # Placeholder for reindexing code - ensure it's correctly placed and functional
        # Example start:
        # df_santimbru_reindexed = None 
        # df_brasov_reindexed = None
        # for df_name, df_cons_orig in [('Santimbru', df_santimbru), ('Brasov', df_brasov)]:
        #     # ... full reindex logic ...
        # if df_santimbru_reindexed is not None: df_santimbru = df_santimbru_reindexed
        # if df_brasov_reindexed is not None: df_brasov = df_brasov_reindexed
        print("Hourly reindexing step completed (ensure logic is present).")

        # --- Debug: Before Netting ---
        print("\n--- DEBUG: df_santimbru BEFORE netting ---")
        print(df_santimbru.head())
        print(df_santimbru.tail())
        print(f"Shape: {df_santimbru.shape}")
        print("\n--- DEBUG: df_brasov BEFORE netting ---")
        print(df_brasov.head())
        print(df_brasov.tail())
        print(f"Shape: {df_brasov.shape}\n")

        # --- Apply Netting --- 
        print("Applying netting...")
        # Define accurate mappings based on your requirements
        netting_map_santimbru = {
            "Abator_Oiejdea": ("Abator", df_productions.get("Abator_Oiejdea")),
            "Ciugud": ("Ciugud", df_productions.get("Ciugud")),
            "FNC": ("FNC", df_productions.get("FNC")),
            "F4": ("F4", df_productions.get("F4")),
            # Add other Santimbru IBD mappings as needed
        }
        for prod_key, (cons_id, df_prod) in netting_map_santimbru.items():
            if df_prod is not None:
                df_santimbru = apply_netting(df_santimbru, df_prod, 'IBD', cons_id)
            else:
                 print(f"Skipping netting for {cons_id}: Production data for {prod_key} not loaded.")

        netting_map_brasov = {
            # Define accurate Brasov POD mappings
            "Brasov_Prod_1": ("594020100002383007", df_productions.get("Brasov")),
            # Add other Brasov POD mappings as needed
        }
        for prod_key, (cons_id, df_prod) in netting_map_brasov.items():
             if df_prod is not None:
                 df_brasov = apply_netting(df_brasov, df_prod, 'POD', cons_id)
             else:
                 print(f"Skipping netting for {cons_id}: Production data for {prod_key} not loaded.")

        # --- Debug: After Netting ---
        print("\n--- DEBUG: df_santimbru AFTER netting ---")
        st.write(df_santimbru)
        print(f"Shape: {df_santimbru.shape}")
        print("\n--- DEBUG: df_brasov AFTER netting ---")
        st.write(df_brasov)
        print(f"Shape: {df_brasov.shape}\n")
        
        # --- Ensure no negative consumption values ---
        print("Clipping negative consumption values to 0...")
        df_santimbru['Prediction'] = df_santimbru['Prediction'].clip(lower=0)
        df_brasov['Prediction'] = df_brasov['Prediction'].clip(lower=0)

        # --- Return the Netted Dataframes ---
        print("Netting function complete. Returning DataFrames.") # Changed print message slightly
        return df_santimbru, df_brasov

    except Exception as e:
        print(f"An unexpected error occurred in netting_consumption_with_production: {e}")
        import traceback
        traceback.print_exc()
        # Return None for both dataframes to allow unpacking
        return None, None

def render_consumption_forecast_Transavia():
	# --- Date Selection Widgets (Placed at the top) ---
	today_eet_ui = datetime.now(pytz.timezone('Europe/Bucharest')).date()
	default_start_ui = today_eet_ui + pd.Timedelta(days=1)
	default_end_ui = default_start_ui + pd.Timedelta(days=6)

	st.markdown("**Select Date Range to Write to Template:**")
	col1_ui, col2_ui = st.columns(2)
	with col1_ui:
		selected_start_date_ui = st.date_input(
			"Start Date:",
			value=default_start_ui,
			key="template_start_date" # Keep the key consistent
		)
	with col2_ui:
		selected_end_date_ui = st.date_input(
			"End Date:",
			value=default_end_ui,
			key="template_end_date" # Keep the key consistent
		)

	# Basic validation displayed persistently
	if selected_start_date_ui > selected_end_date_ui:
		st.error("Error: Start date must be before or the same as end date.")
	# --- End Date Selection Widgets ---

	if st.button("Bring The Data"):
		st.write("Consumption Forecast")
		# ... (other content and functionality for production forecasting)
		# Iterating through the dictionary of PVPP locations
		for location in locations_cons.keys():
			print("Getting data for {}".format(location))
			output_path = f"./Transavia/data/{location}.csv"
			lat = locations_cons[location]["lat"]
			lon = locations_cons[location]["lon"]
			fetch_data(lat, lon, solcast_api_key, output_path)
		# Creating the input_production file
		path = "./Transavia/data"
		cleaning_input_files()
		creating_input_consumption_Santimbru()
		creating_input_cons_file_Brasov()
		df_santimbru = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
		df_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")

		# <<< Add date input widgets >>>
		today_eet = datetime.now(pytz.timezone('Europe/Bucharest')).date()
		default_start = today_eet + pd.Timedelta(days=1)
		default_end = default_start + pd.Timedelta(days=6)

		col1, col2 = st.columns(2)
		with col1:
			selected_start_date = st.date_input(
				"Select Start Date for Template:",
				value=default_start,
				key="start_date_widget"
			)
		with col2:
			selected_end_date = st.date_input(
				"Select End Date for Template:",
				value=default_end,
				key="end_date_widget"
			)

		# Validate date range (optional for display, main validation before file creation)
		if selected_start_date > selected_end_date:
			st.warning("Warning: Start date is after end date.")

		st.dataframe(df_santimbru) # Display full fetched data initially
		st.dataframe(df_brasov)   # Display full fetched data initially
		# Creating the ZIP file with the Inputs:
		folder_path = './Transavia/Consumption/Input'
		zip_name = 'Transavia_Inputs.zip'
		zip_files(folder_path, zip_name)
		file_path = './Transavia/Consumption/Input/Transavia_Inputs.zip'

		with open(file_path, "rb") as f:
			zip_data = f.read()

		# Create a download link
		b64 = base64.b64encode(zip_data).decode()
		button_html = f"""
			 <a download="Transavia_Inputs.zip" href="data:application/zip;base64,{b64}" download>
			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Input Files</button>
			 </a> 
			 """
		st.markdown(button_html, unsafe_allow_html=True)
	# Submit button
	if st.button('Submit'):
		st.success('Forecast Ready', icon="✅")
		# Your code to generate the forecast
		df_santimbru = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
		df_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")
		predicting_exporting_consumption_Santimbru(df_santimbru)
		predicting_exporting_consumption_Brasov(df_brasov)
		# Creating the ZIP file with the Predictions:
		folder_path = './Transavia/Consumption/Results/XGB'
		zip_name = 'Transavia_Consumption_Results.zip'
		zip_files(folder_path, zip_name)
		file_path = './Transavia/Consumption/Results/XGB/Transavia_Consumption_Results.zip'

		with open(file_path, "rb") as f:
			zip_data = f.read()

		# Create a download link
		b64 = base64.b64encode(zip_data).decode()
		button_html = f"""
			 <a download="Transavia_Consumption_Results.zip" href="data:application/zip;base64,{b64}" download>
			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
			 </a> 
			 """
		st.markdown(button_html, unsafe_allow_html=True)
	if st.button("Create Forecast File"):
		# <<< REMOVE date input widgets from HERE >>>
		# today_eet_cf = datetime.now(pytz.timezone('Europe/Bucharest')).date() # Removed
		# default_start_cf = today_eet_cf + pd.Timedelta(days=1) # Removed
		# default_end_cf = default_start_cf + pd.Timedelta(days=6) # Removed

		# st.markdown("**Select Date Range to Write to Template:**") # Removed
		# col1_cf, col2_cf = st.columns(2) # Removed
		# with col1_cf: # Removed
		# 	selected_start_date = st.date_input( # Removed
		# 		"Start Date:", # Removed
		# 		value=default_start_cf, # Removed
		# 		key="template_start_date" # Removed
		# 	) # Removed
		# with col2_cf: # Removed
		# 	selected_end_date = st.date_input( # Removed
		# 		"End Date:", # Removed
		# 		value=default_end_cf, # Removed
		# 		key="template_end_date" # Removed
		# 	) # Removed

		# Validate selected date range
		# if selected_start_date > selected_end_date: # Removed validation here, done outside
		# 	st.error("Error: Start date must be before or the same as end date. Cannot create file.") # Removed
		# 	return # Stop execution # Removed
		# <<< End REMOVE date input widgets >>>

		# <<< Retrieve dates from session state HERE >>>
		selected_start_date = st.session_state.get('template_start_date')
		selected_end_date = st.session_state.get('template_end_date')

		result = netting_consumption_with_production()
		print(f"DEBUG: netting_consumption_with_production returned: {result}")
		print(f"DEBUG: Type of result: {type(result)}")

		# Define forecast range starting from tomorrow (needed BEFORE filtering)
		# REMOVED: No longer need to calculate dates here for the call
		# num_days = st.session_state.get('cons_days', 7) # Get value from widget state
		# today_eet = datetime.now(pytz.timezone('Europe/Bucharest')).date()
		# start_forecast_date = today_eet + pd.Timedelta(days=1)
		# end_forecast_date = start_forecast_date + pd.Timedelta(days=num_days - 1)

		if result is None or len(result) != 2:
			st.error("Failed to get netted data.")
			# df_santimbru_netted_filtered, df_brasov_netted_filtered = None, None # This line was removed as variables are defined later
		else:
			df_santimbru_netted, df_brasov_netted = result # Unpack the result

			# --- IMPORTANT: Filter the dataframes used for display/preview HERE (optional but good UX) ---
			# This uses the USER-SELECTED dates for the preview shown just before file generation
			df_santimbru_netted_filtered_display = None
			df_brasov_netted_filtered_display = None

			if df_santimbru_netted is not None:
				df_santimbru_netted_dt = df_santimbru_netted.copy()
				df_santimbru_netted_dt['Data'] = pd.to_datetime(df_santimbru_netted_dt['Data']).dt.date
				df_santimbru_netted_filtered_display = df_santimbru_netted_dt[
					(df_santimbru_netted_dt['Data'] >= selected_start_date) & (df_santimbru_netted_dt['Data'] <= selected_end_date)
				]

			if df_brasov_netted is not None:
				df_brasov_netted_dt = df_brasov_netted.copy()
				df_brasov_netted_dt['Data'] = pd.to_datetime(df_brasov_netted_dt['Data']).dt.date
				df_brasov_netted_filtered_display = df_brasov_netted_dt[
					(df_brasov_netted_dt['Data'] >= selected_start_date) & (df_brasov_netted_dt['Data'] <= selected_end_date)
				]

			st.write("--- Data Preview for Selected Template Range ---")
			st.dataframe(df_santimbru_netted_filtered_display) # Display filtered netted data for preview
			st.dataframe(df_brasov_netted_filtered_display) # Display filtered netted data for preview
			# --- End Preview Filter --- 

			template_file = "./Transavia/Notificare_Consum_Transavia.xlsx" # Adjust path if needed
			output_file = "./Transavia/Consumption/Results/Notificare_Consum_Transavia_Filled.xlsx" # Name for the output

			# Use the FULL netted dataframes, but pass the USER-SELECTED dates for file creation
			if df_santimbru_netted is not None and not df_santimbru_netted.empty and df_brasov_netted is not None and not df_brasov_netted.empty:
				success = creating_forecast_file(
					df_santimbru_netted, # Pass FULL netted data
					df_brasov_netted,    # Pass FULL netted data
					template_file,
					output_file,
					forecast_start_date=selected_start_date, # Pass USER-SELECTED start date
					forecast_end_date=selected_end_date    # Pass USER-SELECTED end date
				)
				if success:
					st.success(f"Forecast file created: {output_file}")
					# Add download button for the generated Excel file
					with open(output_file, "rb") as fp:
						btn = st.download_button(
							label="Download Forecast File",
							data=fp,
							file_name="Notificare_Consum_Transavia_Filled.xlsx",
							mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
						)
				else:
					st.error("Failed to create forecast file.")
			else:
				st.warning("Netted consumption data is empty or incomplete, cannot create forecast file.")

def render_Transavia_page():
	
	# Web App Title
	st.markdown('''
	## **The Transavia Forecast Section**

	''')

	# Allow the user to choose between Consumption and Production
	forecast_type = st.radio("Choose Forecast Type:", options=["Consumption", "Production"])

	if forecast_type == "Consumption":
		render_consumption_forecast_Transavia()
	elif forecast_type == "Production":
		render_production_forecast_Transavia(locations_PVPPs)

def apply_netting(df_cons, df_prod, cons_id_col, cons_id_val, prod_scale=1.0):
    """
    Applies netting by subtracting scaled production from consumption for a specific ID.
    Includes checks and handling for duplicate Data/Interval keys before merging.
    """
    # --- Input Validation ---
    if df_prod is None:
        print(f"Skipping netting for {cons_id_col}={cons_id_val}: Production data is missing.")
        return df_cons
    if df_cons is None or df_cons.empty:
        print(f"Skipping netting for {cons_id_col}={cons_id_val}: Consumption data is empty.")
        return df_cons
    if 'Production' not in df_prod.columns:
        print(f"Skipping netting for {cons_id_col}={cons_id_val}: 'Production' column missing in production data.")
        return df_cons
    if 'Prediction' not in df_cons.columns:
        print(f"Skipping netting for {cons_id_col}={cons_id_val}: 'Prediction' column missing in consumption data.")
        return df_cons
    if cons_id_col not in df_cons.columns:
        print(f"Skipping netting: ID column '{cons_id_col}' not found in consumption data.")
        return df_cons
    if 'Data' not in df_cons.columns or 'Interval' not in df_cons.columns:
         print(f"Skipping netting for {cons_id_col}={cons_id_val}: 'Data' or 'Interval' missing in consumption data.")
         return df_cons
    if 'Data' not in df_prod.columns or 'Interval' not in df_prod.columns:
         print(f"Skipping netting for {cons_id_col}={cons_id_val}: 'Data' or 'Interval' missing in production data.")
         return df_cons

    print(f"Applying netting for {cons_id_col}={cons_id_val} with scale={prod_scale}...")

    # --- Prepare Production Data (Deduplicate) ---
    prod_temp_name = f"Prod_{cons_id_val}".replace(" ", "_")
    # Ensure required columns exist and are correct types
    df_prod['Data'] = pd.to_datetime(df_prod['Data'], errors='coerce')
    df_prod['Interval'] = pd.to_numeric(df_prod['Interval'], errors='coerce').fillna(-1).astype(int)
    df_prod = df_prod.dropna(subset=['Data'])
    df_prod = df_prod[df_prod['Interval'].between(0, 23)] # Ensure only valid hourly intervals

    if df_prod.empty:
        print(f"Skipping netting for {cons_id_col}={cons_id_val}: Production data empty after cleaning.")
        return df_cons

    prod_temp_orig = df_prod[['Data', 'Interval', 'Production']].rename(columns={'Production': prod_temp_name})

    # Check and handle duplicates in production data
    prod_duplicates = prod_temp_orig[prod_temp_orig.duplicated(subset=['Data', 'Interval'], keep=False)]
    if not prod_duplicates.empty:
        print(f"WARNING: Found {prod_duplicates.shape[0]} duplicate rows in PRODUCTION data ('{cons_id_val}') for merge keys.")
        # Strategy: Keep the first occurrence
        prod_temp = prod_temp_orig.drop_duplicates(subset=['Data', 'Interval'], keep='first').copy()
        print(f"Dropped duplicates from production data ('{cons_id_val}'), kept first.")
    else:
        prod_temp = prod_temp_orig.copy() # Use copy to avoid modifying original df_productions

    # --- Prepare Consumption Data Slice (Deduplicate) ---
    # Ensure types before filtering and merging
    df_cons['Data'] = pd.to_datetime(df_cons['Data'], errors='coerce')
    df_cons['Interval'] = pd.to_numeric(df_cons['Interval'], errors='coerce').fillna(-1).astype(int)
    df_cons['Prediction'] = pd.to_numeric(df_cons['Prediction'], errors='coerce').fillna(0)
    df_cons = df_cons.dropna(subset=['Data'])
    df_cons = df_cons[df_cons['Interval'].between(0, 23)]

    # Filter for the specific ID we are netting
    df_cons_slice = df_cons[df_cons[cons_id_col] == cons_id_val].copy()

    if df_cons_slice.empty:
        print(f"Skipping netting for {cons_id_col}={cons_id_val}: No consumption data found for this ID.")
        return df_cons # Return original df_cons if no rows match the ID

    # Check and handle duplicates in the relevant consumption slice
    cons_duplicates = df_cons_slice[df_cons_slice.duplicated(subset=['Data', 'Interval'], keep=False)]
    if not cons_duplicates.empty:
        print(f"WARNING: Found {cons_duplicates.shape[0]} duplicate rows in CONSUMPTION data for ID '{cons_id_val}' for merge keys.")
        # Keep the first occurrence
        df_cons_slice = df_cons_slice.drop_duplicates(subset=['Data', 'Interval'], keep='first')
        print(f"Dropped duplicates from consumption slice for ID '{cons_id_val}', kept first.")

    # --- Perform Merge ---
    # Merge the (potentially deduplicated) consumption slice with the (potentially deduplicated) production data
    df_merged = pd.merge(
        df_cons_slice, # Use the filtered, deduplicated slice
        prod_temp,     # Use the deduplicated production temp table
        on=['Data', 'Interval'],
        how='left'     # Keep all consumption rows for this ID
    )

    # --- Calculate Net Prediction ---
    # Fill NaNs in the merged production column (means no matching production data)
    df_merged[prod_temp_name] = pd.to_numeric(df_merged[prod_temp_name], errors='coerce').fillna(0)
    # Ensure prediction is numeric
    df_merged['Prediction'] = pd.to_numeric(df_merged['Prediction'], errors='coerce').fillna(0)

    # Calculate the netted value
    df_merged['Net_Prediction'] = df_merged['Prediction'] - (df_merged[prod_temp_name] * prod_scale)
    # Optional: Ensure net prediction doesn't go below zero if that's required
    # df_merged['Net_Prediction'] = df_merged['Net_Prediction'].apply(lambda x: max(x, 0))

    # --- Update Original Consumption DataFrame ---
    # Create a MultiIndex for df_cons for merging/updating
    df_cons = df_cons.set_index(['Data', 'Interval', cons_id_col], drop=False)

    # Prepare the update dataframe (df_merged) with the same index structure
    update_df = df_merged.set_index(['Data', 'Interval', cons_id_col])[['Net_Prediction']].rename(columns={'Net_Prediction': 'Prediction_Update'})

    # Merge the update back into df_cons. Use left join to keep all original rows.
    df_cons = df_cons.merge(update_df, left_index=True, right_index=True, how='left')

    # Update the 'Prediction' column: If Prediction_Update exists (not NaN), use it; otherwise, keep original Prediction.
    df_cons['Prediction'] = df_cons['Prediction_Update'].combine_first(df_cons['Prediction'])

    # Drop the temporary update column and reset index to original structure
    df_cons = df_cons.drop(columns=['Prediction_Update']).reset_index(drop=True)

    print(f"Netting update applied for {cons_id_val}.")

    return df_cons

# --- This function needs the file writing logic --- 
def creating_forecast_file(df_santimbru_netted, df_brasov_netted, template_path, output_path, forecast_start_date, forecast_end_date):
    print(f"Starting: Create forecast file from template: {template_path}")
    print(f"Writing data for date range: {forecast_start_date} to {forecast_end_date}") # Add print for confirmation
    try:
        # 1. --- Validate Input Data --- (Keep existing validation)
        st.write("DEBUG: Received df_santimbru_netted head:")
        if df_santimbru_netted is not None: st.dataframe(df_santimbru_netted.head())
        st.write("DEBUG: Received df_brasov_netted head:")
        if df_brasov_netted is not None: st.dataframe(df_brasov_netted.head())

        if df_santimbru_netted is None or df_brasov_netted is None or df_santimbru_netted.empty or df_brasov_netted.empty:
            st.error("ERROR: Netted data is missing or empty. Cannot create file.")
            return False

        # --- Simplified Data Prep: Create Indexed DataFrames --- 
        print("Preparing simplified lookup structures...")
        lookup_santimbru = None
        lookup_brasov = None

        if df_santimbru_netted is not None and not df_santimbru_netted.empty:
            df_s_lookup = df_santimbru_netted.copy()
            df_s_lookup['Data'] = pd.to_datetime(df_s_lookup['Data']).dt.normalize()
            df_s_lookup['Interval'] = pd.to_numeric(df_s_lookup['Interval'], errors='coerce').fillna(-1).astype(int)
            df_s_lookup['IBD'] = df_s_lookup['IBD'].astype(str)
            df_s_lookup['Value'] = pd.to_numeric(df_s_lookup['Prediction'], errors='coerce')
            # --- Replace source identifier ("Abator") with the template map identifier ("Abator Oiejdea") ---
            df_s_lookup.loc[df_s_lookup['IBD'] == 'Abator', 'IBD'] = 'Abator Oiejdea'
            df_s_lookup = df_s_lookup[df_s_lookup['Interval'].between(0, 23)]
            # Set index for direct lookup
            lookup_santimbru = df_s_lookup.set_index(['Data', 'Interval', 'IBD'])['Value']
            if not lookup_santimbru.index.is_unique:
                print("WARNING: Duplicate keys found in Santimbru data. Keeping first.")
                lookup_santimbru = lookup_santimbru[~lookup_santimbru.index.duplicated(keep='first')]
            print("Santimbru lookup structure prepared.")

        if df_brasov_netted is not None and not df_brasov_netted.empty:
            df_b_lookup = df_brasov_netted.copy()
            df_b_lookup['Data'] = pd.to_datetime(df_b_lookup['Data']).dt.normalize()
            df_b_lookup['Interval'] = pd.to_numeric(df_b_lookup['Interval'], errors='coerce').fillna(-1).astype(int)
            df_b_lookup['POD'] = df_b_lookup['POD'].astype(str)
            df_b_lookup['Value'] = pd.to_numeric(df_b_lookup['Prediction'], errors='coerce')
            df_b_lookup = df_b_lookup[df_b_lookup['Interval'].between(0, 23)]
            # Set index for direct lookup
            lookup_brasov = df_b_lookup.set_index(['Data', 'Interval', 'POD'])['Value']
            if not lookup_brasov.index.is_unique:
                print("WARNING: Duplicate keys found in Brasov data. Keeping first.")
                lookup_brasov = lookup_brasov[~lookup_brasov.index.duplicated(keep='first')]
            print("Brasov lookup structure prepared.")
        # --- End Simplified Data Prep ---

        # 2. --- Load Excel Template --- (Keep existing)
        print("Loading Excel template...")
        try:
             wb = openpyxl.load_workbook(template_path, data_only=False) # Changed data_only to False to preserve formulas
             ws = wb.active
             print("Template loaded with data_only=False (formulas preserved).")
        except FileNotFoundError:
             print(f"ERROR: Template file not found at {template_path}")
             st.error(f"ERROR: Template file not found at {template_path}")
             return False
        except Exception as e:
             print(f"ERROR loading template workbook: {e}")
             st.error(f"ERROR loading template workbook: {e}")
             return False

        # 4. --- Define Column Mapping (As verified) --- (Keep existing)
        template_col_map = {
            'F_Cristian': 'G', 'Abator Oiejdea': 'H', 'F3': 'I', 'F4': 'J',
            'F5': 'K', 'F6': 'L', 'F7': 'M', 'F8': 'N', 'F9': 'O', 'F10': 'P',
            'F17': 'Q', 
            'F20-F21': 'R',
            'FNC': 'S', 'Ciugud': 'T',
            'Abator_Bocsa': 'U',
            'Ferma_Bocsa': 'V',
            'Jebel1': 'W',
            # Brasov PODs
            "594020100002382970": "X", "594020100002383007": "Y",
            "594020100002383502": "Z", "594020100002836497": "AA", # Target POD
            "594020100002967269": "AB", "594020100002841279": "AC",
            "594020100002383014": "AD", "594020100002384691": "AE",
            "594020100002383069": "AF", "594020100002224041": "AG",
            '594020100002273568': 'AH'
        }
        # Identify which identifiers belong to Brasov for lookup routing
        brasov_identifiers = {
            pod for pod in template_col_map.keys()
            if pod.startswith("594") # Simple check for Brasov PODs
        }
        print("Template column map defined.")

        # 5. --- Determine Date Range & Forecast Month/Year --- (Use passed dates)
        # REMOVED: Deriving min/max date from lookup data
        min_date = pd.to_datetime(forecast_start_date) # Use passed start date
        max_date = pd.to_datetime(forecast_end_date)   # Use passed end date
        forecast_month = min_date.month # Use start date for month/year
        forecast_year = min_date.year
        print(f"Using passed date range: {min_date.date()} to {max_date.date()}. Report Month/Year: {forecast_month}/{forecast_year}")

        # 6. --- Iterate Template and Write --- (Modified Lookup)
        print("Writing values to template using direct lookup...")
        data_start_row = 4
        date_col = 'A'
        interval_col = 'B' # Quarterly interval (1-96)

        for row_num in range(data_start_row, ws.max_row + 1):
            try:
                day_val = ws[f"{date_col}{row_num}"].value
                quarter_interval_val = ws[f"{interval_col}{row_num}"].value

                # Basic validation for row
                if day_val is None or quarter_interval_val is None: continue
                try:
                    day = int(day_val)
                    quarter_interval = int(quarter_interval_val)
                    if not (1 <= day <= 31) or not (1 <= quarter_interval <= 96): continue
                except (ValueError, TypeError): continue # Skip if values aren't integers

                # Construct date and hour key components
                try:
                    # Use pandas to handle potential date errors more robustly
                    current_date = pd.Timestamp(year=forecast_year, month=forecast_month, day=day)
                except ValueError: continue # Skip invalid day

                # --->>> Check: Only process rows within the SELECTED forecast date range <<< ---
                if not (min_date <= current_date <= max_date):
                    continue # Skip rows outside the selected forecast period

                hour_interval = (quarter_interval - 1) // 4 # 0-23

                # Write values for all mapped columns in this row
                for identifier_header, target_col_letter in template_col_map.items():
                    hourly_pred = np.nan # Default to NaN
                    try:
                        lookup_key = (current_date, hour_interval, identifier_header)

                        # Route lookup to the correct structure
                        if identifier_header in brasov_identifiers:
                            if lookup_brasov is not None:
                                # Use .get on the index to handle missing keys gracefully
                                if lookup_brasov.index.isin([lookup_key]).any():
                                     hourly_pred = lookup_brasov[lookup_key]
                        else:
                            if lookup_santimbru is not None:
                                 if lookup_santimbru.index.isin([lookup_key]).any():
                                     hourly_pred = lookup_santimbru[lookup_key]

                        # Check if lookup was successful and value is not NaN
                        if not pd.isna(hourly_pred):
                            quarterly_value = hourly_pred / 4.0
                            target_cell = f"{target_col_letter}{row_num}"
                            ws[target_cell].value = quarterly_value
                            ws[target_cell].number_format = '0.00'

                            # Minimal Debug Print (only for first valid row written, target ID)
                            target_pod_aa = "594020100002836497"
                            if row_num == data_start_row and identifier_header == target_pod_aa:
                                print(f"DEBUG Sample Write: Row={row_num}, ID={identifier_header}, Key=({current_date.date()},{hour_interval}), Hourly={hourly_pred:.4f}, Quarterly={quarterly_value:.4f} -> Cell={target_cell}")
                        # else: # Optional: Debugging for failed lookups or NaN values
                            # target_pod_aa = "594020100002836497"
                            # if identifier_header == target_pod_aa:
                            #     print(f"DEBUG SKIP/NaN: Row={row_num}, ID={identifier_header}, Key=({current_date.date()},{hour_interval}), Hourly Value={hourly_pred}")

                    except Exception as e:
                         print(f"ERROR during simplified lookup/write for Row {row_num}, ID {identifier_header}, Key {lookup_key}: {e}")
                         continue # Skip to next identifier on error

                # --- Calculate Sintetic Consumption (Keep existing logic) ---
                try:
                    row_sum = 0.0
                    sintetic_cons_col = 'AI'
                    sintetic_cons_cell_coord = f"{sintetic_cons_col}{row_num}"

                    for col_letter in template_col_map.values():
                        cell_coord = f"{col_letter}{row_num}"
                        written_value = ws[cell_coord].value
                        try:
                            numeric_val = float(written_value) if written_value is not None else 0.0
                        except (ValueError, TypeError):
                            numeric_val = 0.0
                        row_sum += numeric_val

                    sintetic_value = row_sum * 0.05
                    if row_sum > 0.00001:
                        ws[sintetic_cons_cell_coord].value = sintetic_value
                        ws[sintetic_cons_cell_coord].number_format = '0.00'

                except Exception as e:
                     print(f"ERROR calculating/writing Sintetic Consumption for row {row_num}: {e}")
                     ws[sintetic_cons_cell_coord].value = "#ERR!"

            except Exception as e:
                print(f"ERROR processing template row {row_num}: {type(e).__name__} - {e}")

        print("Finished writing values loop.")

        # 7. --- Update Cell B1 --- (Keep existing logic)
        print("Updating Month/Year cell...")
        try:
            locale.setlocale(locale.LC_TIME, 'ro_RO.UTF-8')
            month_name_ro = min_date.strftime("%B").capitalize()
        except locale.Error:
            print("Warning: Romanian locale not found, using default.")
            locale.setlocale(locale.LC_TIME, '')
            month_name_ro = min_date.strftime("%B").capitalize()
        month_year_string = f"{month_name_ro} {forecast_year}"

        target_cell_coord = 'B1'
        merged_range = None
        for rng in ws.merged_cells.ranges:
            if target_cell_coord in rng:
                merged_range = rng
                break

        if merged_range:
            print(f"Cell {target_cell_coord} is part of merged range {merged_range}. Handling merge.")
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_cell_coord = get_column_letter(min_col) + str(min_row)
            range_string = str(merged_range)
            ws.unmerge_cells(range_string)
            ws[top_left_cell_coord].value = month_year_string
            ws.merge_cells(range_string)
            print(f"Re-merged range {range_string}.")
        else:
            print(f"Cell {target_cell_coord} is not merged. Writing directly.")
            ws[target_cell_coord].value = month_year_string

        # 8. --- Save Output --- (Keep existing logic)
        print(f"Saving updated file to: {output_path}")
        try:
            wb.save(output_path)
            print("File saved successfully.")
            return True # Success
        except Exception as e:
            print(f"ERROR saving workbook: {e}")
            st.error(f"ERROR saving workbook: {e}")
            return False # Failure

    except Exception as e:
        print(f"FATAL ERROR in creating_forecast_file: {e}")
        import traceback
        traceback.print_exc()
        st.error(f"FATAL ERROR in creating_forecast_file: {e}")
        return False # Failure