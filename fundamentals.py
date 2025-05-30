import streamlit as st
import pandas as pd
import numpy as np
import requests
from datetime import datetime, timedelta, time
import os
from dotenv import load_dotenv
import wapi
from openpyxl import load_workbook
from datetime import datetime
from entsoe import EntsoePandasClient
import xml.etree.ElementTree as ET
import openpyxl
import base64
import zipfile
import joblib
import xlsxwriter
from datetime import date
import pytz

# ================================================================================VOLUE data==================================================================================

# Set the date dinamically as today
# Get the current date with time set to 00:00
issue_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
print(issue_date)
# Format the date as a string in the desired format
issue_date_str = issue_date.isoformat()
issue_date_str = issue_date.strftime('%Y-%m-%dT%H:%M')
print(issue_date_str)

issue_date_str_yesterday = issue_date - timedelta(days=1)
issue_date_str_yesterday = issue_date_str_yesterday.strftime('%Y-%m-%dT%H:%M')

def get_issue_date():
	issue_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
	print(issue_date)
	# Format the date as a string in the desired format
	issue_date_str = issue_date.strftime('%Y-%m-%dT%H:%M')
	print(issue_date_str)
	return issue_date_str, issue_date

# Fetching the token and store it's expiring timestamp
load_dotenv()
client_id = os.getenv("volue_client_id")
client_secret = os.getenv("volue_client_secret")

# Replace 'client_id' and 'client_secret' with your actual credentials
session = wapi.Session(client_id = client_id, client_secret = client_secret)

def fetch_token(client_id, client_secret):
	"""
	Fetches a new access token using client credentials.

	Args:
		client_id (str): The client ID provided by Volue API.
		client_secret (str): The client secret provided by Volue API.

	Returns:
		dict: A dictionary containing the access token, token type, and expiration timestamp.
	"""
	url = "https://auth.volueinsight.com/oauth2/token"
	auth = requests.auth.HTTPBasicAuth(client_id, client_secret)
	headers = {"Content-Type": "application/x-www-form-urlencoded"}
	data = {"grant_type": "client_credentials"}

	response = requests.post(url, headers=headers, data=data, auth=auth)
	
	if response.status_code == 200:
		token_info = response.json()
		# Calculate expiration timestamp
		expires_in = token_info.get("expires_in", 3600)  # Default to 3600 seconds
		expiration_timestamp = datetime.now() + timedelta(seconds=expires_in)
		
		return {
			"access_token": token_info["access_token"],
			"token_type": token_info["token_type"],
			"expires_at": expiration_timestamp
		}
	else:
		raise Exception("Failed to fetch token: " + response.text)

# Example usage:
# token_info = fetch_token(client_id, client_secret)
# print(token_info)

def is_token_valid(token_info):
	"""
	Checks if the current access token is valid or needs to be refreshed.

	Args:
		token_info (dict): The dictionary containing the token information.

	Returns:
		bool: True if the token is valid, False otherwise.
	"""
	if token_info is None:
		return False
	
	# Consider token as expired if it's close to expiration time (e.g., 5 minutes buffer)
	print(token_info["expires_at"], datetime.now() + timedelta(minutes=5))
	return datetime.now() + timedelta(minutes=5) < token_info["expires_at"]

# Usage example
# if not is_token_valid(token_info):
# 	token_info = fetch_token(client_id, client_secret)
# 	# Update your storage with the new token_info


# Fetching the curves==================================
# Wind curve horizon = days

def fetch_curve(token, curve_name):
	"""
	Fetches curve data by name using the provided access token.

	Args:
		token (str): The access token for authorization.
		curve_name (str): The name of the curve to fetch.

	Returns:
		dict: A dictionary containing the curve data, or an error message.
	"""
	url = "https://api.volueinsight.com/api/instances/"
	headers = {
		"Authorization": f"Bearer {token}",
		"Content-Type": "application/json"
	}
	params = {
		"name": curve_name  # Filter by curve name
	}
	
	response = requests.get(url, headers=headers, params=params)
	
	if response.status_code == 200:
		return response.json()  # Returns the curve data
	else:
		return {"error": "Failed to fetch curve data", "status_code": response.status_code, "message": response.text}

# Usage example (Assuming `token_info` is a dictionary containing your valid access token)
# curve_name = "pro ro wnd ec00 mwh/h cet min15 f"
# curve_data = fetch_curve(token_info['access_token'], curve_name)
# print(curve_data)


def fetch_time_series_data(token, curve_id, start_date, end_date, time_zone=None, output_time_zone=None, filter=None, function=None, frequency=None):
	"""
	Fetches time series data for a specified curve.

	Args:
		token (str): The access token for authorization.
		curve_id (int): The ID of the curve to fetch data for.
		start_date (str): The start date for the data range in YYYY-MM-DD format.
		end_date (str): The end date for the data range in YYYY-MM-DD format.
		time_zone (str): Optional. The curve time zone before filtering and frequency change.
		output_time_zone (str): Optional. The curve time zone after filtering and frequency change.
		filter (str): Optional. Filter out parts of the time series.
		function (str): Optional. The aggregation/split function to use when changing frequency.
		frequency (str): Optional. The required frequency of the output.

	Returns:
		dict: A dictionary containing the time series data for the curve, or an error message.
	"""
	url = f"https://api.volueinsight.com/api/series/{curve_id}"
	headers = {
		"Authorization": f"Bearer {token}",
		"Content-Type": "application/json"
	}
	params = {
		"from": start_date,
		"to": end_date,
	}
	
	# Optional parameters
	if time_zone:
		params['time_zone'] = time_zone
	if output_time_zone:
		params['output_time_zone'] = output_time_zone
	if filter:
		params['filter'] = filter
	if function:
		params['function'] = function
	if frequency:
		params['frequency'] = frequency

	response = requests.get(url, headers=headers, params=params)

	if response.status_code == 200:
		return response.json()  # Returns the time series data for the curve
	else:
		return {"error": "Failed to fetch time series data", "status_code": response.status_code, "message": response.text}

#=================================Fetching the Wind Power data============================================
def fetch_volue_wind_data(issue_date_str):
	# INSTANCES curve 15 min
	curve = session.get_curve(name='pro ro wnd ec00 mwh/h cet min15 f')
	# INSTANCES curves contain a timeseries for each defined issue dates
	# Get a list of available curves with issue dates within a timerange with:
	# curve.search_instances(issue_date_from='2018-01-01', issue_date_to='2018-01-01')
	ts_15min = curve.get_instance(issue_date=issue_date_str)
	df_wind_15min = ts_15min.to_pandas() # convert TS object to pandas.Series object
	df_wind_15min = df_wind_15min.to_frame() # convert pandas.Series to pandas.DataFrame
	st.dataframe(df_wind_15min)
	df_wind_15min.to_csv("./Market Fundamentals/Wind_data_15min.csv")
	# INSTANCES curve hour
	# curve = session.get_curve(name='pro ro wnd fwd mw cet h f')
	# # INSTANCES curves contain a timeseries for each defined issue dates
	# # Get a list of available curves with issue dates within a timerange with:
	# # curve.search_instances(issue_date_from='2018-01-01', issue_date_to='2018-01-01')
	# ts_h = curve.get_instance(issue_date=issue_date_str)
	# pd_s_h = ts_h.to_pandas() # convert TS object to pandas.Series object
	# pd_df_h = pd_s_h.to_frame() # convert pandas.Series to pandas.DataFrame
	# st.dataframe(pd_df_h)
	# pd_df_h.to_csv("./Market Fundamentals/Wind_data_hourly.csv")

	# Writing the hourly values to the Trading Tool file
	# Load the wind data from CSV without altering the date format
	wind_data_path = './Market Fundamentals/Wind_data_hourly.csv'  # Update with the actual path
	wind_data = pd.read_csv(wind_data_path)

	# Determine tomorrow's date as a string to match your CSV format
	tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')

	# Filter rows based on the string representation of tomorrow's date
	# Assuming the date in your CSV is in the format 'YYYY-MM-DD' and is in the first column
	# tomorrow_data = wind_data[wind_data.iloc[:, 0].str.startswith(tomorrow)]
	# st.dataframe(tomorrow_data)
	# # Write to Excel
	# excel_file_path = './Market Fundamentals/Volue_data.xlsx'  # Update with the actual path
	# workbook = load_workbook(excel_file_path)
	# sheet = workbook["Volue_Data_eng"] 

	# # Confirm we have data to write
	# if len(tomorrow_data) > 0:
	# 	excel_row = 4
	# 	for index, row in tomorrow_data.iterrows():
	# 		if excel_row == 12 or excel_row == 25:
	# 			excel_row += 1
	# 		cell = f'E{excel_row}'
	# 		sheet[cell] = row[1]  # Assuming data to write is in the second column
	# 		print(f"Writing {row[1]} to {cell}")  # Diagnostic print to confirm writing
	# 		excel_row += 1
	# 	workbook.save(filename=excel_file_path)
	# 	print("Excel file has been updated.")
	# else:
	# 	print("No data available for tomorrow to write into the Excel file.")

	# Writing the quarterly wind data to the Volue data Excel file
	wind_data_path = './Market Fundamentals/Wind_data_15min.csv'  # Update with the actual path
	wind_data = pd.read_csv(wind_data_path)
	# Write to Excel
	excel_file_path = './Market Fundamentals/Volue_data.xlsx'  # Update with the actual path
	workbook = load_workbook(excel_file_path)
	sheet = workbook["Volue_Data_eng"] 
	excel_row = 4
	for index, row in wind_data.iterrows():
		cell = f'AQ{excel_row}'
		sheet[cell] = row[0]  # Assuming data to write is in the second column
		cell = f'AR{excel_row}'
		sheet[cell] = row[1]  # Assuming data to write is in the second column
		print(f"Writing {row[1]} to {cell}")  # Diagnostic print to confirm writing
		excel_row += 1
	workbook.save(filename=excel_file_path)
	print("Excel file has been updated.")

	return df_wind_15min
#=================================Fetching the Solar Power data============================================
def fetch_volue_solar_data(issue_date_str):
	# INSTANCE curve 15 min
	curve = session.get_curve(name='pro ro spv ec00 mwh/h cet min15 f')
	# INSTANCES curves contain a timeseries for each defined issue dates
	# Get a list of available curves with issue dates within a timerange with:
	# curve.search_instances(issue_date_from='2018-01-01', issue_date_to='2018-01-01')
	ts_15min = curve.get_instance(issue_date=issue_date_str)
	pd_s_15min = ts_15min.to_pandas() # convert TS object to pandas.Series object
	df_solar_15min = pd_s_15min.to_frame() # convert pandas.Series to pandas.DataFrame
	st.dataframe(df_solar_15min)
	df_solar_15min.to_csv("./Market Fundamentals/PV_data_15min.csv")
	## INSTANCE curve hour
	# curve = session.get_curve(name='pro ro spv fwd mw cet h f')
	# INSTANCES curves contain a timeseries for each defined issue dates
	# Get a list of available curves with issue dates within a timerange with:
	# curve.search_instances(issue_date_from='2018-01-01', issue_date_to='2018-01-01')
	# ts_h = curve.get_instance(issue_date=issue_date_str)
	# pd_s_h = ts_h.to_pandas() # convert TS object to pandas.Series object
	# pd_df_h = pd_s_h.to_frame() # convert pandas.Series to pandas.DataFrame
	# st.dataframe(pd_df_h)
	# # Writing the hourly values to the Trading Tool file
	# pd_df_h.to_csv("./Market Fundamentals/PV_data_hourly.csv")
	# # Load the wind data from CSV without altering the date format
	# pv_data_path = './Market Fundamentals/PV_data_hourly.csv'  # Update with the actual path
	# pv_data = pd.read_csv(pv_data_path)

	# # Determine tomorrow's date as a string to match your CSV format
	# tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')

	# # Filter rows based on the string representation of tomorrow's date
	# # Assuming the date in your CSV is in the format 'YYYY-MM-DD' and is in the first column
	# tomorrow_data = pv_data[pv_data.iloc[:, 0].str.startswith(tomorrow)]
	# st.dataframe(tomorrow_data)
	# # Write to Excel
	# excel_file_path = './Market Fundamentals/Volue_data.xlsx'  # Update with the actual path
	# workbook = load_workbook(excel_file_path)
	# sheet = workbook["Volue_Data_eng"] 

	# # Confirm we have data to write
	# if len(tomorrow_data) > 0:
	# 	excel_row = 4
	# 	for index, row in tomorrow_data.iterrows():
	# 		if excel_row == 12 or excel_row == 25:
	# 			excel_row += 1
	# 		cell = f'P{excel_row}'
	# 		sheet[cell] = row[1]  # Assuming data to write is in the second column
	# 		print(f"Writing {row[1]} to {cell}")  # Diagnostic print to confirm writing
	# 		excel_row += 1
	# 	workbook.save(filename=excel_file_path)
	# 	print("Excel file has been updated.")
	# else:
	# 	print("No data available for tomorrow to write into the Excel file.")

	# Writing the quarterly pv data to the Volue data Excel file
	pv_data_path = './Market Fundamentals/PV_data_15min.csv'  # Update with the actual path
	pv_data = pd.read_csv(pv_data_path)
	# Write to Excel
	excel_file_path = './Market Fundamentals/Volue_data.xlsx'  # Update with the actual path
	workbook = load_workbook(excel_file_path)
	sheet = workbook["Volue_Data_eng"] 
	excel_row = 4
	for index, row in pv_data.iterrows():
		cell = f'AY{excel_row}'
		sheet[cell] = row[0]  # Assuming data to write is in the second column
		cell = f'AZ{excel_row}'
		sheet[cell] = row[1]  # Assuming data to write is in the second column
		print(f"Writing {row[1]} to {cell}")  # Diagnostic print to confirm writing
		excel_row += 1
	workbook.save(filename=excel_file_path)
	print("Excel file has been updated.")
	return df_solar_15min
#=================================Fetching the Hydro Power data============================================
def fetch_volue_hydro_data(issue_date_str):
	# INSTANCE curve hour
	curve = session.get_curve(name='pro ro hydro tot mwh/h cet h f')
	# INSTANCES curves contain a timeseries for each defined issue dates
	# Get a list of available curves with issue dates within a timerange with:
	# curve.search_instances(issue_date_from='2018-01-01', issue_date_to='2018-01-01')
	try:
		ts_h = curve.get_instance(issue_date=issue_date_str)
		pd_s_h = ts_h.to_pandas() # convert TS object to pandas.Series object
	except:
		st.write("No data available for tomorrow to write into the Excel file.")
		ts_h = curve.get_instance(issue_date=issue_date_str_yesterday)
		pd_s_h = ts_h.to_pandas() # convert TS object to pandas.Series object

	pd_df_h = pd_s_h.to_frame() # convert pandas.Series to pandas.DataFrame
	st.dataframe(pd_df_h)
	# Writing the hourly values to the Trading Tool file
	pd_df_h.to_csv("./Market Fundamentals/Hydro_data_hourly.csv")
	# Load the wind data from CSV without altering the date format
	hydro_data_path = './Market Fundamentals/Hydro_data_hourly.csv'  # Update with the actual path
	hydro_data = pd.read_csv(hydro_data_path)

	# Writing the quarterly hydro data to the Volue data Excel file
	hydro_data_path = './Market Fundamentals/Hydro_data_hourly.csv'  # Update with the actual path
	hydro_data = pd.read_csv(hydro_data_path)
	# Write to Excel
	excel_file_path = './Market Fundamentals/Volue_data.xlsx'  # Update with the actual path
	workbook = load_workbook(excel_file_path)
	sheet = workbook["Volue_Data_eng"] 
	excel_row = 4
	for index, row in hydro_data.iterrows():
		cell = f'AI{excel_row}'
		sheet[cell] = row[0]  # Assuming data to write is in the second column
		cell = f'AJ{excel_row}'
		sheet[cell] = row[1]  # Assuming data to write is in the second column
		print(f"Writing {row[1]} to {cell}")  # Diagnostic print to confirm writing
		excel_row += 1
	workbook.save(filename=excel_file_path)
	print("Excel file has been updated.")
	# Creating the 15 data dataframe in order to add it to the Volue dataframe
	# Load the CSV file
	data = pd_df_h.copy()
	data.reset_index(inplace=True)

	# Rename columns for clarity
	data.rename(columns={
		data.columns[0]: 'Date',
		data.columns[1]: 'Hydro Power'
	}, inplace=True)

	# Convert the date column to datetime
	data['Date'] = pd.to_datetime(data['Date'])

	# Set the date column as the index
	data.set_index('Date', inplace=True)

	# Resample the data to quarter-hourly intervals, filling forward the hydro power values
	quarterly_data = data.resample('15T').ffill()

	# Reset the index to move the date back to a column
	quarterly_data.reset_index(inplace=True)

	# Add the Interval column which represents the quarter-hourly intervals in the day (1 to 96)
	quarterly_data['Interval'] = quarterly_data['Date'].dt.hour * 4 + quarterly_data['Date'].dt.minute // 15 + 1

	# Create a column for the formatted date without time information
	quarterly_data['Formatted Date'] = quarterly_data['Date'].dt.strftime('%d.%m.%Y')

	# Reorder and rename columns to match the specified output
	quarterly_data = quarterly_data[['Formatted Date', 'Interval', 'Hydro Power']]
	quarterly_data.rename(columns={'Formatted Date': 'Date'}, inplace=True)

	return quarterly_data
#=================================Fetching the Temperature data============================================
def fetch_volue_temperature_data(issue_date_str):
	# INSTANCE curve 15 min
	curve = session.get_curve(name='tt ro con ec00 °c cet min15 f')
	# INSTANCES curves contain a timeseries for each defined issue dates
	# Get a list of available curves with issue dates within a timerange with:
	# curve.search_instances(issue_date_from='2018-01-01', issue_date_to='2018-01-01')
	ts_15min = curve.get_instance(issue_date=issue_date_str)
	pd_s_15min = ts_15min.to_pandas() # convert TS object to pandas.Series object
	pd_df_15min = pd_s_15min.to_frame() # convert pandas.Series to pandas.DataFrame
	st.dataframe(pd_df_15min)
	pd_df_15min.to_csv("./Market Fundamentals/Temperature_data_15min.csv")

	# Writing the quarterly pv data to the Volue data Excel file
	temp_data_path = './Market Fundamentals/Temperature_data_15min.csv'  # Update with the actual path
	temp_data = pd.read_csv(temp_data_path)
	# Write to Excel
	excel_file_path = './Market Fundamentals/Volue_data.xlsx'  # Update with the actual path
	workbook = load_workbook(excel_file_path)
	sheet = workbook["Volue_Data_eng"] 
	excel_row = 4
	for index, row in temp_data.iterrows():
		cell = f'BN{excel_row}'
		sheet[cell] = row[0]  # Assuming data to write is in the second column
		cell = f'BO{excel_row}'
		sheet[cell] = row[1]  # Assuming data to write is in the second column
		print(f"Writing {row[1]} to {cell}")  # Diagnostic print to confirm writing
		excel_row += 1
	workbook.save(filename=excel_file_path)
	print("Excel file has been updated.")
	return pd_df_15min
#=================================Fetching the Price data==================================================
def fetch_volue_price_data(issue_date_str):
	# INSTANCE curve hour
	curve = session.get_curve(name='pri ro spot merged ron/mwh cet h f')
	# INSTANCES curves contain a timeseries for each defined issue dates
	# Get a list of available curves with issue dates within a timerange with:
	# curve.search_instances(issue_date_from='2018-01-01', issue_date_to='2018-01-01')
	ts_h = curve.get_instance(issue_date=issue_date_str)
	pd_s_h = ts_h.to_pandas() # convert TS object to pandas.Series object
	pd_df_h = pd_s_h.to_frame() # convert pandas.Series to pandas.DataFrame
	st.dataframe(pd_df_h)
	# Writing the hourly values to the Trading Tool file
	pd_df_h.to_csv("./Market Fundamentals/Price_data_hourly.csv")
	# Load the wind data from CSV without altering the date format
	price_data_path = './Market Fundamentals/Price_data_hourly.csv'  # Update with the actual path
	price_data = pd.read_csv(price_data_path)

	# Determine tomorrow's date as a string to match your CSV format
	tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
	two_days_ahead = (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d')
	three_days_ahead = (datetime.now() + timedelta(days=3)).strftime('%Y-%m-%d')

	# Filter rows based on the string representation of tomorrow's date
	# Assuming the date in your CSV is in the format 'YYYY-MM-DD' and is in the first column
	tomorrow_data = price_data[price_data.iloc[:, 0].str.startswith(tomorrow)]
	two_days_ahead_data = price_data[price_data.iloc[:, 0].str.startswith(two_days_ahead)]
	three_days_ahead_data = price_data[price_data.iloc[:, 0].str.startswith(three_days_ahead)]
	st.dataframe(tomorrow_data)
	st.dataframe(two_days_ahead_data)
	st.dataframe(three_days_ahead_data)
	# Write to Excel
	excel_file_path = './Market Fundamentals/Volue_data.xlsx'  # Update with the actual path
	workbook = load_workbook(excel_file_path)
	sheet = workbook["Pret_PZU"] 

	# Confirm we have data to write
	if len(tomorrow_data) > 0:
		excel_row = 3
		for index, row in tomorrow_data.iterrows():
			if excel_row == 11 or excel_row == 24:
				excel_row += 1
			cell = f'AE{excel_row}'
			sheet[cell] = row[1]  # Assuming data to write is in the second column
			print(f"Writing {row[1]} to {cell}")  # Diagnostic print to confirm writing
			excel_row += 1
	if len(two_days_ahead_data) > 0:
		excel_row = 3
		for index, row in two_days_ahead_data.iterrows():
			if excel_row == 11 or excel_row == 24:
				excel_row += 1
			cell = f'AF{excel_row}'
			sheet[cell] = row[1]  # Assuming data to write is in the second column
			print(f"Writing {row[1]} to {cell}")  # Diagnostic print to confirm writing
			excel_row += 1
	if len(three_days_ahead_data) > 0:
		excel_row = 3
		for index, row in three_days_ahead_data.iterrows():
			if excel_row == 11 or excel_row == 24:
				excel_row += 1
			cell = f'AG{excel_row}'
			sheet[cell] = row[1]  # Assuming data to write is in the second column
			print(f"Writing {row[1]} to {cell}")  # Diagnostic print to confirm writing
			excel_row += 1
		workbook.save(filename=excel_file_path)
		print("Excel file has been updated.")
	else:
		print("No data available for tomorrow to write into the Excel file.")
	# Creating the 15 data dataframe in order to add it to the Volue dataframe
	# Load the CSV file
	data = pd_df_h.copy()
	data.reset_index(inplace=True)

	# Rename columns for clarity
	data.rename(columns={
		data.columns[0]: 'Date',
		data.columns[1]: 'Price'
	}, inplace=True)

	# Convert the date column to datetime
	data['Date'] = pd.to_datetime(data['Date'])

	# Set the date column as the index
	data.set_index('Date', inplace=True)

	# Resample the data to quarter-hourly intervals, filling forward the hydro power values
	quarterly_data = data.resample('15T').ffill()

	# Reset the index to move the date back to a column
	quarterly_data.reset_index(inplace=True)

	# Add the Interval column which represents the quarter-hourly intervals in the day (1 to 96)
	quarterly_data['Interval'] = quarterly_data['Date'].dt.hour * 4 + quarterly_data['Date'].dt.minute // 15 + 1

	# Create a column for the formatted date without time information
	quarterly_data['Formatted Date'] = quarterly_data['Date'].dt.strftime('%d.%m.%Y')

	# Reorder and rename columns to match the specified output
	quarterly_data = quarterly_data[['Formatted Date', 'Interval', 'Price']]
	quarterly_data.rename(columns={'Formatted Date': 'Date'}, inplace=True)

	return quarterly_data

#=====================================================================================Input Data=======================================================================================

# Updating the date on the Pret_PZU sheet
# Step 2: Create a new workbook (or you can open an existing one)
def updating_PZU_date():
	wb = load_workbook('./Market Fundamentals/Volue_data.xlsx')

	# Step 3: Select the active worksheet
	ws = wb["Pret_PZU"]

	# Step 4: Calculate the next day's date
	next_day = datetime.now() + timedelta(days=1)

	# Step 5: Write dates in the format "dd.mm.yyyy" from A2 to A25
	for row in range(2, 26):  # Note: range(2, 26) will loop from 2 to 25
		# Format the date as "dd.mm.yyyy"
		formatted_date = next_day.strftime("%d.%m.%Y")
		ws[f'A{row}'] = formatted_date

	# Step 6: Save the workbook
	wb.save("./Market Fundamentals/Volue_data.xlsx")

# Fetching the Consumption and Production from Transelectrica===================================================
# 1. Consumption
def process_file_consumption_transelectrica(file_path):
	# Read the file, skipping initial rows to get to the actual data
	data_cleaned = pd.read_excel(file_path, skiprows=4)
	
	# Drop the unnecessary first two columns (index and "Update day")
	data_cleaned.drop(columns=data_cleaned.columns[:2], inplace=True)
	
	# Rename the first column to 'Date' for clarity
	data_cleaned.rename(columns={data_cleaned.columns[0]: 'Date'}, inplace=True)
	
	# Determine the number of intervals (subtracting the Date column)
	num_intervals = len(data_cleaned.columns) - 1
	
	# Generate interval identifiers
	interval_identifiers = list(range(1, num_intervals + 1))
	
	# Melt the dataframe with generated interval identifiers
	data_long = pd.melt(data_cleaned, id_vars=["Date"], value_vars=data_cleaned.columns[1:], var_name="Interval", value_name="Value")
	
	# Replace the interval column with the generated interval identifiers
	data_long['Interval'] = data_long.groupby('Date').cumcount() + 1
	
	# Sort values to ensure they are ordered by Date and then by Interval
	data_long_sorted = data_long.sort_values(by=["Date", "Interval"]).reset_index(drop=True)
	
	return data_long_sorted

# 2. Production
def process_file_production_transelectrica(file_path):
	# Read the CSV, skipping initial rows to get to the actual data
	data_cleaned = pd.read_excel(file_path, skiprows=4)
	st.write(data_cleaned)
	# Drop the unnecessary first two columns (index and "Update day")
	data_cleaned.drop(columns=data_cleaned.columns[:2], inplace=True)
	
	# Rename the first column to 'Date' for clarity
	data_cleaned.rename(columns={data_cleaned.columns[0]: 'Date'}, inplace=True)
	
	# Determine the number of intervals (subtracting the Date column)
	num_intervals = len(data_cleaned.columns) - 1
	
	# Generate interval identifiers
	interval_identifiers = list(range(1, num_intervals + 1))
	
	# Melt the dataframe with generated interval identifiers
	data_long = pd.melt(data_cleaned, id_vars=["Date"], value_vars=data_cleaned.columns[1:], var_name="Interval", value_name="Value")
	
	# Replace the interval column with the generated interval identifiers
	data_long['Interval'] = data_long.groupby('Date').cumcount() + 1
	
	# Sort values to ensure they are ordered by Date and then by Interval
	data_long_sorted = data_long.sort_values(by=["Date", "Interval"]).reset_index(drop=True)
	st.write(data_long_sorted)
	return data_long_sorted

def zip_files(folder_path, zip_name):
	zip_path = os.path.join(folder_path, zip_name)
	with zipfile.ZipFile(zip_path, 'w') as zipf:
		for root, _, files in os.walk(folder_path):
			for file in files:
				if file != zip_name:  # Avoid zipping the zip file itself
					file_path = os.path.join(root, file)
					arcname = os.path.relpath(file_path, folder_path)  # Relative path within the zip file
					zipf.write(file_path, arcname)
#====================================================================================ENTSOE newAPI data=======================================================================================
api_key_entsoe = os.getenv("api_key_entsoe")
if api_key_entsoe is None:
	api_key_entsoe = "api_key_entsoe = ad0626ec-bccb-4d39-9b90-30b2e50c18e0"
client = EntsoePandasClient(api_key=api_key_entsoe)


# ===========================Imbalance Volumes and Prices================================================================

# Function to fetch imbalance volumes for the intraday scenario
def fetching_imbalance_volumes(start_date, end_date):
	# Setting up the start and end dates (today and tomorrow)
	today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-1)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Format the start and end dates to match the API requirements (yyyymmddhhmm)
	period_start = start_cet.strftime('%Y%m%d%H%M')
	period_end = end_cet.strftime('%Y%m%d%H%M')

	# Define the endpoint and parameters for fetching imbalance volumes via ENTSO-E API
	url = "https://web-api.tp.entsoe.eu/api"

	# Parameters for the API request
	params = {
		"securityToken": api_key_entsoe,
		"documentType": "A86",  # Document type for total imbalance volumes
		"controlArea_Domain": "10YRO-TEL------P",  # Romania's area EIC code
		"periodStart": period_start,  # Start date in yyyymmddhhmm format
		"periodEnd": period_end,  # End date in yyyymmddhhmm format
	}

	# Headers for the API request
	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	# Make the request to the ENTSO-E API
	try:
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()  # Raise an error for bad status codes

		# Save the response content to a file
		with open("./data_fetching/Entsoe/entsoe_response.zip", "wb") as f:
			f.write(response.content)

		print("File saved as entsoe_response.zip")

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")

# Define namespaces (try without namespaces if it doesn't match)
namespaces = {'ns': 'urn:iec62325.351:tc57wg16:451-6:balancingdocument:3:0'}

# Fetch and process imbalance volumes data from the existing zip file
def process_imbalance_volumes(start_date, end_date, zip_filepath='./data_fetching/Entsoe/entsoe_response.zip'):
	"""
	Process imbalance volume data from the given zip file and return a DataFrame
	for the specified date range (start_date to end_date).
	"""
	# Extract the .zip file to access the XML content
	with zipfile.ZipFile(zip_filepath, 'r') as zip_ref:
		extracted_files = zip_ref.namelist()
		xml_filename = extracted_files[0]  # Assuming there's only one file in the .zip
		zip_ref.extract(xml_filename)

	# Parse the XML content
	tree = ET.parse(xml_filename)
	root = tree.getroot()

	# Extract the namespace dynamically
	ns_match = root.tag[root.tag.find("{"):root.tag.find("}")+1]
	namespaces = {'ns': ns_match.strip("{}")} if ns_match else {}

	# Initialize lists to store parsed data
	timestamps_utc = []
	volumes = []

	# Iterate over all TimeSeries elements
	for timeseries in root.findall('ns:TimeSeries', namespaces):
		flow_direction_tag = timeseries.find('ns:flowDirection.direction', namespaces)
		if flow_direction_tag is not None:
			flow_direction = flow_direction_tag.text
		else:
			continue  # Skip this timeseries if no flow direction is found

		# Determine if the series is a deficit or an excedent based on flow direction
		direction_sign = -1 if flow_direction == 'A02' else (1 if flow_direction == 'A01' else None)
		if direction_sign is None:
			continue

		for period in timeseries.findall('ns:Period', namespaces):
			start = period.find('ns:timeInterval/ns:start', namespaces)
			if start is None:
				continue

			start_time_utc = datetime.strptime(start.text, '%Y-%m-%dT%H:%MZ')  # Start time is in UTC

			for point in period.findall('ns:Point', namespaces):
				position_tag = point.find('ns:position', namespaces)
				quantity_tag = point.find('ns:quantity', namespaces)

				if position_tag is None or quantity_tag is None:
					continue

				try:
					position = int(position_tag.text)
					quantity = float(quantity_tag.text) * direction_sign
				except ValueError as e:
					print(f"Error converting position or quantity: {e}, skipping.")
					continue

				point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))
				timestamps_utc.append(point_time_utc)
				volumes.append(quantity)

	# Create DataFrame from parsed data
	df_imbalance = pd.DataFrame({
		'Timestamp': timestamps_utc,
		'Imbalance Volume': volumes
	})

	if df_imbalance.empty:
		print("DataFrame is empty after parsing XML. No data to process.")
		return df_imbalance

	# Convert UTC timestamps to CET
	df_imbalance['Timestamp'] = pd.to_datetime(df_imbalance['Timestamp']).dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')

	# Filter by user-provided date range
	if start_date and end_date:
		start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
		end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')
		df_imbalance = df_imbalance[(df_imbalance['Timestamp'] >= start_cet) & (df_imbalance['Timestamp'] <= end_cet)]

	# Ensure all timestamps are covered for the range
	full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')
	df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')

	# Merge actual data with the full index, keeping original values intact
	df_imbalance = df_full.merge(df_imbalance.set_index('Timestamp'), how='left', left_index=True, right_index=True)
	df_imbalance.fillna(0, inplace=True)  # Fill missing values with 0

	# Reset index to return a clean DataFrame
	df_imbalance.reset_index(inplace=True)

	# Clean up: remove the extracted file after processing
	os.remove(xml_filename)

	return df_imbalance

def fetch_imbalance_prices(start_date, end_date):
	# Setting up the start and end dates (today and tomorrow)
	today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-1)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Define the endpoint and parameters for fetching imbalance prices via ENTSO-E API
	url = "https://web-api.tp.entsoe.eu/api"

	# Parameters for the API request
	params = {
		"securityToken": api_key_entsoe,
		"documentType": "A85",  # Document type for imbalance prices
		"controlArea_Domain": "10YRO-TEL------P",  # Romania's area EIC code
		"periodStart": start_cet.strftime('%Y%m%d%H%M'),  # Start date in yyyymmddhhmm format
		"periodEnd": end_cet.strftime('%Y%m%d%H%M'),  # End date in yyyymmddhhmm format
	}

	# Headers for the API request
	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	# Make the request to the ENTSO-E API
	try:
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()  # Raise an error for bad status codes

		# Save the response content to a file
		with open("./data_fetching/Entsoe/imbalance_prices_response.zip", "wb") as f:
			f.write(response.content)

		print("File saved as imbalance_prices_response.zip")

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")

def process_imbalance_prices(start_date, end_date, zip_filepath='./data_fetching/Entsoe/imbalance_prices_response.zip'):
	"""
	Process imbalance price data from the given zip file for the specified date range.
	"""
	# Extract the .zip file to access the XML content
	with zipfile.ZipFile(zip_filepath, 'r') as zip_ref:
		extracted_files = zip_ref.namelist()
		xml_filename = extracted_files[0]  # Assuming there's only one file in the .zip
		zip_ref.extract(xml_filename)

	# Parse the XML content
	tree = ET.parse(xml_filename)
	root = tree.getroot()

	# Extract the namespace dynamically if needed
	ns_match = root.tag[root.tag.find("{"):root.tag.find("}")+1]
	namespaces = {'ns': ns_match.strip("{}")} if ns_match else {}
	print(f"Using namespace: {namespaces}")

	# Dictionary to store data by timestamp
	data_dict = {}

	# Iterate over all TimeSeries elements
	for timeseries in root.findall('ns:TimeSeries', namespaces):
		# Extract Business Type
		business_type_tag = timeseries.find('ns:businessType', namespaces)

		if business_type_tag is None or business_type_tag.text != "A19":
			continue  # Skip if business type is not "A19" (imbalance prices)

		# Iterate over all Period elements within each TimeSeries
		for period in timeseries.findall('ns:Period', namespaces):
			start = period.find('ns:timeInterval/ns:start', namespaces)

			if start is None:
				continue

			start_time_utc = datetime.strptime(start.text, '%Y-%m-%dT%H:%MZ')  # Start time is in UTC

			# Iterate over all Point elements within each Period
			for point in period.findall('ns:Point', namespaces):
				position_tag = point.find('ns:position', namespaces)
				price_tag = point.find('ns:imbalance_Price.amount', namespaces)
				category_tag = point.find('ns:imbalance_Price.category', namespaces)

				# Skip if required fields are missing
				if position_tag is None or price_tag is None or category_tag is None:
					continue

				try:
					position = int(position_tag.text)
					price = float(price_tag.text)
					category = category_tag.text
				except ValueError:
					continue  # Skip if conversion fails

				# Calculate timestamp for the point based on position in UTC
				point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))

				# Initialize a dictionary entry for the timestamp if not already present
				if point_time_utc not in data_dict:
					data_dict[point_time_utc] = {'Excedent Price': None, 'Deficit Price': None}

				# Assign the price to the appropriate category
				if category == 'A04':  # Excedent Price
					data_dict[point_time_utc]['Excedent Price'] = price
				elif category == 'A05':  # Deficit Price
					data_dict[point_time_utc]['Deficit Price'] = price

	# Convert dictionary to DataFrame
	df_prices = pd.DataFrame.from_dict(data_dict, orient='index').reset_index()
	df_prices.rename(columns={'index': 'Timestamp_UTC'}, inplace=True)

	# Drop duplicate timestamps to avoid issues with reindexing
	df_prices['Timestamp_UTC'] = pd.to_datetime(df_prices['Timestamp_UTC'])
	df_prices = df_prices.drop_duplicates(subset='Timestamp_UTC', keep='first')

	# Check if DataFrame is empty before further processing
	if df_prices.empty:
		print("DataFrame is empty after parsing XML. No data to process.")
		return df_prices

	# Convert the UTC timestamps to CET (Europe/Berlin, which handles DST)
	df_prices['Timestamp_CET'] = df_prices['Timestamp_UTC'].dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')

	# Remove the UTC column and rename CET to Timestamp for simplicity
	df_prices.drop(columns=['Timestamp_UTC'], inplace=True)
	df_prices.rename(columns={'Timestamp_CET': 'Timestamp'}, inplace=True)

	# Filter by user-provided date range
	start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
	end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')
	df_prices = df_prices[(df_prices['Timestamp'] >= start_cet) & (df_prices['Timestamp'] <= end_cet)]

	# Ensure all timestamps are covered for the range (15-min intervals)
	full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')
	df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')

	# Merge actual data with the full index, keeping original values intact
	df_prices = df_full.merge(df_prices.set_index('Timestamp'), how='left', left_index=True, right_index=True)
	df_prices.fillna(0, inplace=True)  # Fill missing values with 0

	# Reset index to return a clean DataFrame
	df_prices.reset_index(inplace=True)

	# Clean up: remove the extracted file after processing
	os.remove(xml_filename)

	return df_prices

def create_combined_imbalance_dataframe(df_prices, df_volumes):
	"""
	This function combines the imbalance prices and volumes into a single DataFrame.
	
	Parameters:
	df_prices (DataFrame): DataFrame containing timestamps, Excedent Price, and Deficit Price.
	df_volumes (DataFrame): DataFrame containing timestamps and Imbalance Volume.

	Returns:
	DataFrame: Combined DataFrame with Timestamp, Excedent Price, Deficit Price, and Imbalance Volume.
	"""
	# Merge prices and volumes on the Timestamp column using an outer join to ensure all data is included.
	df_combined = pd.merge(df_prices, df_volumes, on='Timestamp', how='outer')

	# Sort the combined DataFrame by Timestamp to keep it in chronological order.
	df_combined = df_combined.sort_values(by='Timestamp')

	# Fill any missing values with 0.0 to ensure consistency in analysis.
	df_combined.fillna(0.0, inplace=True)

	# Saving the data to the folder
	df_combined_excel = df_combined.copy()
	df_combined_excel["Timestamp"] = df_combined_excel["Timestamp"].dt.tz_localize(None)
	# Add a "Quarter" column based on the Timestamp column
	df_combined_excel['Quarter'] = (df_combined_excel.index % 96) + 1
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df_combined_excel['Lookup'] = df_combined_excel["Timestamp"].dt.strftime('%d.%m.%Y') + df_combined_excel["Quarter"].astype(str)
	df_combined_excel.to_excel("./data_fetching/Entsoe/Imbalance_Historical.xlsx", index=False)

	return df_combined

# ===========================Actual and Notified Wind Production================================================================

def fetch_process_wind_notified(start_date, end_date):
	"""
	Fetch and process wind notified production data from ENTSO-E API for a given date range.
	"""
	# Convert start_date and end_date to CET timestamps
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-2)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Format the start and end dates to match the API requirements (yyyymmddhhmm)
	period_start = start_cet.strftime('%Y%m%d%H%M')
	period_end = end_cet.strftime('%Y%m%d%H%M')

	url = "https://web-api.tp.entsoe.eu/api"

	params = {
		"securityToken": api_key_entsoe,
		"documentType": "A69",  # Document type for wind and solar generation forecast
		"processType": "A18",   # A18 represents Intraday forecast
		"in_Domain": "10YRO-TEL------P",  # EIC code for the desired country/region
		"periodStart": period_start,
		"periodEnd": period_end,
		"PsrType": "B19"  # B19 corresponds to Wind Onshore
	}

	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	try:
		# Make the API request
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()

		# Parse the XML response
		root = ET.fromstring(response.content)

		# Extract namespace dynamically
		ns_match = root.tag[root.tag.find("{"):root.tag.find("}")+1]
		namespaces = {'ns': ns_match.strip("{}")} if ns_match else {}

		# Initialize lists to store parsed data
		timestamps_utc = []
		quantities = []

		# Iterate over TimeSeries to extract the points
		for timeseries in root.findall('ns:TimeSeries', namespaces):
			for period in timeseries.findall('ns:Period', namespaces):
				start = period.find('ns:timeInterval/ns:start', namespaces)

				if start is None:
					continue

				# Extract the start time and assume a 15-min resolution
				start_time_utc = datetime.strptime(start.text, '%Y-%m-%dT%H:%MZ')  # Start time is in UTC

				# Iterate over all Point elements within each Period
				for point in period.findall('ns:Point', namespaces):
					position_tag = point.find('ns:position', namespaces)
					quantity_tag = point.find('ns:quantity', namespaces)

					if position_tag is None or quantity_tag is None:
						continue

					try:
						position = int(position_tag.text)
						quantity = float(quantity_tag.text)
					except ValueError:
						continue

					# Calculate timestamp for the point based on position in UTC
					point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))
					timestamps_utc.append(point_time_utc)
					quantities.append(quantity)

		# Create DataFrame from the parsed data
		df_notified_production = pd.DataFrame({
			'Timestamp_UTC': timestamps_utc,
			'Notified Production (MW)': quantities
		})

		# Convert UTC to CET
		df_notified_production['Timestamp_UTC'] = pd.to_datetime(df_notified_production['Timestamp_UTC'])
		df_notified_production['Timestamp_CET'] = df_notified_production['Timestamp_UTC'].dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')
		df_notified_production.drop(columns=['Timestamp_UTC'], inplace=True)
		df_notified_production.rename(columns={'Timestamp_CET': 'Timestamp'}, inplace=True)

		# Filter by user-provided date range
		start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
		end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')

		# Create a complete index for 15-minute intervals
		full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')

		# Merge actual data with the full index to ensure all intervals are covered
		df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')
		df_notified_production = df_full.merge(
			df_notified_production.set_index('Timestamp'),
			how='left',
			left_index=True,
			right_index=True
		)
		df_notified_production.fillna(0, inplace=True)  # Replace missing values with 0
		df_notified_production.reset_index(inplace=True)

		# Write the DataFrame to Excel
		df_notified_production_excel = df_notified_production.copy()
		df_notified_production_excel['Timestamp'] = df_notified_production_excel['Timestamp'].dt.tz_localize(None)
		df_notified_production_excel.to_excel("./data_fetching/Entsoe/Notified_Production_Wind.xlsx", index=False)

		return df_notified_production

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")
		return pd.DataFrame()

def fetch_process_wind_actual_production(start_date, end_date):
	"""
	Fetch and process actual wind production data from ENTSO-E API for a given date range.
	"""
	# Convert start_date and end_date to CET timestamps
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-2)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Format the start and end dates to match the API requirements (yyyymmddhhmm)
	period_start = start_cet.strftime('%Y%m%d%H%M')
	period_end = end_cet.strftime('%Y%m%d%H%M')

	url = "https://web-api.tp.entsoe.eu/api"

	params = {
		"securityToken": api_key_entsoe,  # Replace with your actual API key
		"documentType": "A75",  # Document type for actual generation per type (all production types)
		"processType": "A16",   # A16 represents Realized generation
		"in_Domain": "10YRO-TEL------P",  # EIC code for the desired country/region (Romania in this case)
		"periodStart": period_start,
		"periodEnd": period_end,
		"PsrType": "B19"  # B19 corresponds to Wind Onshore
	}

	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	try:
		# Make the API request
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()

		# Parse the XML response
		root = ET.fromstring(response.content)

		# Extract namespace dynamically
		ns_match = root.tag[root.tag.find("{"):root.tag.find("}")+1]
		namespaces = {'ns': ns_match.strip("{}")} if ns_match else {}

		# Initialize lists to store parsed data
		timestamps_utc = []
		quantities = []

		# Iterate over TimeSeries to extract the points
		for timeseries in root.findall('ns:TimeSeries', namespaces):
			for period in timeseries.findall('ns:Period', namespaces):
				start = period.find('ns:timeInterval/ns:start', namespaces)
				resolution = period.find('ns:resolution', namespaces)

				if start is None or resolution is None:
					continue

				# Extract the start time and resolution
				start_time_utc = datetime.strptime(start.text, '%Y-%m-%dT%H:%MZ')  # Start time is in UTC

				# Iterate over all Point elements within each Period
				for point in period.findall('ns:Point', namespaces):
					position_tag = point.find('ns:position', namespaces)
					quantity_tag = point.find('ns:quantity', namespaces)

					if position_tag is None or quantity_tag is None:
						continue

					try:
						position = int(position_tag.text)
						quantity = float(quantity_tag.text)
					except ValueError:
						continue

					# Calculate timestamp for the point based on position in UTC
					point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))
					timestamps_utc.append(point_time_utc)
					quantities.append(quantity)

		# Create DataFrame from the parsed data
		df_actual_production = pd.DataFrame({
			'Timestamp_UTC': timestamps_utc,
			'Actual Production (MW)': quantities
		})

		# Convert UTC to CET
		df_actual_production['Timestamp_UTC'] = pd.to_datetime(df_actual_production['Timestamp_UTC'])
		df_actual_production['Timestamp_CET'] = df_actual_production['Timestamp_UTC'].dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')
		df_actual_production.drop(columns=['Timestamp_UTC'], inplace=True)
		df_actual_production.rename(columns={'Timestamp_CET': 'Timestamp'}, inplace=True)

		# Filter by user-provided date range
		start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
		end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')

		# Create a complete index for 15-minute intervals
		full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')

		# Merge actual data with the full index to ensure all intervals are covered
		df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')
		df_actual_production = df_full.merge(
			df_actual_production.set_index('Timestamp'),
			how='left',
			left_index=True,
			right_index=True
		)
		df_actual_production.fillna(0, inplace=True)  # Replace missing values with 0
		df_actual_production.reset_index(inplace=True)

		# Write the DataFrame to Excel
		df_actual_production_excel = df_actual_production.copy()
		df_actual_production_excel['Timestamp'] = df_actual_production_excel['Timestamp'].dt.tz_localize(None)
		df_actual_production_excel.to_excel("./data_fetching/Entsoe/Actual_Production_Wind.xlsx", index=False)

		return df_actual_production

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")
		return pd.DataFrame()

def combine_wind_production_data(df_notified, df_actual):
	# Step 1: Sort and align the wind notified and actual production DataFrames
	df_notified = df_notified.sort_values(by='Timestamp').set_index('Timestamp')
	df_actual = df_actual.sort_values(by='Timestamp').set_index('Timestamp')
	
	# Step 2: Concatenate the notified and actual data based on Timestamp
	df_combined = pd.concat([df_notified, df_actual], axis=1)

	# Step 3: Reset index to make Timestamp a column again
	df_combined.reset_index(inplace=True)
	
	# Step 4: Rename the columns for better readability
	df_combined.columns = ['Timestamp', 'Notified Production (MW)', 'Actual Production (MW)']

	# Step 5: Add the volue forecast to the combined dataframe by aligning timestamps
	# df_forecast = df_forecast.sort_values(by='Timestamp').set_index('Timestamp')

	# Aligning the forecast data based on the timestamps in the combined dataframe
	# df_combined = df_combined.set_index('Timestamp')
	# df_combined['Volue Forecast (MW)'] = df_forecast['Volue Forecast (MW)']
	# df_combined.reset_index(inplace=True)

	# # Step 6: Fill any missing values if required (e.g., with 0 or 'NaN')
	# df_combined.fillna(method='ffill', inplace=True)  # Forward fill any missing values if necessary

	# Saving the data to the folder
	df_combined_excel = df_combined.copy()
	df_combined_excel["Timestamp"] = df_combined_excel["Timestamp"].dt.tz_localize(None)
	# Add a "Quarter" column based on the Timestamp column
	df_combined_excel['Quarter'] = (df_combined_excel.index % 96) + 1
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df_combined_excel['Lookup'] = df_combined_excel["Timestamp"].dt.strftime('%d.%m.%Y') + df_combined_excel["Quarter"].astype(str)
	df_combined_excel.to_excel("./data_fetching/Entsoe/Wind_Production_Historical.xlsx", index=False)

	return df_combined

# ===========================Actual and Notified Solar Production================================================================

def fetch_process_solar_notified(start_date, end_date):
	"""
	Fetch and process solar notified production data from ENTSO-E API for a given date range.
	"""
	# Convert start_date and end_date to CET timestamps
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-2)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Format the start and end dates to match the API requirements (yyyymmddhhmm)
	period_start = start_cet.strftime('%Y%m%d%H%M')
	period_end = end_cet.strftime('%Y%m%d%H%M')

	url = "https://web-api.tp.entsoe.eu/api"

	params = {
		"securityToken": api_key_entsoe,
		"documentType": "A69",  # Document type for wind and solar generation forecast
		"processType": "A18",   # A18 represents Intraday forecast
		"in_Domain": "10YRO-TEL------P",  # EIC code for the desired country/region
		"periodStart": period_start,
		"periodEnd": period_end,
		"PsrType": "B16"  # B16 corresponds to Solar
	}

	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	try:
		# Make the API request
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()

		# Parse the XML response
		root = ET.fromstring(response.content)

		# Extract namespace dynamically
		ns_match = root.tag[root.tag.find("{"):root.tag.find("}")+1]
		namespaces = {'ns': ns_match.strip("{}")} if ns_match else {}

		# Initialize lists to store parsed data
		timestamps_utc = []
		quantities = []

		# Iterate over TimeSeries to extract the points
		for timeseries in root.findall('ns:TimeSeries', namespaces):
			for period in timeseries.findall('ns:Period', namespaces):
				start = period.find('ns:timeInterval/ns:start', namespaces)

				if start is None:
					continue

				# Extract the start time and assume a 15-min resolution
				start_time_utc = datetime.strptime(start.text, '%Y-%m-%dT%H:%MZ')  # Start time is in UTC

				# Iterate over all Point elements within each Period
				for point in period.findall('ns:Point', namespaces):
					position_tag = point.find('ns:position', namespaces)
					quantity_tag = point.find('ns:quantity', namespaces)

					if position_tag is None or quantity_tag is None:
						continue

					try:
						position = int(position_tag.text)
						quantity = float(quantity_tag.text)
					except ValueError:
						continue

					# Calculate timestamp for the point based on position in UTC
					point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))
					timestamps_utc.append(point_time_utc)
					quantities.append(quantity)

		# Create DataFrame from the parsed data
		df_notified_production = pd.DataFrame({
			'Timestamp_UTC': timestamps_utc,
			'Notified Production (MW)': quantities
		})

		# Convert UTC to CET
		df_notified_production['Timestamp_UTC'] = pd.to_datetime(df_notified_production['Timestamp_UTC'])
		df_notified_production['Timestamp_CET'] = df_notified_production['Timestamp_UTC'].dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')
		df_notified_production.drop(columns=['Timestamp_UTC'], inplace=True)
		df_notified_production.rename(columns={'Timestamp_CET': 'Timestamp'}, inplace=True)

		# Filter by user-provided date range
		start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
		end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')

		# Create a complete index for 15-minute intervals
		full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')

		# Merge actual data with the full index to ensure all intervals are covered
		df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')
		df_notified_production = df_full.merge(
			df_notified_production.set_index('Timestamp'),
			how='left',
			left_index=True,
			right_index=True
		)
		df_notified_production.fillna(0, inplace=True)  # Replace missing values with 0
		df_notified_production.reset_index(inplace=True)

		# Write the DataFrame to Excel
		df_notified_production_excel = df_notified_production.copy()
		df_notified_production_excel['Timestamp'] = df_notified_production_excel['Timestamp'].dt.tz_localize(None)
		df_notified_production_excel.to_excel("./data_fetching/Entsoe/Notified_Production_Solar.xlsx", index=False)

		return df_notified_production

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")
		return pd.DataFrame()

def fetch_process_solar_actual_production(start_date, end_date):
	"""
	Fetch and process solar actual production data from ENTSO-E API for a given date range.
	"""
	# Convert start_date and end_date to CET timestamps
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-2)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Format the start and end dates to match the API requirements (yyyymmddhhmm)
	period_start = start_cet.strftime('%Y%m%d%H%M')
	period_end = end_cet.strftime('%Y%m%d%H%M')

	url = "https://web-api.tp.entsoe.eu/api"

	params = {
		"securityToken": api_key_entsoe,  # Replace with your actual API key
		"documentType": "A75",  # Document type for actual generation per type (all production types)
		"processType": "A16",   # A16 represents Realized generation
		"in_Domain": "10YRO-TEL------P",  # EIC code for the desired country/region (Romania in this case)
		"periodStart": period_start,
		"periodEnd": period_end,
		"PsrType": "B16"  # B16 corresponds to Solar
	}

	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	try:
		# Make the API request
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()

		# Parse the XML response
		root = ET.fromstring(response.content)

		# Extract namespace dynamically
		ns_match = root.tag[root.tag.find("{"):root.tag.find("}")+1]
		namespaces = {'ns': ns_match.strip("{}")} if ns_match else {}

		# Initialize lists to store parsed data
		timestamps_utc = []
		quantities = []

		# Iterate over TimeSeries to extract the points
		for timeseries in root.findall('ns:TimeSeries', namespaces):
			for period in timeseries.findall('ns:Period', namespaces):
				start = period.find('ns:timeInterval/ns:start', namespaces)
				resolution = period.find('ns:resolution', namespaces)

				if start is None or resolution is None:
					continue

				# Extract the start time and resolution
				start_time_utc = datetime.strptime(start.text, '%Y-%m-%dT%H:%MZ')  # Start time is in UTC

				# Iterate over all Point elements within each Period
				for point in period.findall('ns:Point', namespaces):
					position_tag = point.find('ns:position', namespaces)
					quantity_tag = point.find('ns:quantity', namespaces)

					if position_tag is None or quantity_tag is None:
						continue

					try:
						position = int(position_tag.text)
						quantity = float(quantity_tag.text)
					except ValueError:
						continue

					# Calculate timestamp for the point based on position in UTC
					point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))
					timestamps_utc.append(point_time_utc)
					quantities.append(quantity)

		# Create DataFrame from the parsed data
		df_actual_production = pd.DataFrame({
			'Timestamp_UTC': timestamps_utc,
			'Actual Production (MW)': quantities
		})

		# Convert UTC to CET
		df_actual_production['Timestamp_UTC'] = pd.to_datetime(df_actual_production['Timestamp_UTC'])
		df_actual_production['Timestamp_CET'] = df_actual_production['Timestamp_UTC'].dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')
		df_actual_production.drop(columns=['Timestamp_UTC'], inplace=True)
		df_actual_production.rename(columns={'Timestamp_CET': 'Timestamp'}, inplace=True)

		# Filter by user-provided date range
		start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
		end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')

		# Create a complete index for 15-minute intervals
		full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')

		# Merge actual data with the full index to ensure all intervals are covered
		df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')
		df_actual_production = df_full.merge(
			df_actual_production.set_index('Timestamp'),
			how='left',
			left_index=True,
			right_index=True
		)
		df_actual_production.fillna(0, inplace=True)  # Replace missing values with 0
		df_actual_production.reset_index(inplace=True)

		# Write the DataFrame to Excel
		df_actual_production_excel = df_actual_production.copy()
		df_actual_production_excel['Timestamp'] = df_actual_production_excel['Timestamp'].dt.tz_localize(None)
		df_actual_production_excel.to_excel("./data_fetching/Entsoe/Actual_Production_Solar.xlsx", index=False)

		return df_actual_production

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")
		return pd.DataFrame()

def combine_solar_production_data(df_notified, df_actual):
	# Step 1: Sort and align the wind notified and actual production DataFrames
	df_notified = df_notified.sort_values(by='Timestamp').set_index('Timestamp')
	df_actual = df_actual.sort_values(by='Timestamp').set_index('Timestamp')
	
	# Step 2: Concatenate the notified and actual data based on Timestamp
	df_combined = pd.concat([df_notified, df_actual], axis=1)

	# Step 3: Reset index to make Timestamp a column again
	df_combined.reset_index(inplace=True)
	
	# Step 4: Rename the columns for better readability
	df_combined.columns = ['Timestamp', 'Notified Production (MW)', 'Actual Production (MW)']

	# Step 5: Add the volue forecast to the combined dataframe by aligning timestamps
	# df_forecast = df_forecast.sort_values(by='Timestamp').set_index('Timestamp')

	# Aligning the forecast data based on the timestamps in the combined dataframe
	# df_combined = df_combined.set_index('Timestamp')
	# df_combined['Volue Forecast (MW)'] = df_forecast['Volue Forecast (MW)']
	# df_combined.reset_index(inplace=True)

	# Step 6: Fill any missing values if required (e.g., with 0 or 'NaN')
	# df_combined.fillna(method='ffill', inplace=True)  # Forward fill any missing values if necessary

	# Saving the data to the folder
	df_combined_excel = df_combined.copy()
	# Remove timezone information from the 'Timestamp' column
	df_combined_excel["Timestamp"] = df_combined_excel["Timestamp"].dt.tz_localize(None)
	# Add a "Quarter" column based on the Timestamp column
	df_combined_excel['Quarter'] = (df_combined_excel.index % 96) + 1
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df_combined_excel['Lookup'] = df_combined_excel["Timestamp"].dt.strftime('%d.%m.%Y') + df_combined_excel["Quarter"].astype(str)
	df_combined_excel.to_excel("./data_fetching/Entsoe/Solar_Production_Historical.xlsx", index=False)

	return df_combined

# ===========================Actual and Notified Hydro Production================================================================

def fetch_process_hydro_water_reservoir_actual_production(start_date, end_date):
	"""
	Fetch and process hydro water reservoir actual production data from ENTSO-E API for a given date range.
	"""
	# Convert start_date and end_date to CET timestamps
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-2)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Format the start and end dates to match the API requirements (yyyymmddhhmm)
	period_start = start_cet.strftime('%Y%m%d%H%M')
	period_end = end_cet.strftime('%Y%m%d%H%M')

	url = "https://web-api.tp.entsoe.eu/api"

	params = {
		"securityToken": api_key_entsoe,  # Replace with your actual API key
		"documentType": "A75",  # Document type for actual generation per type
		"processType": "A16",   # A16 represents Realized generation
		"in_Domain": "10YRO-TEL------P",  # EIC code for Romania
		"periodStart": period_start,
		"periodEnd": period_end,
		"PsrType": "B12"  # B12 corresponds to Hydro Water Reservoir
	}

	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	try:
		# Make the API request
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()

		# Parse the XML response
		root = ET.fromstring(response.content)

		# Extract namespace dynamically
		ns_match = root.tag[root.tag.find("{"):root.tag.find("}")+1]
		namespaces = {'ns': ns_match.strip("{}")} if ns_match else {}

		# Initialize lists to store parsed data
		timestamps_utc = []
		quantities = []

		# Iterate over TimeSeries to extract the points
		for timeseries in root.findall('ns:TimeSeries', namespaces):
			for period in timeseries.findall('ns:Period', namespaces):
				start = period.find('ns:timeInterval/ns:start', namespaces)
				resolution = period.find('ns:resolution', namespaces)

				if start is None or resolution is None:
					continue

				# Extract the start time and resolution
				start_time_utc = datetime.strptime(start.text, '%Y-%m-%dT%H:%MZ')  # Start time is in UTC

				# Iterate over all Point elements within each Period
				for point in period.findall('ns:Point', namespaces):
					position_tag = point.find('ns:position', namespaces)
					quantity_tag = point.find('ns:quantity', namespaces)

					if position_tag is None or quantity_tag is None:
						continue

					try:
						position = int(position_tag.text)
						quantity = float(quantity_tag.text)
					except ValueError:
						continue

					# Calculate timestamp for the point based on position in UTC
					point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))
					timestamps_utc.append(point_time_utc)
					quantities.append(quantity)

		# Create DataFrame from the parsed data
		df_actual_production = pd.DataFrame({
			'Timestamp_UTC': timestamps_utc,
			'Actual Production (MW)': quantities
		})

		# Convert UTC to CET
		df_actual_production['Timestamp_UTC'] = pd.to_datetime(df_actual_production['Timestamp_UTC'])
		df_actual_production['Timestamp_CET'] = df_actual_production['Timestamp_UTC'].dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')
		df_actual_production.drop(columns=['Timestamp_UTC'], inplace=True)
		df_actual_production.rename(columns={'Timestamp_CET': 'Timestamp'}, inplace=True)

		# Filter by user-provided date range
		start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
		end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')

		# Create a complete index for 15-minute intervals
		full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')

		# Merge actual data with the full index to ensure all intervals are covered
		df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')
		df_actual_production = df_full.merge(
			df_actual_production.set_index('Timestamp'),
			how='left',
			left_index=True,
			right_index=True
		)
		df_actual_production.fillna(0, inplace=True)  # Replace missing values with 0
		df_actual_production.reset_index(inplace=True)

		# Write the DataFrame to Excel
		df_actual_production_excel = df_actual_production.copy()
		df_actual_production_excel['Timestamp'] = df_actual_production_excel['Timestamp'].dt.tz_localize(None)
		df_actual_production_excel.to_excel("./data_fetching/Entsoe/Actual_Production_Hydro_Water_Reservoir.xlsx", index=False)

		return df_actual_production

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")
		return pd.DataFrame()


def fetch_process_hydro_river_actual_production(start_date, end_date):
	"""
	Fetch and process hydro river actual production data from ENTSO-E API for a given date range.
	"""
	# Convert start_date and end_date to CET timestamps
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-2)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Format the start and end dates to match the API requirements (yyyymmddhhmm)
	period_start = start_cet.strftime('%Y%m%d%H%M')
	period_end = end_cet.strftime('%Y%m%d%H%M')

	url = "https://web-api.tp.entsoe.eu/api"

	params = {
		"securityToken": api_key_entsoe,  # Replace with your actual API key
		"documentType": "A75",  # Document type for actual generation per type
		"processType": "A16",   # A16 represents Realized generation
		"in_Domain": "10YRO-TEL------P",  # EIC code for Romania
		"periodStart": period_start,
		"periodEnd": period_end,
		"PsrType": "B11"  # B11 corresponds to Hydro River
	}

	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	try:
		# Make the API request
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()

		# Parse the XML response
		root = ET.fromstring(response.content)

		# Extract namespace dynamically
		ns_match = root.tag[root.tag.find("{"):root.tag.find("}")+1]
		namespaces = {'ns': ns_match.strip("{}")} if ns_match else {}

		# Initialize lists to store parsed data
		timestamps_utc = []
		quantities = []

		# Iterate over TimeSeries to extract the points
		for timeseries in root.findall('ns:TimeSeries', namespaces):
			for period in timeseries.findall('ns:Period', namespaces):
				start = period.find('ns:timeInterval/ns:start', namespaces)
				resolution = period.find('ns:resolution', namespaces)

				if start is None or resolution is None:
					continue

				# Extract the start time and resolution
				start_time_utc = datetime.strptime(start.text, '%Y-%m-%dT%H:%MZ')  # Start time is in UTC

				# Iterate over all Point elements within each Period
				for point in period.findall('ns:Point', namespaces):
					position_tag = point.find('ns:position', namespaces)
					quantity_tag = point.find('ns:quantity', namespaces)

					if position_tag is None or quantity_tag is None:
						continue

					try:
						position = int(position_tag.text)
						quantity = float(quantity_tag.text)
					except ValueError:
						continue

					# Calculate timestamp for the point based on position in UTC
					point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))
					timestamps_utc.append(point_time_utc)
					quantities.append(quantity)

		# Create DataFrame from the parsed data
		df_actual_production = pd.DataFrame({
			'Timestamp_UTC': timestamps_utc,
			'Actual Production (MW)': quantities
		})

		# Convert UTC to CET
		df_actual_production['Timestamp_UTC'] = pd.to_datetime(df_actual_production['Timestamp_UTC'])
		df_actual_production['Timestamp_CET'] = df_actual_production['Timestamp_UTC'].dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')
		df_actual_production.drop(columns=['Timestamp_UTC'], inplace=True)
		df_actual_production.rename(columns={'Timestamp_CET': 'Timestamp'}, inplace=True)

		# Filter by user-provided date range
		start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
		end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')

		# Create a complete index for 15-minute intervals
		full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')

		# Merge actual data with the full index to ensure all intervals are covered
		df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')
		df_actual_production = df_full.merge(
			df_actual_production.set_index('Timestamp'),
			how='left',
			left_index=True,
			right_index=True
		)
		df_actual_production.fillna(0, inplace=True)  # Replace missing values with 0
		df_actual_production.reset_index(inplace=True)

		# Write the DataFrame to Excel
		df_actual_production_excel = df_actual_production.copy()
		df_actual_production_excel['Timestamp'] = df_actual_production_excel['Timestamp'].dt.tz_localize(None)
		df_actual_production_excel.to_excel("./data_fetching/Entsoe/Actual_Production_Hydro_River.xlsx", index=False)

		return df_actual_production

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")
		return pd.DataFrame()


def align_and_combine_hydro_data(df_water_reservoir, df_river):
	# Ensure notified and actual dataframes are sorted and set the index to Timestamp
	df_water_reservoir = df_water_reservoir.sort_values(by='Timestamp').set_index('Timestamp')
	df_river = df_river.sort_values(by='Timestamp').set_index('Timestamp')

	# Load volue forecast data and rename the columns appropriately
	# df_volue_forecast = df_volue_forecast.rename(columns={df_volue_forecast.columns[0]: 'Volue Forecast (MW)'})
	
	# # Convert Volue forecast index to datetime and set the timezone to CET (Europe/Berlin)
	# df_volue_forecast.index = pd.to_datetime(df_volue_forecast.index, utc=True).tz_convert('Europe/Berlin')
	
	# # Resample the volue forecast to 15-minute intervals using linear interpolation
	# df_volue_forecast_resampled = df_volue_forecast.resample('15T').interpolate(method='linear')
	
	# # Reset the index of the resampled forecast and rename the index to Timestamp
	# df_volue_forecast_resampled = df_volue_forecast_resampled.reset_index().rename(columns={'index': 'Timestamp'})

	# Align the resampled volue forecast DataFrame to the combined notified and actual hydro production data
	df_combined = pd.concat([df_water_reservoir, df_river], axis=1).reset_index()
	# df_combined['Volue Forecast (MW)'] = df_volue_forecast_resampled.set_index('Timestamp')['Volue Forecast (MW)'].reindex(df_combined['Timestamp']).values

	# Sort the final DataFrame by Timestamp and fill any NaN values with zeros
	df_combined.sort_values(by='Timestamp', inplace=True)
	df_combined.fillna(0, inplace=True)

	# Rename columns for clarity
	df_combined.columns = ['Timestamp', 'Hydro Reservoir Actual (MW)', 'Hydro River Actual (MW)']
	df_combined["Hydro Actual (MW)"] = df_combined["Hydro Reservoir Actual (MW)"] + df_combined["Hydro River Actual (MW)"]
	# Writing the Hydro Production to Excel
	df_combined_excel = df_combined.copy()
	# Remove timezone information from the 'Timestamp' column
	df_combined_excel['Timestamp'] = df_combined_excel['Timestamp'].dt.tz_localize(None)
	# Add a "Quarter" column based on the Timestamp column
	df_combined_excel['Quarter'] = (df_combined_excel.index % 96) + 1
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df_combined_excel['Lookup'] = df_combined_excel["Timestamp"].dt.strftime('%d.%m.%Y') + df_combined_excel["Quarter"].astype(str)
	# Save to Excel
	df_combined_excel.to_excel("./data_fetching/Entsoe/Hydro_Production_Historical.xlsx", index=False)
	
	return df_combined

# ===========================Actual and Forecasted Consumption================================================================

def fetch_consumption_forecast(start_date, end_date):
	"""
	Fetch forecasted consumption data for the specified date range and save it to an Excel file.
	"""
	# Convert start_date and end_date to CET timestamps
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-2)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Format the start and end dates to match the API requirements (yyyymmddhhmm)
	period_start = start_cet.strftime('%Y%m%d%H%M')
	period_end = end_cet.strftime('%Y%m%d%H%M')

	# API request parameters
	url = "https://web-api.tp.entsoe.eu/api"
	params = {
		"securityToken": api_key_entsoe,  # Replace with your actual API key
		"documentType": "A65",           # Forecasted consumption document type
		"processType": "A01",            # Process type for notified consumption
		"outBiddingZone_Domain": "10YRO-TEL------P",  # Romania's area EIC code
		"periodStart": period_start,
		"periodEnd": period_end
	}
	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	try:
		# Make the API request
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()  # Raise error for bad response

		# Parse XML response
		root = ET.fromstring(response.content)
		namespaces = {'ns': root.tag[root.tag.find("{"):root.tag.find("}")+1].strip("{}")}

		timestamps_utc = []
		quantities = []

		# Extract values from XML
		for timeseries in root.findall('ns:TimeSeries', namespaces):
			for period in timeseries.findall('ns:Period', namespaces):
				start_time = period.find('ns:timeInterval/ns:start', namespaces)

				if start_time is None:
					continue

				start_time_utc = datetime.strptime(start_time.text, '%Y-%m-%dT%H:%MZ')

				for point in period.findall('ns:Point', namespaces):
					position_tag = point.find('ns:position', namespaces)
					quantity_tag = point.find('ns:quantity', namespaces)

					if position_tag is None or quantity_tag is None:
						continue

					try:
						position = int(position_tag.text)
						quantity = float(quantity_tag.text)
					except ValueError:
						continue

					point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))
					timestamps_utc.append(point_time_utc)
					quantities.append(quantity)

		# Create DataFrame
		df_forecast = pd.DataFrame({
			'Timestamp_UTC': timestamps_utc,
			'Forecasted Consumption (MW)': quantities
		})

		# Convert UTC to CET
		df_forecast['Timestamp_UTC'] = pd.to_datetime(df_forecast['Timestamp_UTC'])
		df_forecast['Timestamp_CET'] = df_forecast['Timestamp_UTC'].dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')
		df_forecast.drop(columns=['Timestamp_UTC'], inplace=True)
		df_forecast.rename(columns={'Timestamp_CET': 'Timestamp'}, inplace=True)

		# Filter by user-provided date range
		start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
		end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')

		# Create a complete index for 15-minute intervals
		full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')
		df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')

		# Merge the full index with actual data to ensure all intervals are included
		df_forecast = df_full.merge(df_forecast.set_index('Timestamp'), how='left', left_index=True, right_index=True)
		df_forecast.fillna(0, inplace=True)  # Fill missing values with 0
		df_forecast.reset_index(inplace=True)

		# Save the DataFrame to an Excel file
		df_forecast_excel = df_forecast.copy()
		df_forecast_excel['Timestamp'] = df_forecast_excel['Timestamp'].dt.tz_localize(None)  # Remove timezone for Excel
		# Add a "Quarter" column based on the Timestamp column
		df_forecast_excel['Quarter'] = (df_forecast_excel.index % 96) + 1
		# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
		df_forecast_excel['Lookup'] = df_forecast_excel["Timestamp"].dt.strftime('%d.%m.%Y') + df_forecast_excel["Quarter"].astype(str)
		df_forecast_excel.to_excel("./data_fetching/Entsoe/Consumption_Forecast_Historical.xlsx", index=False)

		return df_forecast

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")
		return pd.DataFrame()

def fetch_actual_consumption(start_date, end_date):
	# Setting up the start and end dates (today and tomorrow)
	today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
	start_cet = pd.Timestamp(start_date.strftime('%Y%m%d') + '0000', tz='Europe/Budapest') + timedelta(hours=-2)
	end_cet = pd.Timestamp((end_date + timedelta(days=1)).strftime('%Y%m%d') + '0000', tz='Europe/Budapest')

	# Format the start and end dates to match the API requirements (yyyymmddhhmm)
	period_start = start_cet.strftime('%Y%m%d%H%M')
	period_end = end_cet.strftime('%Y%m%d%H%M')

	url = "https://web-api.tp.entsoe.eu/api"

	params = {
		"securityToken": api_key_entsoe,
		"documentType": "A65",  # Document type for actual load
		"processType": "A16",   # A16 represents Realized consumption
		"outBiddingZone_Domain": "10YRO-TEL------P",  # EIC code for Romania
		"periodStart": period_start,
		"periodEnd": period_end,
	}

	headers = {
		"Content-Type": "application/xml",
		"Accept": "application/xml",
	}

	try:
		response = requests.get(url, params=params, headers=headers)
		response.raise_for_status()
		
		# Parse XML response
		root = ET.fromstring(response.content)
		namespaces = {'ns': root.tag[root.tag.find("{"):root.tag.find("}")+1].strip("{}")}

		timestamps_utc = []
		quantities = []

		# Extract values from XML
		for timeseries in root.findall('ns:TimeSeries', namespaces):
			for period in timeseries.findall('ns:Period', namespaces):
				start_time = period.find('ns:timeInterval/ns:start', namespaces)
				resolution = period.find('ns:resolution', namespaces)

				if start_time is None or resolution is None:
					continue

				start_time_utc = datetime.strptime(start_time.text, '%Y-%m-%dT%H:%MZ')

				for point in period.findall('ns:Point', namespaces):
					position_tag = point.find('ns:position', namespaces)
					quantity_tag = point.find('ns:quantity', namespaces)

					if position_tag is None or quantity_tag is None:
						continue

					try:
						position = int(position_tag.text)
						quantity = float(quantity_tag.text)
					except ValueError as e:
						print(f"Error converting position or quantity: {e}, skipping.")
						continue

					point_time_utc = start_time_utc + timedelta(minutes=15 * (position - 1))
					timestamps_utc.append(point_time_utc)
					quantities.append(quantity)

		# Create DataFrame
		df_actual = pd.DataFrame({
			'Timestamp_UTC': timestamps_utc,
			'Actual Consumption (MW)': quantities
		})

		df_actual['Timestamp_UTC'] = pd.to_datetime(df_actual['Timestamp_UTC'])
		df_actual['Timestamp_CET'] = df_actual['Timestamp_UTC'].dt.tz_localize('UTC').dt.tz_convert('Europe/Berlin')
		df_actual.drop(columns=['Timestamp_UTC'], inplace=True)
		df_actual.rename(columns={'Timestamp_CET': 'Timestamp'}, inplace=True)

		# Filter by user-provided date range
		start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
		end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')

		# Create a complete index for 15-minute intervals
		full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')
		df_full = pd.DataFrame({'Timestamp': full_index_cet}).set_index('Timestamp')

		# Merge the full index with actual data to ensure all intervals are included
		df_actual = df_full.merge(df_actual.set_index('Timestamp'), how='left', left_index=True, right_index=True)
		df_actual.fillna(0, inplace=True)  # Fill missing values with 0
		df_actual.reset_index(inplace=True)

		# Writing the Consumption Actual to Excel
		df_actual_excel = df_actual.copy()
		# Remove timezone information from the 'Timestamp' column
		df_actual_excel['Timestamp'] = df_actual_excel['Timestamp'].dt.tz_localize(None)
		# Add a "Quarter" column based on the Timestamp column
		df_actual_excel['Quarter'] = (df_actual_excel.index % 96) + 1
		# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
		df_actual_excel['Lookup'] = df_actual_excel["Timestamp"].dt.strftime('%d.%m.%Y') + df_actual_excel["Quarter"].astype(str)
		# Save to Excel
		df_actual_excel.to_excel("./data_fetching/Entsoe/Consumption_Actual_Historical.xlsx", index=False)

		return df_actual

	except requests.exceptions.RequestException as e:
		print(f"An error occurred: {e}")
		return pd.DataFrame()

def combine_consumption_data(df_forecast, df_actual):
	# Step 1: Sort and align the wind notified and actual production DataFrames
	df_forecast = df_forecast.sort_values(by='Timestamp').set_index('Timestamp')
	df_actual = df_actual.sort_values(by='Timestamp').set_index('Timestamp')
	
	# Step 2: Concatenate the notified and actual data based on Timestamp
	df_combined = pd.concat([df_forecast, df_actual], axis=1)

	# Step 3: Reset index to make Timestamp a column again
	df_combined.reset_index(inplace=True)
	
	# Step 4: Rename the columns for better readability
	df_combined.columns = ['Timestamp', 'Consumption Forecast (MW)', 'Actual Consumption (MW)']

	# Step 6: Fill any missing values if required (e.g., with 0 or 'NaN')
	# df_combined.fillna(method='ffill', inplace=True)  # Forward fill any missing values if necessary

	return df_combined


# ===========================Activation Energy=======================================================================================

def fetch_igcc_netting_flows_range(start_date, end_date):
    """
    Fetch IGCC netting import/export values from Transelectrica over a range of CET dates.
    Aligns timestamps in CET and computes quarter-hour indices (1–96 per day).
    """
    cet = pytz.timezone("Europe/Berlin")
    utc = pytz.utc
    rows = []
    current_day = start_date

    while current_day <= end_date:
        # Time window: from CET midnight minus 2 hours to next CET midnight
        start_cet = datetime.combine(current_day, datetime.min.time()).replace(tzinfo=cet)
        end_cet = start_cet + timedelta(days=1)

        start_utc = start_cet.astimezone(utc) - timedelta(hours=2)
        end_utc = end_cet.astimezone(utc)

        url = (
            "https://newmarkets.transelectrica.ro/usy-durom-publicreportg01/00121002500000000000000000000100/publicReport/estimatedPowerSystemImbalance"
            f"?timeInterval.from={start_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')}"
            f"&timeInterval.to={end_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')}"
            f"&pageInfo.pageSize=3000"
        )

        response = requests.get(url)
        if response.status_code != 200:
            print(f"❌ Failed for {current_day}: {response.status_code}")
            current_day += timedelta(days=1)
            continue

        try:
            data = response.json().get("itemList", [])
            if not data:
                print(f"⚠️ No IGCC data for {current_day}")
                current_day += timedelta(days=1)
                continue

            for item in data:
                ts = datetime.fromisoformat(item['timeInterval']['from'].replace("Z", "+00:00")).astimezone(cet)
                import_val = float(item.get("imbalanceNettingImport", 0) or 0)
                export_val = float(item.get("imbalanceNettingExport", 0) or 0)
                rows.append([ts.replace(second=0, microsecond=0), import_val, export_val])

        except Exception as e:
            print(f"❌ Error for {current_day}: {e}")

        current_day += timedelta(days=1)

    # --- Build final DataFrame
    df = pd.DataFrame(rows, columns=["Timestamp (CET)", "IGCC Import (MW)", "IGCC Export (MW)"])
    if df.empty:
        print("⚠️ No IGCC data collected.")
        return df

    df["Timestamp (CET)"] = pd.to_datetime(df["Timestamp (CET)"]).dt.tz_localize(None)
    df = df[df["Timestamp (CET)"].dt.date.between(start_date, end_date)]
    df = df.sort_values("Timestamp (CET)").reset_index(drop=True)

    df["Quarter"] = df["Timestamp (CET)"].apply(lambda ts: (ts.hour * 60 + ts.minute) // 15 + 1)
    df["Lookup"] = df["Timestamp (CET)"].dt.strftime('%d.%m.%Y') + df["Quarter"].astype(str)

    # --- Save to Excel
    df.to_excel("./data_fetching/Entsoe/IGCC_Netting_Flows_Historical.xlsx", index=False)
    print("✅ Saved IGCC Netting data.")

    return df

	
# def fetch_intraday_balancing_activations(start_date, end_date):
#     """
#     Fetch activated balancing energy data (aFRR & mFRR) from Transelectrica API
#     for a user-defined date range.
#     """

#     # Define CET timezone
#     cet_timezone = pytz.timezone("Europe/Berlin")

#     # Adjust `start_date` to include the last quarter of the previous day (23:45)
#     start_cet_adjusted = datetime.combine(start_date - timedelta(days=1), time(23, 45)).replace(tzinfo=cet_timezone)
#     end_cet_midnight = datetime.combine(end_date, datetime.min.time()).replace(tzinfo=cet_timezone) + timedelta(days=1)

#     # Convert CET times to UTC (API operates in UTC)
#     start_utc = start_cet_adjusted.astimezone(pytz.utc)
#     end_utc = end_cet_midnight.astimezone(pytz.utc)

#     # Format timestamps for API request
#     from_time = start_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")
#     to_time = end_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")

#     # ✅ Debugging: Print time range
#     print(f"\n🔍 Requesting data from {from_time} UTC to {to_time} UTC")

#     # ✅ API Request (Transelectrica)
#     url = f"https://newmarkets.transelectrica.ro/usy-durom-publicreportg01/00121002500000000000000000000100/publicReport/activatedBalancingEnergyOverview?timeInterval.from={from_time}&timeInterval.to={to_time}&pageInfo.pageSize=3000"
    
#     response = requests.get(url)
    
#     # ✅ Check response status
#     if response.status_code != 200:
#         print(f"❌ Failed to fetch data. Status code: {response.status_code}")
#         return pd.DataFrame()

#     # ✅ Parse JSON response
#     data = response.json()
#     items = data.get("itemList", [])

#     if not items:
#         print(f"⚠️ No data found for the selected date range ({start_date} to {end_date}).")
#         return pd.DataFrame()

#     print(f"✅ Successfully fetched {len(items)} records from API.")

#     # ✅ Process and convert timestamps
#     rows = []
#     for item in items:
#         try:
#             # Convert timestamps from UTC to CET
#             utc_from = datetime.fromisoformat(item['timeInterval']['from'].replace('Z', '+00:00'))
#             cet_from = utc_from.astimezone(cet_timezone)

#             # **✅ Round timestamps to the nearest minute**
#             cet_from = cet_from.replace(second=0, microsecond=0)

#             # ✅ Extract energy values (convert None to 0)
#             afrr_up = item.get("aFRR_Up", 0) or 0
#             afrr_down = item.get("aFRR_Down", 0) or 0
#             mfrr_up = item.get("mFRR_Up", 0) or 0
#             mfrr_down = item.get("mFRR_Down", 0) or 0

#             # ✅ Store processed row
#             rows.append([cet_from, afrr_up, afrr_down, mfrr_up, mfrr_down])

#         except Exception as e:
#             print(f"❌ Error processing record: {e}")

#     # ✅ Convert to DataFrame
#     df = pd.DataFrame(rows, columns=["Timestamp (CET)", "aFRR Up (MW)", "aFRR Down (MW)", "mFRR Up (MW)", "mFRR Down (MW)"])

#     # ✅ Ensure timestamps are in the correct format and remove extra precision
#     df["Timestamp (CET)"] = pd.to_datetime(df["Timestamp (CET)"]).dt.floor("min")

#     # ✅ Adjust start/end times based on the extracted data
#     if not df.empty:
#         start_cet = df["Timestamp (CET)"].min()  # Get earliest timestamp from extracted data
#         end_cet = df["Timestamp (CET)"].max()  # Get latest timestamp from extracted data
#     else:
#         print("⚠️ DataFrame is empty, using default time range!")
#         start_cet = pd.Timestamp(start_date.strftime('%Y-%m-%dT00:00:00'), tz='Europe/Berlin')
#         end_cet = pd.Timestamp(end_date.strftime('%Y-%m-%dT23:45:00'), tz='Europe/Berlin')

#     # ✅ Generate the corrected timestamp index
#     full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')

#     # ✅ Merge with real data, keeping extracted values
#     df_full = pd.DataFrame({'Timestamp (CET)': full_index_cet}).set_index("Timestamp (CET)")
#     df = df.set_index("Timestamp (CET)")

#     # ✅ Merge instead of reindexing to prevent overwriting
#     df = df_full.merge(df, how='left', left_index=True, right_index=True)

#     # ✅ Fill missing values *only* where data was not present
#     df.fillna(0, inplace=True)

#     # Reset index
#     df.reset_index(inplace=True)

#     # ✅ Ensure first quarter (`00:00`) is included
#     if df["Timestamp (CET)"].iloc[0].time() != time(0, 0):
#         print("⚠️ Missing 00:00 quarter, adding it manually.")
#         first_quarter = pd.DataFrame([[datetime.combine(start_date, time(0, 0)), 0, 0, 0, 0]],
#                                      columns=df.columns)
#         df = pd.concat([first_quarter, df], ignore_index=True)

#     # ✅ Debug: Print first and last timestamps to verify range
#     if not df.empty:
#         print("\n✅ First Timestamp in DataFrame:", df.iloc[0]["Timestamp (CET)"])
#         print("✅ Last Timestamp in DataFrame:", df.iloc[-1]["Timestamp (CET)"])

#     print("📊 Processed DataFrame:")
#     print(df.head())

#     # Remove timezone information from the 'Timestamp' column
#     df['Timestamp (CET)'] = pd.to_datetime(df['Timestamp (CET)'], utc=False, errors='coerce')

#     if pd.api.types.is_datetime64_any_dtype(df['Timestamp (CET)']):
#         if df['Timestamp (CET)'].dt.tz is not None:
#             df['Timestamp (CET)'] = df['Timestamp (CET)'].dt.tz_localize(None)
#     # Add a "Quarter" column based on the Timestamp column
#     df['Quarter'] = ((df.index % 96) + 1)

#     # Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
#     df['Lookup'] = df["Timestamp (CET)"].dt.strftime('%d.%m.%Y') + df["Quarter"].astype(str)

#     # Save to Excel
#     df.to_excel("./data_fetching/Entsoe/Balancing_Energy_Activation.xlsx", index=False)

#     return df

# def fetch_intraday_balancing_activations_2(start_date, end_date):
#     cet_timezone = pytz.timezone("Europe/Berlin")

#     # Adjust `start_date` to include the last quarter of the previous day (23:45)
#     start_cet_adjusted = datetime.combine(start_date - timedelta(days=1), time(23, 45)).replace(tzinfo=cet_timezone)
#     end_cet_midnight = datetime.combine(end_date, datetime.min.time()).replace(tzinfo=cet_timezone) + timedelta(days=1)

#     start_utc = start_cet_adjusted.astimezone(pytz.utc)
#     end_utc = end_cet_midnight.astimezone(pytz.utc)

#     from_time = start_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")
#     to_time = end_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")

#     print(f"\n🔍 Requesting data from {from_time} UTC to {to_time} UTC")

#     url = f"https://newmarkets.transelectrica.ro/usy-durom-publicreportg01/00121002500000000000000000000100/publicReport/activatedBalancingEnergyOverview?timeInterval.from={from_time}&timeInterval.to={to_time}&pageInfo.pageSize=3000"
#     response = requests.get(url)

#     if response.status_code != 200:
#         print(f"❌ Failed to fetch data. Status code: {response.status_code}")
#         return pd.DataFrame()

#     data = response.json()
#     items = data.get("itemList", [])
#     if not items:
#         print(f"⚠️ No data found for the selected date range ({start_date} to {end_date}).")
#         return pd.DataFrame()

#     print(f"✅ Successfully fetched {len(items)} records from API.")

#     rows = []
#     for item in items:
#         try:
#             utc_from = datetime.fromisoformat(item['timeInterval']['from'].replace('Z', '+00:00'))
#             cet_from = utc_from.astimezone(cet_timezone).replace(second=0, microsecond=0)

#             afrr_up = item.get("aFRR_Up", 0) or 0
#             afrr_down = item.get("aFRR_Down", 0) or 0
#             mfrr_up = item.get("mFRR_Up", 0) or 0
#             mfrr_down = item.get("mFRR_Down", 0) or 0

#             rows.append([cet_from, afrr_up, afrr_down, mfrr_up, mfrr_down])
#         except Exception as e:
#             print(f"❌ Error processing record: {e}")

#     df = pd.DataFrame(rows, columns=["Timestamp (CET)", "aFRR Up (MW)", "aFRR Down (MW)", "mFRR Up (MW)", "mFRR Down (MW)"])

#     # Ensure datetime column
#     df["Timestamp (CET)"] = pd.to_datetime(df["Timestamp (CET)"], utc=False, errors='coerce')
#     df = df[df["Timestamp (CET)"].notnull()]  # Drop parsing failures

#     # Floor to minute
#     df["Timestamp (CET)"] = df["Timestamp (CET)"].dt.floor("min")

#     # Generate full 15-minute range
#     start_cet = df["Timestamp (CET)"].min()
#     end_cet = df["Timestamp (CET)"].max()
#     full_index_cet = pd.date_range(start=start_cet, end=end_cet, freq='15T', tz='Europe/Berlin')

#     df_full = pd.DataFrame({'Timestamp (CET)': full_index_cet})
#     df_full["Timestamp (CET)"] = pd.to_datetime(df_full["Timestamp (CET)"]).dt.tz_localize(None)

#     # Prepare base DataFrame and merge
#     df["Timestamp (CET)"] = df["Timestamp (CET)"].dt.tz_localize(None)
#     df = df_full.merge(df, on="Timestamp (CET)", how="left").fillna(0)

#     # Add 00:00 quarter if missing
#     first_timestamp = pd.Timestamp(datetime.combine(start_date, time(0, 0)))
#     if first_timestamp not in df["Timestamp (CET)"].values:
#         print("⚠️ Adding missing 00:00 quarter manually.")
#         df = pd.concat([
#             pd.DataFrame([[first_timestamp, 0, 0, 0, 0]], columns=df.columns),
#             df
#         ]).drop_duplicates(subset="Timestamp (CET)").sort_values("Timestamp (CET)").reset_index(drop=True)

#     # Add Quarter and Lookup columns
#     df['Quarter'] = ((df.index % 96) + 1)
#     df['Lookup'] = df["Timestamp (CET)"].dt.strftime('%d.%m.%Y') + df["Quarter"].astype(str)

#     # Export and return
#     df.to_excel("./data_fetching/Entsoe/Balancing_Energy_Activation.xlsx", index=False)
#     print("✅ Final DataFrame:")
#     print(df.head())

#     return df


# def fetch_intraday_balancing_activations_3(start_date, end_date, retries=3, delay=5):
#     """
#     Fetch activated balancing energy data (aFRR & mFRR) from Transelectrica API
#     for a user-defined date range (in CET). Handles misaligned timestamps,
#     aligns to 15-min quarters, and saves to Excel.
#     """

#     cet = pytz.timezone("Europe/Berlin")
#     start_cet = datetime.combine(start_date, time(0, 0)).replace(tzinfo=cet)
#     end_cet = datetime.combine(end_date, time(23, 45)).replace(tzinfo=cet)
#     start_utc = start_cet.astimezone(pytz.utc)
#     end_utc = (end_cet + timedelta(minutes=15)).astimezone(pytz.utc)

#     from_time = start_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")
#     to_time = end_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")

#     url = f"https://newmarkets.transelectrica.ro/usy-durom-publicreportg01/00121002500000000000000000000100/publicReport/activatedBalancingEnergyOverview?timeInterval.from={from_time}&timeInterval.to={to_time}&pageInfo.pageSize=3000"

#     for attempt in range(retries):
#         try:
#             response = requests.get(url)
#             response.raise_for_status()
#             break
#         except requests.exceptions.RequestException as e:
#             print(f"Attempt {attempt + 1}/{retries} - Network error: {e}")
#             if attempt < retries - 1:
#                 time_module.sleep(delay)
#             else:
#                 print("❌ Max retries reached. Returning empty DataFrame.")
#                 return pd.DataFrame()

#     data = response.json()
#     items = data.get("itemList", [])
#     if not items:
#         print("⚠️ No data returned.")
#         return pd.DataFrame()

#     rows = []
#     for item in items:
#         try:
#             utc_from = datetime.fromisoformat(item['timeInterval']['from'].replace('Z', '+00:00'))
#             cet_from = utc_from.astimezone(cet)

#             # ✅ Round to nearest lower 15-min interval
#             cet_from = cet_from.floor("15min")

#             rows.append([
#                 cet_from,
#                 float(item.get("aFRR_Up", 0) or 0),
#                 float(item.get("aFRR_Down", 0) or 0),
#                 float(item.get("mFRR_Up", 0) or 0),
#                 float(item.get("mFRR_Down", 0) or 0)
#             ])
#         except Exception as e:
#             print(f"❌ Error processing record: {e}")

#     df = pd.DataFrame(rows, columns=[
#         "Timestamp (CET)", "aFRR Up (MW)", "aFRR Down (MW)", "mFRR Up (MW)", "mFRR Down (MW)"
#     ])

#     # ✅ Generate full index with all quarters
#     full_index = pd.date_range(start=start_cet, end=end_cet, freq="15T", tz=cet).tz_localize(None)
#     df_full = pd.DataFrame({'Timestamp (CET)': full_index})
#     df = pd.merge(df_full, df, on="Timestamp (CET)", how="left").fillna(0)

#     # ✅ Remove tz info
#     df["Timestamp (CET)"] = pd.to_datetime(df["Timestamp (CET)"]).dt.tz_localize(None)

#     # ✅ Quarter calculation: 1–96
#     df["Quarter"] = df["Timestamp (CET)"].dt.hour * 4 + (df["Timestamp (CET)"].dt.minute // 15) + 1

#     # ✅ Lookup format: dd.mm.yyyy + quarter
#     df["Lookup"] = df["Timestamp (CET)"].dt.strftime('%d.%m.%Y') + df["Quarter"].astype(str)

#     # ✅ Save to Excel
#     df.to_excel("./data_fetching/Entsoe/Balancing_Energy_Activation.xlsx", index=False)
#     print("✅ Saved to ./data_fetching/Entsoe/Balancing_Energy_Activation.xlsx")

#     return df


# def fetch_intraday_balancing_activations_4(start_date, end_date):
#     """Fetch activated balancing energy (aFRR & mFRR) with DST-aware UTC→CET alignment."""

#     cet_tz = pytz.timezone("Europe/Bucharest")
#     all_rows = []

#     current_day = start_date
#     while current_day <= end_date:
#         # Determine DST offset for this day
#         cet_midnight = cet_tz.localize(datetime.combine(current_day, time(0, 0)))
#         utc_equivalent = cet_midnight.astimezone(pytz.utc)

#         # Calculate how many 15-min quarters we need from previous UTC day
#         quarter_shift = int((cet_midnight - utc_equivalent).total_seconds() // (15 * 60))

#         # Fetch from (midnight - shift) UTC to next day midnight UTC
#         utc_start = utc_equivalent - timedelta(minutes=15 * quarter_shift)
#         utc_end = utc_equivalent + timedelta(days=1)

#         from_time = utc_start.strftime("%Y-%m-%dT%H:%M:%S.000Z")
#         to_time = utc_end.strftime("%Y-%m-%dT%H:%M:%S.000Z")

#         print(f"📡 Fetching UTC data from {from_time} to {to_time} for CET date {current_day}")

#         url = f"https://newmarkets.transelectrica.ro/usy-durom-publicreportg01/00121002500000000000000000000100/publicReport/activatedBalancingEnergyOverview?timeInterval.from={from_time}&timeInterval.to={to_time}&pageInfo.pageSize=3000"
#         response = requests.get(url)

#         if response.status_code != 200:
#             print(f"❌ Failed to fetch data for {current_day} | Status: {response.status_code}")
#             current_day += timedelta(days=1)
#             continue

#         try:
#             items = response.json().get("itemList", [])
#             if not items:
#                 print(f"⚠️ No data returned for {current_day}")
#                 current_day += timedelta(days=1)
#                 continue

#             for item in items:
#                 try:
#                     utc_from = datetime.fromisoformat(item['timeInterval']['from'].replace('Z', '+00:00'))
#                     cet_from = utc_from.astimezone(cet_tz).replace(second=0, microsecond=0)

#                     afrr_up = item.get("aFRR_Up", 0) or 0
#                     afrr_down = item.get("aFRR_Down", 0) or 0
#                     mfrr_up = item.get("mFRR_Up", 0) or 0
#                     mfrr_down = item.get("mFRR_Down", 0) or 0

#                     all_rows.append([cet_from, afrr_up, afrr_down, mfrr_up, mfrr_down])
#                 except Exception as e:
#                     print(f"❌ Error parsing item: {e}")

#         except Exception as e:
#             print(f"❌ JSON error: {e}")

#         current_day += timedelta(days=1)

#     # Construct DataFrame
#     df = pd.DataFrame(all_rows, columns=["Timestamp (CET)", "aFRR Up (MW)", "aFRR Down (MW)", "mFRR Up (MW)", "mFRR Down (MW)"])

#     # Drop timezone info, sort
#     df["Timestamp (CET)"] = pd.to_datetime(df["Timestamp (CET)"]).dt.tz_localize(None)
#     df = df.sort_values("Timestamp (CET)").reset_index(drop=True)

#     # Filter only quarters in target range
#     full_range = pd.date_range(start=start_date, end=end_date + timedelta(days=1) - timedelta(minutes=15), freq="15min", tz=cet_tz)
#     full_range = full_range.tz_localize(None)

#     df = df[df["Timestamp (CET)"].isin(full_range)].reset_index(drop=True)

#     # Add Quarter and Lookup
#     df['Quarter'] = ((df.index % 96) + 1)
#     df['Lookup'] = df["Timestamp (CET)"].dt.strftime('%d.%m.%Y') + df["Quarter"].astype(str)

#     # Save to Excel
#     df.to_excel("./data_fetching/Entsoe/Balancing_Energy_Activation.xlsx", index=False)
#     print("✅ Saved to Excel: Balancing_Energy_Activation.xlsx")

#     return df

# def fetch_intraday_balancing_activations(start_date, end_date):
#     """Fetch activated balancing energy data and align it to CET."""

#     # Timezones
#     utc = pytz.utc
#     cet = pytz.timezone("CET")

#     # Fetch window: 2 hours before CET midnight
#     request_start = datetime.combine(start_date - timedelta(days=1), time(22, 0)).replace(tzinfo=cet)
#     request_end = datetime.combine(end_date + timedelta(days=1), time(0, 0)).replace(tzinfo=cet)

#     # Convert to UTC for API
#     from_time = request_start.astimezone(utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
#     to_time = request_end.astimezone(utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")

#     url = f"https://newmarkets.transelectrica.ro/usy-durom-publicreportg01/00121002500000000000000000000100/publicReport/activatedBalancingEnergyOverview?timeInterval.from={from_time}&timeInterval.to={to_time}&pageInfo.pageSize=3000"
#     response = requests.get(url)

#     if response.status_code != 200:
#         print(f"Failed to fetch data: {response.status_code}")
#         return pd.DataFrame()

#     data = response.json().get("itemList", [])
#     records = []

#     for item in data:
#         try:
#             ts_utc = datetime.fromisoformat(item["timeInterval"]["from"].replace("Z", "+00:00")).astimezone(cet)
#             ts_utc = ts_utc.replace(second=0, microsecond=0)
#             records.append({
#                 "Timestamp (CET)": ts_utc,
#                 "aFRR Up (MW)": item.get("aFRR_Up", 0) or 0,
#                 "aFRR Down (MW)": item.get("aFRR_Down", 0) or 0,
#                 "mFRR Up (MW)": item.get("mFRR_Up", 0) or 0,
#                 "mFRR Down (MW)": item.get("mFRR_Down", 0) or 0,
#             })
#         except Exception as e:
#             print(f"Error processing record: {e}")

#     df = pd.DataFrame(records)

#     if df.empty:
#         print("No data available.")
#         return df

#     # Filter strictly for CET date range
#     df = df[(df["Timestamp (CET)"] >= pd.Timestamp(start_date, tz=cet)) &
#             (df["Timestamp (CET)"] < pd.Timestamp(end_date + timedelta(days=1), tz=cet))]

#     df["Timestamp (CET)"] = df["Timestamp (CET)"].dt.tz_localize(None)

#     df["Quarter"] = ((df.index % 96) + 1)
#     df["Lookup"] = df["Timestamp (CET)"].dt.strftime('%d.%m.%Y') + df["Quarter"].astype(str)

#     # Save to Excel
#     output_path = "./data_fetching/Entsoe/Balancing_Energy_Activation_CET.xlsx"
#     os.makedirs(os.path.dirname(output_path), exist_ok=True)
#     df.to_excel(output_path, index=False)

#     return df

def fetch_intraday_balancing_activations(start_date, end_date):
    """
    Fetch activated aFRR and mFRR balancing energy from Transelectrica API 
    and align the data in CET with proper quarter-hour indexing (1 to 96).

    Each day is built from UTC data starting 2 hours earlier to account for 
    Daylight Saving Time (when CET = UTC+2), ensuring the CET day starts at 00:00.
    """
    cet = pytz.timezone("Europe/Berlin")
    utc = pytz.utc

    rows = []
    current_day = start_date

    while current_day <= end_date:
        # --- Build the UTC time window: 2 hours before CET midnight to next CET midnight
        start_cet_dt = datetime.combine(current_day, datetime.min.time()).replace(tzinfo=cet)
        end_cet_dt = start_cet_dt + timedelta(days=1)

        start_utc = start_cet_dt - timedelta(hours=2)  # Covers 00:00 CET = 22:00 UTC
        end_utc = end_cet_dt

        url = (
            "https://newmarkets.transelectrica.ro/usy-durom-publicreportg01/"
            "00121002500000000000000000000100/publicReport/activatedBalancingEnergyOverview"
            f"?timeInterval.from={start_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')}"
            f"&timeInterval.to={end_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')}"
            f"&pageInfo.pageSize=3000"
        )

        response = requests.get(url)
        if response.status_code != 200:
            print(f"❌ Failed for {current_day}: {response.status_code}")
            current_day += timedelta(days=1)
            continue

        try:
            data = response.json().get("itemList", [])
            if not data:
                print(f"⚠️ No data for {current_day}")
                current_day += timedelta(days=1)
                continue

            for item in data:
                start = datetime.fromisoformat(item['timeInterval']['from'].replace("Z", "+00:00")).astimezone(cet)
                afrr_up = item.get("aFRR_Up", 0) or 0
                afrr_down = item.get("aFRR_Down", 0) or 0
                mfrr_up = item.get("mFRR_Up", 0) or 0
                mfrr_down = item.get("mFRR_Down", 0) or 0

                rows.append([start.replace(second=0, microsecond=0), afrr_up, afrr_down, mfrr_up, mfrr_down])

        except Exception as e:
            print(f"❌ Error processing data for {current_day}: {e}")

        current_day += timedelta(days=1)

    # --- Create DataFrame
    df = pd.DataFrame(rows, columns=["Timestamp (CET)", "aFRR Up (MW)", "aFRR Down (MW)", "mFRR Up (MW)", "mFRR Down (MW)"])

    if df.empty:
        print("⚠️ No data collected.")
        return df

    # --- Strip timezone info and sort
    df["Timestamp (CET)"] = pd.to_datetime(df["Timestamp (CET)"]).dt.tz_localize(None)
    df = df.sort_values("Timestamp (CET)").reset_index(drop=True)

    # --- Filter for timestamps that belong to actual requested CET days
    df = df[df["Timestamp (CET)"].dt.date.between(start_date, end_date)]

    # --- Add quarter column (1 to 96 for each day)
    df["Quarter"] = df["Timestamp (CET)"].apply(lambda ts: (ts.hour * 60 + ts.minute) // 15 + 1)

    # --- Create lookup key: e.g., "01.04.202512" for 12th quarter on April 1st
    df["Lookup"] = df["Timestamp (CET)"].dt.strftime('%d.%m.%Y') + df["Quarter"].astype(str)

    # --- Save to Excel
    df.to_excel("./data_fetching/Entsoe/Balancing_Energy_Activation.xlsx", index=False)
    print("✅ Saved aligned balancing activation data in CET.")

    return df

# 1 Imbalance Prices
def imbalance_prices(start, end):
	# Define the start and end timestamps for the query
	# start = pd.Timestamp('20240401', tz='Europe/Bucharest')  # Adjust the start date as needed
	# end = pd.Timestamp('20240411', tz='Europe/Bucharest')    # Adjust the end date as needed

	# Set the country code for Romania
	country_code = 'RO'  # ISO-3166 alpha-2 code for Romania

	# Fetch imbalance prices
	imbalance_prices = client.query_imbalance_prices(country_code, start=start, end=end, psr_type=None)

	# Display or analyze the fetched data
	return imbalance_prices

# 2. Imbalance Volumes
def imbalance_volumes(start, end):
	# Define the start and end timestamps for the query
	# start = pd.Timestamp('20240401', tz='Europe/Bucharest')  # Adjust the start date as needed
	# end = pd.Timestamp('20240411', tz='Europe/Bucharest')    # Adjust the end date as needed

	# Set the country code for Romania
	country_code = 'RO'  # ISO-3166 alpha-2 code for Romania

	# Fetch imbalance prices
	imbalance_volumes = client.query_imbalance_volumes(country_code, start=start, end=end, psr_type=None)

	# Display or analyze the fetched data
	return imbalance_volumes

#3. Contracted Aggragated mFRR Volumes
def activated_mFRR_energy(start, end):
	# Update parameters with Czech control area domain
	params = {
		'documentType': 'A83',
		"controlArea_Domain": "10YRO-TEL------P",
		"businessType": "A97"
	}

	# Make the API request
	response = client._base_request(params=params, start=start, end=end)

	# Process the response
	if response.status_code == 200:
		# Check if the response is empty
		if response.text:
			# Check the content type
			content_type = response.headers.get('Content-Type', '')
			if 'application/json' in content_type:
				try:
					data = response.json()
					# st.write(data)
				except ValueError as e:
					st.error(f"Invalid JSON data received: {e}")
			else:
				st.error(f"Unexpected content type: {content_type}")
				# st.text(f"Response content: {response.text}")
		else:
			st.warning("Empty response received.")
	else:
		st.error(f"Failed to retrieve data: HTTP Status {response.status_code}")

	return response.text

# Defining the function to process the server response
def creating_mFRR_dfs(data):
	# Creating two dataframes, one for Up and another for Down
	# The XML content you provided
	xml_data = data

	# Define the namespace map to use with ElementTree
	namespaces = {
		'ns': 'urn:iec62325.351:tc57wg16:451-6:balancingdocument:3:0'  # 'ns' is a placeholder for the namespace
	}

	# Parse the XML
	root = ET.fromstring(xml_data)

	# Dictionary to hold DataFrames, one for each TimeSeries
	dfs = {}

	# Counter to identify TimeSeries
	time_series_id = 1

	# Navigate through the XML structure and extract data
	for ts in root.findall('.//ns:TimeSeries', namespaces):
		timestamps = []
		positions = []
		quantities = []

		for point in ts.findall('.//ns:Point', namespaces):
			position = int(point.find('ns:position', namespaces).text)
			quantity = point.find('ns:quantity', namespaces).text
			# Assuming the timestamp is the same for all points in a TimeSeries for simplicity
			time_interval = ts.find('.//ns:Period/ns:timeInterval', namespaces)
			start = time_interval.find('ns:start', namespaces).text
			end = time_interval.find('ns:end', namespaces).text
			start_date = pd.to_datetime(start).strftime('%d.%m.%Y')  # Format as "dd.mm.yyyy"
			end_date = pd.to_datetime(end).strftime('%d.%m.%Y')  # Format as "dd.mm.yyyy"
			timestamps.append(end_date)
			positions.append(position)
			quantities.append(quantity)

		# Create a DataFrame for the current TimeSeries
		df = pd.DataFrame({
			'Date': timestamps,
			'Interval': positions,
			'Quantity': quantities
		})
		
		# Rename columns based on TimeSeries number
		if time_series_id == 1:
			df.rename(columns={'Quantity': 'Activated Energy Down'}, inplace=True)
		else:
			df.rename(columns={'Quantity': 'Activated Energy Up'}, inplace=True)
		
		# Store the DataFrame in the dictionary with a unique key
		dfs[f"TimeSeries_{time_series_id}"] = df
		time_series_id += 1

	# Now dfs['TimeSeries_1'] will hold the DataFrame for the first TimeSeries
	# and dfs['TimeSeries_2'] for the second, etc.

	# Example to print DataFrame for TimeSeries 1 and 2
	st.write("DataFrame for TimeSeries 1:")
	st.dataframe(dfs['TimeSeries_1'])
	st.write("\nDataFrame for TimeSeries 2:")
	st.dataframe(dfs['TimeSeries_2'])

	# Merge the dataframes on 'Date' and 'Interval'
	activated_energy_df = pd.merge(dfs['TimeSeries_1'], dfs['TimeSeries_2'], on=['Date', 'Interval'], how='outer')
	return activated_energy_df

# Defining the functions fetch the data for a range of days
def make_api_request(params, start, end):
	# Replace with your actual API request logic
	response = client._base_request(params=params, start=start, end=end)
	return response

def query_daily_mFRR(start_date, end_date):
	current_date = start_date
	final_dfs = []  # List to store each day's DataFrame

	while current_date < end_date:
		# Generate start and end timestamps for the current day
		start = pd.Timestamp(f"{current_date.strftime('%Y%m%d')}0000", tz='Europe/Bucharest')
		end = pd.Timestamp(f"{(current_date + timedelta(days=1)).strftime('%Y%m%d')}0000", tz='Europe/Bucharest')
		
		# Update parameters with the control area domain
		params = {
			'documentType': 'A83',
			"controlArea_Domain": "10YRO-TEL------P",
			"businessType": "A97"
		}
		
		# Make the API request for the current day
		response = make_api_request(params, start, end)

		# Process the API response into a DataFrame
		if response.status_code == 200 and response.text:
			day_df = creating_mFRR_dfs(response.text)  # Assume this function returns a DataFrame
			final_dfs.append(day_df)
		else:
			st.error(f"Failed to retrieve data for {current_date.strftime('%Y-%m-%d')}: HTTP Status {response.status_code}")

		# Move to the next day
		current_date += timedelta(days=1)

	# Concatenate all daily DataFrames into a single DataFrame
	if final_dfs:
		combined_df = pd.concat(final_dfs, ignore_index=True)
		return combined_df
	else:
		return pd.DataFrame()  # Return an empty DataFrame if no data was collected

#4. Fethcing the Actual Load 
def actual_load(start, end):
	# Define the start and end timestamps for the query
	# start = pd.Timestamp('20240401', tz='Europe/Bucharest')  # Adjust the start date as needed
	# end = pd.Timestamp('20240411', tz='Europe/Bucharest')    # Adjust the end date as needed

	# Set the country code for Romania
	country_code = 'RO'  # ISO-3166 alpha-2 code for Romania

	# Fetch imbalance prices
	actual_load = client.query_load(country_code, start=start, end=end)

	# Display or analyze the fetched data
	return actual_load

#5. Fetching the Generation Forecast
def generation_forecast(start, end):
	# Define the start and end timestamps for the query
	# start = pd.Timestamp('20240401', tz='Europe/Bucharest')  # Adjust the start date as needed
	# end = pd.Timestamp('20240411', tz='Europe/Bucharest')    # Adjust the end date as needed

	# Set the country code for Romania
	country_code = 'RO'  # ISO-3166 alpha-2 code for Romania

	# Fetch generation forecast
	generation_forecast = client.query_generation_forecast(country_code, start=start, end=end)

	# Display or analyze the fetched data
	return generation_forecast

#6. Fetching the Actual Geneeration per Source
def actual_generation_source(start, end):
	# Define the start and end timestamps for the query
	# start = pd.Timestamp('20240401', tz='Europe/Bucharest')  # Adjust the start date as needed
	# end = pd.Timestamp('20240411', tz='Europe/Bucharest')    # Adjust the end date as needed

	# Set the country code for Romania
	country_code = 'RO'  # ISO-3166 alpha-2 code for Romania

	# Fetch imbalance prices
	actual_generation_source = client.query_generation(country_code, start=start, end=end, psr_type=None, include_eic=False)

	# Display or analyze the fetched data
	return actual_generation_source

#7. Fetching the Forecast Load
def load_forecast(start, end):
	# Define the start and end timestamps for the query
	# start = pd.Timestamp('20240401', tz='Europe/Bucharest')  # Adjust the start date as needed
	# end = pd.Timestamp('20240411', tz='Europe/Bucharest')    # Adjust the end date as needed

	# Set the country code for Romania
	country_code = 'RO'  # ISO-3166 alpha-2 code for Romania

	# Fetch Load Forecast
	load_forecast = client.query_load_forecast(country_code, start=start, end=end)

	# Display or analyze the fetched data
	return load_forecast

#8. Fetching the Forecast Load in CET for the Trading_Tool Update
def load_forecast_CET(start_cet, end_cet):
	# Define the start and end timestamps for the query
	# start = pd.Timestamp('20240401', tz='Europe/Bucharest')  # Adjust the start date as needed
	# end = pd.Timestamp('20240411', tz='Europe/Bucharest')    # Adjust the end date as needed

	# Set the country code for Romania
	country_code = 'RO'  # ISO-3166 alpha-2 code for Romania
	st.session_state.start_cet = start_cet
	st.session_state.end_cet = end_cet
	st.write(start_cet, end_cet)
	# Fetch Load Forecast
	load_forecast_cet = client.query_load_forecast(country_code, start=start_cet, end=end_cet)
	load_forecast_cet.reset_index(inplace=True)
	load_forecast_cet.columns = ["Timestamp", "Load"]

	# Convert 'Timestamp' to datetime format
	load_forecast_cet['Timestamp'] = pd.to_datetime(load_forecast_cet['Timestamp'])
	# Convert datetime to CET timezone
	load_forecast_cet['Timestamp'] = load_forecast_cet['Timestamp'].dt.tz_convert('CET')

	# Remove timezone information but keep the time in CET
	load_forecast_cet['Timestamp'] = load_forecast_cet['Timestamp'].dt.tz_localize(None)
	load_forecast_cet.to_excel("./Market Fundamentals/Entsoe_data/load_forecast.xlsx", index=False)
	# Display or analyze the fetched data
	return load_forecast_cet

# Unintended Deviations ================================================================================================================
def fetch_unintended_deviation_data(start_cet, end_cet):
    """Fetch unintended deviation data (import/export) over multiple days, converting timestamps to CET."""

    # Define CET timezone
    cet_timezone = pytz.timezone("Europe/Berlin")

    # Prepare list to store data
    all_rows = []

    # Loop through each day in the range
    current_day = start_cet
    while current_day <= end_cet:
        # Set start & end of the day in CET
        start_cet_adjusted = datetime.combine(current_day - timedelta(days=1), time(23, 45)).replace(tzinfo=cet_timezone)
        end_cet_midnight = datetime.combine(current_day, datetime.min.time()).replace(tzinfo=cet_timezone) + timedelta(days=1)

        # Convert CET times to UTC for API request
        start_utc = start_cet_adjusted.astimezone(pytz.utc)
        end_utc = end_cet_midnight.astimezone(pytz.utc)

        # Format timestamps for API request
        from_time = start_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")
        to_time = end_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")

        # Debug: Print request time range
        print(f"\n🔍 Requesting data from {from_time} UTC to {to_time} UTC for {current_day.strftime('%Y-%m-%d')}")

        # API Request
        url = f"https://newmarkets.transelectrica.ro/usy-durom-publicreportg01/00121002500000000000000000000100/publicReport/estimatedPowerSystemImbalance?timeInterval.from={from_time}&timeInterval.to={to_time}&pageInfo.pageSize=3000"

        response = requests.get(url)

        if response.status_code != 200:
            print(f"⚠️ Failed to fetch data for {current_day.strftime('%Y-%m-%d')}. Status code: {response.status_code}")
            current_day += timedelta(days=1)  # Move to the next day
            continue

        # Parse JSON response
        try:
            data = response.json()
            items = data.get("itemList", [])

            print(f"✅ Fetched {len(items)} records for {current_day.strftime('%Y-%m-%d')}.")

            if len(items) == 0:
                print("⚠️ No data found in itemList.")
                current_day += timedelta(days=1)
                continue

            # Process and convert timestamps
            for item in items:
                try:
                    # Convert timestamps from UTC to CET
                    utc_from = datetime.fromisoformat(item['timeInterval']['from'].replace('Z', '+00:00'))
                    utc_to = datetime.fromisoformat(item['timeInterval']['to'].replace('Z', '+00:00'))

                    cet_from = utc_from.astimezone(cet_timezone)
                    cet_to = utc_to.astimezone(cet_timezone)

                    # Store in formatted string
                    time_period = f"{cet_from.strftime('%Y-%m-%d %H:%M:%S')} - {cet_to.strftime('%Y-%m-%d %H:%M:%S')}"

                    # Extract values (handling potential missing values)
                    unintended_import = float(item.get("estimatedUnintendedDeviationINArea", 0) or 0)
                    unintended_export = float(item.get("estimatedUnintendedDeviationOUTArea", 0) or 0)

                    # Debugging - Print added records
                    print(f"ADDING: {time_period} | IN: {unintended_import}, OUT: {unintended_export}")

                    # Store processed row
                    all_rows.append([cet_from, unintended_import, unintended_export])

                except Exception as e:
                    print(f"❌ Error processing record: {e}")

        except Exception as e:
            print(f"❌ JSON Parsing Error for {current_day.strftime('%Y-%m-%d')}: {e}")

        # Move to the next day
        current_day += timedelta(days=1)

    # Convert to DataFrame
    df_unintended_deviation = pd.DataFrame(all_rows, columns=['Timestamp', 'Unintended_Import (MW)', 'Unintended_Export (MW)'])

    if df_unintended_deviation.empty:
        print("⚠️ No data fetched for the requested period.")
        return df_unintended_deviation

    # Ensure Timestamp is properly formatted in CET
    df_unintended_deviation['Timestamp'] = pd.to_datetime(df_unintended_deviation['Timestamp']).dt.tz_localize(None)

    # Add "Quarter" and "Lookup" fields for easy reference
    df_unintended_deviation['Quarter'] = ((df_unintended_deviation.index % 96) + 1)
    df_unintended_deviation['Lookup'] = df_unintended_deviation["Timestamp"].dt.strftime('%d.%m.%Y') + df_unintended_deviation["Quarter"].astype(str)

    # Save to Excel
    df_unintended_deviation.to_excel("./data_fetching/Entsoe/Unintended_Deviation.xlsx", index=False)

    print("✅ Successfully saved unintended deviation data.")

    return df_unintended_deviation

#==================================================================================== Wind Prouction Forecast===================================================================================
solcast_api_key = os.getenv("solcast_api_key")

def fetching_Cogealac_data():
	lat = 44.561156
	lon = 28.562586
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=wind_direction_100m,wind_direction_10m,wind_speed_100m,wind_speed_10m&period=PT60M&format=csv&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_raw.csv", 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")
	# Adjusting the values to EET time
	data = pd.read_csv("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_raw.csv")
	# Assuming 'period_end' is the column to keep fixed and all other columns are to be shifted
	columns_to_shift = data.columns.difference(['period_end'])

	# Shift the data columns by 2 intervals
	data_shifted = data[columns_to_shift].shift(2)

	# Combine the fixed 'period_end' with the shifted data columns
	data_adjusted = pd.concat([data[['period_end']], data_shifted], axis=1)

	# Optionally, handle the NaN values in the first two rows after shifting
	data_adjusted.fillna(0, inplace=True)  # Or use another method as appropriate

	# Save the adjusted DataFrame
	data_adjusted.to_csv("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_raw.csv", index=False)

def fetching_Cogealac_data_15min():
	lat = 44.561156
	lon = 28.562586
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,wind_direction_100m,wind_direction_10m,wind_speed_100m,wind_speed_10m&period=PT15M&format=csv&time_zone=3&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_Cogealac_raw_15min.csv", 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")
	# Saving the Cogealac wind dataset
	data = pd.read_csv("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_Cogealac_raw_15min.csv")
	data.to_csv("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_Cogealac_raw_15min.xlsx", index=False)

def fetching_Focsani_data_15min():
	lat = 45.696475
	lon = 27.184043
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,wind_direction_100m,wind_direction_10m,wind_speed_100m,wind_speed_10m&period=PT15M&format=csv&time_zone=3&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_Focsani_raw_15min.csv", 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")
	# Adjusting the values to EET time
	data = pd.read_csv("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_Focsani_raw_15min.csv")
	data.to_csv("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_Focsani_raw_15min.xlsx", index=False)

def predicting_wind_production():
	# Creating the forecast_dataset df
	data_Cogealac = pd.read_csv("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_raw.csv")
	forecast_dataset = pd.read_excel("./Market Fundamentals/Wind_Production_Forecast/Input_Wind_dataset.xlsx")
	# Convert 'period_end' in santimbru to datetime
	data['period_end'] = pd.to_datetime(data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	dates = data['period_end'].dt.strftime('%Y-%m-%d')
	# Write the dates to the Input file
	forecast_dataset['Data'] = dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	forecast_dataset['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	intervals = data["period_end"].dt.hour + 1
	forecast_dataset["Interval"] = intervals
	# Replace NaNs in the 'Interval' column with 0
	forecast_dataset['Interval'].fillna(1, inplace=True)
	# Completing the wind_direction_100m column
	forecast_dataset["wind_direction_100m"] = data["wind_direction_100m"].values
	# Completing the wind_direction_10m column
	forecast_dataset["wind_direction_10m"] = data["wind_direction_10m"].values
	# Completing the wind_speed_100m column
	forecast_dataset["wind_speed_100m"] = data["wind_speed_100m"].values
	# Completing the wind_speed_10m column
	forecast_dataset["wind_speed_10m"] = data["wind_speed_10m"].values


	xgb_loaded = joblib.load("./Market Fundamentals/Wind_Production_Forecast/rs_xgb_wind_production_0924.pkl")

	forecast_dataset["Month"] = pd.to_datetime(forecast_dataset.Data).dt.month
	dataset = forecast_dataset.copy()
	forecast_dataset = forecast_dataset.drop("Data", axis=1)

	preds = xgb_loaded.predict(forecast_dataset.values)
	
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Market Fundamentals/Wind_Production_Forecast/Wind_Forecast_Production.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
		worksheet.write(row, col + 2, value, decimal_format)
		row += 1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
		worksheet.write(row, col + 0, Data, date_format)
		worksheet.write(row, col + 1, Interval)
		row += 1

	workbook.close()
	file_path = "./Market Fundamentals/Wind_Production_Forecast/Wind_Forecast_Production.xlsx"
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset

def predicting_wind_production_15min():
	# Creating the forecast_dataset df
	data_Cogealac = pd.read_csv("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_Cogealac_raw_15min.csv")
	data_Focsani = pd.read_csv("./Market Fundamentals/Wind_Production_Forecast/Wind_dataset_Focsani_raw_15min.csv")
	forecast_dataset = pd.read_excel("./Market Fundamentals/Wind_Production_Forecast/Input_Wind_dataset_15min.xlsx")
	# Convert the 'period_end' column to datetime, handling errors
	data_Cogealac['period_end'] = pd.to_datetime(data_Cogealac['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')
	data_Focsani['period_end'] = pd.to_datetime(data_Focsani['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')
	# Shift the 'period_end' column by 2 hours
	data_Cogealac['period_end'] = data_Cogealac['period_end'] + pd.Timedelta(hours=3)
	data_Focsani['period_end'] = data_Focsani['period_end'] + pd.Timedelta(hours=3)

	forecast_dataset['Data'] = data_Cogealac.period_end.dt.strftime('%Y-%m-%d %H:%M:%S').values
	# Creating the Interval column
	forecast_dataset['Interval'] = data_Cogealac.period_end.dt.hour * 4 + data_Cogealac.period_end.dt.minute // 15 + 1
	# Replace NaNs in the 'Interval' column with 0
	forecast_dataset['Interval'].fillna(9, inplace=True)
	# Replace NaNs in the 'Date' column with the previous valid observation
	forecast_dataset['Data'].fillna(method='ffill', inplace=True)
	# Filling the air_temp Cogealac column
	forecast_dataset["air_temp_Cogealac"] = data_Cogealac["air_temp"].values
	# Completing the wind_direction_100m Cogealac column
	forecast_dataset["wind_direction_100m_Cogealac"] = data_Cogealac["wind_direction_100m"].values
	# Completing the wind_direction_10m Cogealac column
	forecast_dataset["wind_direction_10m_Cogealac"] = data_Cogealac["wind_direction_10m"].values
	# Completing the wind_speed_100m Cogealac column
	forecast_dataset["wind_speed_100m_Cogealac"] = data_Cogealac["wind_speed_100m"].values
	# Completing the wind_speed_10m Cogealac column
	forecast_dataset["wind_speed_10m_Cogealac"] = data_Cogealac["wind_speed_10m"].values
	# Completing the wind_gust Cogealac column
	# forecast_dataset["wind_gust_Cogealac"] = data_Cogealac["wind_gust"].values
	# Filling the air_temp Focsani column
	forecast_dataset["air_temp_Focsani"] = data_Focsani["air_temp"].values
	# Completing the wind_direction_100m Focsani column
	forecast_dataset["wind_direction_100m_Focsani"] = data_Focsani["wind_direction_100m"].values
	# Completing the wind_direction_10m Focsani column
	forecast_dataset["wind_direction_10m_Focsani"] = data_Focsani["wind_direction_10m"].values
	# Completing the wind_speed_100m Focsani column
	forecast_dataset["wind_speed_100m_Focsani"] = data_Focsani["wind_speed_100m"].values
	# Completing the wind_speed_10m Focsani column
	forecast_dataset["wind_speed_10m_Focsani"] = data_Focsani["wind_speed_10m"].values
	# Completing the wind_gust Focsani column
	# forecast_dataset["wind_gust_Focsani"] = data_Focsani["wind_gust"].values

	xgb_loaded = joblib.load("./Market Fundamentals/Wind_Production_Forecast/rs_xgb_wind_production_0325.pkl")

	forecast_dataset["Month"] = pd.to_datetime(forecast_dataset.Data).dt.month
	dataset = forecast_dataset.copy()
	forecast_dataset = forecast_dataset.drop("Data", axis=1)
	forecast_dataset = forecast_dataset[["Interval", "air_temp_Cogealac", "wind_direction_100m_Cogealac", "wind_direction_10m_Cogealac", "wind_speed_100m_Cogealac", "wind_speed_10m_Cogealac", "air_temp_Focsani", "wind_direction_100m_Focsani", "wind_direction_10m_Focsani", "wind_speed_100m_Focsani", "wind_speed_10m_Focsani", "Month"]]
	preds = xgb_loaded.predict(forecast_dataset.values)

	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]

	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Market Fundamentals/Wind_Production_Forecast/Wind_Forecast_Production_15min.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
		worksheet.write(row, col + 2, value, decimal_format)
		row += 1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
		worksheet.write(row, col + 0, Data, date_format)
		worksheet.write(row, col + 1, Interval)
		row += 1

	workbook.close()
	file_path = "./Market Fundamentals/Wind_Production_Forecast/Wind_Forecast_Production_15min.xlsx"
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset
#=================================================================================PZU Price Forecast=========================================================================================
# Defining the function for forecasting
def predicting_price_forecast():
	# Creating the forecast_dataset df
	forecast_dataset = pd.read_excel("./Market Fundamentals/Spot_Price_Forecast/Input_Price_dataset.xlsx")
	xgb_loaded = joblib.load("./Market Fundamentals/Spot_Price_Forecast/rs_xgb_price_forecast_0924.pkl")

	forecast_dataset["Month"] = pd.to_datetime(forecast_dataset.Data).dt.month
	dataset = forecast_dataset.copy()
	forecast_dataset = forecast_dataset.drop("Data", axis=1)

	preds = xgb_loaded.predict(forecast_dataset.values)
	
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Market Fundamentals/Spot_Price_Forecast/Results_Price_Forecast.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
		worksheet.write(row, col + 2, value, decimal_format)
		row += 1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
		worksheet.write(row, col + 0, Data, date_format)
		worksheet.write(row, col + 1, Interval)
		row += 1

	workbook.close()
	# Adding the Lookup column
	file_path = "./Market Fundamentals/Spot_Price_Forecast/Results_Price_Forecast.xlsx"
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset
#====================================================================================Rendering into App================================================================================

if "df_wind_15min" and "df_solar_15min" not in st.session_state:
	st.session_state["df_wind_15min"] = []
	st.session_state["df_solar_15min"] = []

def render_fundamentals_page():
	
	# Page tittle
	st.header("Market Fundamentals")
	st.divider()
	# Get Volue data
	st.subheader("Volue Data", divider = "rainbow")
	if st.button("Fetch data"):
		issue_date_str = get_issue_date()[0]
		df_wind_15min = fetch_volue_wind_data(issue_date_str)
		df_solar_15min = fetch_volue_solar_data(issue_date_str)
		df_hydro_15min = fetch_volue_hydro_data(issue_date_str)
		# df_temps_15min = fetch_volue_temperature_data(issue_date_str)
		df_price_15min = fetch_volue_price_data(issue_date_str)
		st.write("Volue Price dataframe")
		st.dataframe(df_price_15min)
		# Creatiung the Volue dataframe
		# 1. Creating the first layer of the big dataframe containing the Wind and Solar Energy
		df_wind_15min.reset_index(inplace=True)
		df_wind_15min.columns = ['Timestamp', "Wind Power"]

		# Convert 'Timestamp' to datetime format
		df_wind_15min['Timestamp'] = pd.to_datetime(df_wind_15min['Timestamp'])

		# Create 'Date' and 'Interval' columns from 'Timestamp'
		df_wind_15min['Date'] = df_wind_15min['Timestamp'].dt.strftime('%d.%m.%Y')
		df_wind_15min['Interval'] = ((df_wind_15min['Timestamp'].dt.hour * 60 + df_wind_15min['Timestamp'].dt.minute) // 15 + 1)
		df_wind_15min.drop(columns=["Timestamp"], inplace=True)
		st.dataframe(df_wind_15min)

		df_solar_15min.reset_index(inplace=True)
		df_solar_15min.columns = ['Timestamp', "Solar Power"]

		# Convert 'Timestamp' to datetime format
		df_solar_15min['Timestamp'] = pd.to_datetime(df_solar_15min['Timestamp'])

		# Create 'Date' and 'Interval' columns from 'Timestamp'
		df_solar_15min['Date'] = df_solar_15min['Timestamp'].dt.strftime('%d.%m.%Y')
		df_solar_15min['Interval'] = ((df_solar_15min['Timestamp'].dt.hour * 60 + df_solar_15min['Timestamp'].dt.minute) // 15 + 1)
		df_solar_15min.drop(columns=["Timestamp"], inplace=True)
		st.dataframe(df_solar_15min)

		df_wind_15min['Date'] = pd.to_datetime(df_wind_15min['Date'], format="%d.%m.%Y")
		df_solar_15min['Date'] = pd.to_datetime(df_solar_15min['Date'], format="%d.%m.%Y")

		df_wind_15min['Interval'] = df_wind_15min['Interval'].astype('int64')
		df_solar_15min['Interval'] = df_solar_15min['Interval'].astype('int64')

		# Now proceed with the merge as previously described
		df_final = pd.merge(df_wind_15min, df_solar_15min, on=['Date', 'Interval'], how='left')

		df_hydro_15min['Date'] = pd.to_datetime(df_hydro_15min['Date'], format="%d.%m.%Y")
		df_hydro_15min['Interval'] = df_hydro_15min['Interval'].astype('int64')
		# Adding the Hydro Power to the Volue dataframe
		df_final_2 = pd.merge(df_final, df_hydro_15min, on=['Date', 'Interval'], how='left')

		# Adding the Temperatures to the Volue dataframe
		# df_temps_15min = fetch_volue_temperature_data(issue_date_str)
		# df_temps_15min.reset_index(inplace=True)
		# df_temps_15min.columns = ['Timestamp', "Temperature"]

		# # Convert 'Timestamp' to datetime format
		# df_temps_15min['Timestamp'] = pd.to_datetime(df_temps_15min['Timestamp'])

		# # Create 'Date' and 'Interval' columns from 'Timestamp'
		# df_temps_15min['Date'] = df_temps_15min['Timestamp'].dt.strftime('%d.%m.%Y')
		# df_temps_15min['Interval'] = ((df_temps_15min['Timestamp'].dt.hour * 60 + df_temps_15min['Timestamp'].dt.minute) // 15 + 1)
		# df_temps_15min.drop(columns=["Timestamp"], inplace=True)
		# st.dataframe(df_temps_15min)
		# # Creating the Volue dataframe containing Wind, Solar, Hydro and Temperatures
		# df_final_3 = pd.merge(df_final_2, df_temps_15min, on=['Date', 'Interval'], how='left')
		# st.dataframe(df_final_3)

		# Adding the Price to the Volue dataframe
		df_price_15min['Date'] = pd.to_datetime(df_price_15min['Date'], format="%d.%m.%Y")
		df_price_15min['Interval'] = df_price_15min['Interval'].astype('int64')
		df_final_3 = pd.merge(df_final_2, df_price_15min, on=['Date', 'Interval'], how='left')
		st.dataframe(df_final_3)

		# Creating a download button for the Volue data
		file_path = './Market Fundamentals/Volue_data.xlsx'
		with open(file_path, "rb") as f:
			excel_data = f.read()

		# Create a download link
		b64 = base64.b64encode(excel_data).decode()
		button_html = f"""
			 <a download="Volue_data.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Export Volue Data</button>
			 </a> 
			 """
		st.markdown(button_html, unsafe_allow_html=True)

	# Fething the Transelectrica data
	st.subheader("Transelectrica Data", divider="violet")
	st.text("Upload the weekly consumption and production from Transelectrica. Don't change anything.")
	# Uploading the consumption ad production files from Transelectrica
	uploaded_files = st.file_uploader("Upload Weekly Consumption and Production files", type=['csv', 'xlsx', "xls"], accept_multiple_files=True)
	if st.button("Process Transelectrica"):
		if uploaded_files:
			for uploaded_file in uploaded_files:
				if uploaded_file.type == 'application/vnd.ms-excel' or uploaded_file.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
					# 1. Consumption
					if uploaded_file.name == "weekly consumtion 2023.xlsx":
						print(uploaded_file.name)
						# Now you can call this function with the path to your consumption data file
						processed_data = process_file_consumption_transelectrica(uploaded_file)

						# First, ensure we're working with string types to avoid errors on non-string types
						processed_data['Date'] = processed_data['Date'].astype(str)

						# Then, filter out any rows where 'Date' does not contain a valid date string (e.g., the string "Data")
						processed_data = processed_data[processed_data['Date'].str.contains(r'\d{4}-\d{2}-\d{2}')]

						# Now, safely convert 'Date' column to datetime format
						processed_data['Date'] = pd.to_datetime(processed_data['Date'])

						# Filter for current montha and year
						# Get the current date
						current_date = datetime.now()

						# Extract the current year and month
						current_year = current_date.year
						current_month = current_date.month

						# Filter the data for the current year and month
						filtered_data = processed_data[(processed_data['Date'].dt.year == current_year) & (processed_data['Date'].dt.month == current_month)]

						filtered_data.dropna(inplace=True)
						# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
						filtered_data['Lookup'] = filtered_data["Date"].dt.strftime('%d.%m.%Y') + filtered_data["Interval"].astype(str)
						filtered_data.to_excel('./Market Fundamentals/Transelectrica_data/Cons_Prod_final/Weekly_Consumption_2025.xlsx', index=False)
					elif uploaded_file.name == "weekly_production_2023.xls":
						production = pd.read_excel("./Market Fundamentals/Transelectrica_data/weekly_production_2023.xls")
						production.to_excel('./Market Fundamentals/Transelectrica_data/weekly_production_2023.xlsx', engine='openpyxl', index=False)
						# Example usage
						file_path = './Market Fundamentals/Transelectrica_data/weekly_production_2023.xlsx'  # Change this to your actual file path
						processed_data = process_file_production_transelectrica(file_path)

						# First, ensure we're working with string types to avoid errors on non-string types
						processed_data['Date'] = processed_data['Date'].astype(str)

						# Then, filter out any rows where 'Date' does not contain a valid date string (e.g., the string "Data")
						processed_data = processed_data[processed_data['Date'].str.contains(r'\d{4}-\d{2}-\d{2}')]

						# Now, safely convert 'Date' column to datetime format
						processed_data['Date'] = pd.to_datetime(processed_data['Date'])

						# Filter for current montha and year
						# Get the current date
						current_date = datetime.now()

						# Extract the current year and month
						current_year = current_date.year
						current_month = current_date.month

						# Filter for April 2024
						filtered_data = processed_data[(processed_data['Date'].dt.year == current_year) & (processed_data['Date'].dt.month == current_month)]
						filtered_data.dropna(inplace=True)
						# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
						filtered_data['Lookup'] = filtered_data["Date"].dt.strftime('%d.%m.%Y') + filtered_data["Interval"].astype(str)
						filtered_data.to_excel('./Market Fundamentals/Transelectrica_data/Cons_Prod_final/Weekly_Production_2025.xlsx', index=False)

		# Downloading the Files
		folder_path = './Market Fundamentals/Transelectrica_data/Cons_Prod_final/'
		zip_name = 'Transelectrica_Input.zip'
		zip_files(folder_path, zip_name)
		file_path = './Market Fundamentals/Transelectrica_data/Cons_Prod_final/Transelectrica_Input.zip'

		with open(file_path, "rb") as f:
			zip_data = f.read()

		# Create a download link
		b64 = base64.b64encode(zip_data).decode()
		button_html = f"""
			 <a download="Transelectrica_Input.zip" href="data:application/zip;base64,{b64}" download>
			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Input Files</button>
			 </a> 
			 """
		st.markdown(button_html, unsafe_allow_html=True)

	# Fetching the Entsoe data
	st.subheader("Entsoe Data", divider="violet")
	issue_date = get_issue_date()[1]
	start_date = st.date_input("Select Start Date", value=pd.to_datetime(issue_date + timedelta(days=1)))
	end_date = st.date_input("Select End Date", value=pd.to_datetime(issue_date + timedelta(days=2)))
	# start_date = pd.to_datetime('2024-04-14')
	# end_date = pd.to_datetime('2024-04-15')
	# Convert the input dates to Timestamps with time set to '0000' hours
	start = pd.Timestamp(f"{start_date.strftime('%Y%m%d')}0000", tz='Europe/Bucharest')
	end = pd.Timestamp(f"{end_date.strftime('%Y%m%d')}0000", tz='Europe/Bucharest')
	# Switching to CET time for the Load Forecast
	start_cet = pd.Timestamp(f"{start_date.strftime('%Y%m%d')}0000", tz='Europe/Budapest')
	end_cet = pd.Timestamp(f"{end_date.strftime('%Y%m%d')}0000", tz='Europe/Budapest')

	# start = pd.Timestamp('202404110000', tz='Europe/Bucharest')  # Adjust the start date as needed
	# end = pd.Timestamp('202404120000', tz='Europe/Bucharest')    # Adjust the end date as needed
	if st.button("Entsoe"):
		# Fetch the DataFrames
		df_imbalance_prices = imbalance_prices(start, end)
		df_imbalance_volumes = imbalance_volumes(start, end)
		# Merge the DataFrames on their indices
		df_imbalance = pd.merge(df_imbalance_prices, df_imbalance_volumes, left_index=True, right_index=True, how='inner')
		# Assuming df_imbalance is your merged DataFrame
		df_imbalance = df_imbalance.rename(columns={'Long': 'Excedent Price',
													'Short': 'Deficit Price'})
		st.dataframe(df_imbalance)
		# Fetching the Activated mFRR capacities
		if (end_date-start_date).days > 1:
			st.text("Fetching data for more than one day")
			df_activated_energy = query_daily_mFRR(start_date, end_date)
		else:
			data = activated_mFRR_energy(start, end)
			df_activated_energy = creating_mFRR_dfs(data)
		# st.dataframe(df_activated_energy)

		# Fethcing the Actual Load
		df_actual_load = actual_load(start, end)
		# st.dataframe(df_actual_load)

		# Fethcing the Forecast Load
		df_load_forecast = load_forecast(start, end)
		df_load_forecast_CET = load_forecast_CET(start_cet, end_cet)
		st.dataframe(df_load_forecast_CET)

		# Fetching the Generation Forecast
		df_generation_forecast = generation_forecast(start, end)
		# st.dataframe(df_generation_forecast)

		# Fetching the Generation per Source
		df_actual_generation_source = actual_generation_source(start, end)
		st.dataframe(df_actual_generation_source)
		st.write(df_actual_generation_source.columns)
		
		# Creating the dataframe with all the Entsoe data
		# 1. Concatenating the df_activated_energy and df_imbalance
		df_imbalance.reset_index(inplace=True)
		df_imbalance.columns = ['Timestamp', 'Excedent Price', 'Deficit Price', 'Imbalance Volume']

		# Convert 'Timestamp' to datetime format
		df_imbalance['Timestamp'] = pd.to_datetime(df_imbalance['Timestamp'])

		# Create 'Date' and 'Interval' columns from 'Timestamp'
		df_imbalance['Date'] = df_imbalance['Timestamp'].dt.strftime('%d.%m.%Y')
		df_imbalance['Interval'] = ((df_imbalance['Timestamp'].dt.hour * 60 + df_imbalance['Timestamp'].dt.minute) // 15 + 1)

		# Now proceed with the merge as previously described
		df_final = pd.merge(df_activated_energy, df_imbalance, on=['Date', 'Interval'], how='left')

		# Optionally remove the 'Timestamp' column if it's no longer needed
		df_final.drop(columns=['Timestamp'], inplace=True)

		# 2. Adding the Actual Load
		df_actual_load.reset_index(inplace=True)
		df_actual_load.rename(columns={'index': 'Timestamp'}, inplace=True)
		# Convert 'Timestamp' to datetime format
		df_actual_load['Timestamp'] = pd.to_datetime(df_actual_load['Timestamp'])

		# Create 'Date' and 'Interval' columns from 'Timestamp'
		df_actual_load['Date'] = df_actual_load['Timestamp'].dt.strftime('%d.%m.%Y')
		df_actual_load['Interval'] = ((df_actual_load['Timestamp'].dt.hour * 60 + df_actual_load['Timestamp'].dt.minute) // 15 + 1)

		# Now proceed with the merge as previously described
		df_final_2 = pd.merge(df_final, df_actual_load, on=['Date', 'Interval'], how='left')

		# Optionally remove the 'Timestamp' column if it's no longer needed
		df_final_2.drop(columns=['Timestamp'], inplace=True)

		# 3. Adding the Generation Forecast
		# Convert the Series to DataFrame
		df_generation_forecast = df_generation_forecast.to_frame('Generation Forecast')
		# If your Series had the timestamp as the index (which is common in time series data), you can reset the index to make it a column:
		df_generation_forecast.reset_index(inplace=True)
		df_generation_forecast.rename(columns={'index': 'Timestamp'}, inplace=True)

		# Create 'Date' and 'Interval' columns from 'Timestamp'
		df_generation_forecast['Date'] = df_generation_forecast['Timestamp'].dt.strftime('%d.%m.%Y')
		df_generation_forecast['Interval'] = ((df_generation_forecast['Timestamp'].dt.hour * 60 + df_generation_forecast['Timestamp'].dt.minute) // 15 + 1)

		# Now proceed with the merge as previously described
		df_final_3 = pd.merge(df_final_2, df_generation_forecast, on=['Date', 'Interval'], how='left')

		# Optionally remove the 'Timestamp' column if it's no longer needed
		df_final_3.drop(columns=['Timestamp'], inplace=True)

		# 4. Adding the Generation Forecast per Source
		df_actual_generation_source.reset_index(inplace=True)
		df_actual_generation_source.columns = ['Timestamp', "Biomass", "Fossil Brown coal/Lignite", "Fossil Gas", "Fossil Hard coal", 
		"Hydro Run-of-river and poundage", "Hydro Water Reservoir", "Nuclear", "Solar", "Wind Onshore"]

		# Convert 'Timestamp' to datetime format
		df_actual_generation_source['Timestamp'] = pd.to_datetime(df_actual_generation_source['Timestamp'])

		# Create 'Date' and 'Interval' columns from 'Timestamp'
		df_actual_generation_source['Date'] = df_actual_generation_source['Timestamp'].dt.strftime('%d.%m.%Y')
		df_actual_generation_source['Interval'] = ((df_actual_generation_source['Timestamp'].dt.hour * 60 + df_actual_generation_source['Timestamp'].dt.minute) // 15 + 1)

		# Now proceed with the merge as previously described
		df_final_4 = pd.merge(df_final_3, df_actual_generation_source, on=['Date', 'Interval'], how='left')

		# Optionally remove the 'Timestamp' column if it's no longer needed
		df_final_4.drop(columns=['Timestamp'], inplace=True)

		# Display the merged DataFrame
		st.dataframe(df_final_4)

		# Exporting the final Entsoe dataframe to Excel
		df_final_4.to_excel("./Market Fundamentals/Entsoe_data/Entsoe_data.xlsx")

		# Formatting the Date column
		# Assuming df_final_4 is your DataFrame
		# Create an Excel writer object
		with pd.ExcelWriter("./Market Fundamentals/Entsoe_data/Entsoe_data.xlsx", 
							engine='xlsxwriter', 
							datetime_format='dd.mm.yyyy') as writer:
			df_final_4.to_excel(writer, index=False)

			# Access the xlsxwriter workbook and worksheet objects from the dataframe
			workbook  = writer.book
			worksheet = writer.sheets['Sheet1']

			# Get the number of rows in the DataFrame
			num_rows = len(df_final_4.index)

			# Define a format object for Excel to use
			date_format = workbook.add_format({'num_format': 'dd.mm.yyyy'})

			# Apply the date format to the column with dates (assuming it's the first column)
			worksheet.set_column(0, 0, None, date_format)  # Column 'A:A' if your dates are in the first column

	if st.button("Load Forecast CET"):
		load_forecast_CET(start_cet, end_cet)

		# Creating a download button for the Volue data
		file_path = './Market Fundamentals/Entsoe_data/Load_Forecast.xlsx'
		with open(file_path, "rb") as f:
			excel_data = f.read()

		# Create a download link
		b64 = base64.b64encode(excel_data).decode()
		button_html = f"""
			 <a download="Load_Forecast.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Load</button>
			 </a> 
			 """
		st.markdown(button_html, unsafe_allow_html=True)

	st.header("Wind Production Forecast", divider = "green")
	# if st.button("Forecast Wind Production"):
	# 	fetching_Cogealac_data()
	# 	st.dataframe(predicting_wind_production())
	# 	with open("./Market Fundamentals/Wind_Production_Forecast/Wind_Forecast_Production.xlsx", "rb") as f:
	# 		excel_data = f.read()

	# 		# Create a download link
	# 		b64 = base64.b64encode(excel_data).decode()
	# 		button_html = f"""
	# 			 <a download="Wind_Production.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
	# 			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Wind Production Forecast</button>
	# 			 </a> 
	# 			 """
	# 		st.markdown(button_html, unsafe_allow_html=True)
	if st.button("Forecast Wind Production Quarterly"):
		fetching_Cogealac_data_15min()
		fetching_Focsani_data_15min()
		st.dataframe(predicting_wind_production_15min())
		with open("./Market Fundamentals/Wind_Production_Forecast/Wind_Forecast_Production_15min.xlsx", "rb") as f:
			excel_data = f.read()

			# Create a download link
			b64 = base64.b64encode(excel_data).decode()
			button_html = f"""
				 <a download="Wind_Production_15min.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
				 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Wind Production Forecast</button>
				 </a> 
				 """
			st.markdown(button_html, unsafe_allow_html=True)

	st.header("Spot Price Forecast", divider = "gray")
	if st.button("Forecast Price"):
		st.dataframe(predicting_price_forecast())
		with open("./Market Fundamentals/Spot_Price_Forecast/Results_Price_Forecast.xlsx", "rb") as f:
			excel_data = f.read()

			# Create a download link
			b64 = base64.b64encode(excel_data).decode()
			button_html = f"""
				 <a download="Price_Forecast.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
				 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Price Forecast</button>
				 </a> 
				 """
			st.markdown(button_html, unsafe_allow_html=True)

	# Streamlit UI for date input
	st.header("Fetch Fundamentals from ENTSO-E")

	# Default values for the date range
	today = date.today()
	default_start_date = today
	default_end_date = today

	# Date range input widgets
	start_date = st.date_input("Start Date", value=default_start_date)
	end_date = st.date_input("End Date", value=default_end_date)

	# Validate the date range
	if start_date > end_date:
		st.error("Error: Start Date cannot be after End Date.")
	else:
		st.success(f"Selected date range: {start_date} to {end_date}")

	if st.button("Fetch Data"):
		# Fetching the Imbalance Volumes and Prices
		fetching_imbalance_volumes(start_date, end_date)
		fetch_imbalance_prices(start_date, end_date)
		df_volumes = process_imbalance_volumes(start_date, end_date)
		df_prices = process_imbalance_prices(start_date, end_date)
		df_imbalance = create_combined_imbalance_dataframe(df_volumes, df_prices)

		# Fetching the Wind Historical Production
		df_wind_notified = fetch_process_wind_notified(start_date, end_date)
		df_wind_actual = fetch_process_wind_actual_production(start_date, end_date)
		df_wind_data = combine_wind_production_data(df_wind_notified, df_wind_actual)
		
		# Fetching the Solar Historical Prodution
		df_solar_notified = fetch_process_solar_notified(start_date, end_date)
		df_solar_actual = fetch_process_solar_actual_production(start_date, end_date)
		df_solar_data = combine_solar_production_data(df_solar_notified, df_solar_actual)

		# Hhydro Historical data
		df_reservoir = fetch_process_hydro_water_reservoir_actual_production(start_date, end_date)
		df_river = fetch_process_hydro_river_actual_production(start_date, end_date)
		df_hydro = align_and_combine_hydro_data(df_reservoir, df_river)

		# Fetching the Consumption data
		df_consumption_forecast = fetch_consumption_forecast(start_date, end_date)
		df_consumption_actual = fetch_actual_consumption(start_date, end_date)

		# Fetching the Balancing Energy Activation
		st.dataframe(fetch_intraday_balancing_activations(start_date, end_date))

		# Fetching the Unintended Deviation Data
		st.dataframe(fetch_unintended_deviation_data(start_date, end_date))
		
		# Fetching the IGCC flows
		st.dataframe(fetch_igcc_netting_flows_range(start_date, end_date))










